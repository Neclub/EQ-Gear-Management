"""Export team equipped gear to Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from inventory_parser.team_report import CharacterGear, TeamGearReport
from inventory_parser.achievement_report import AchievementReport
from inventory_parser.achievement_parser import format_expansion_label
from inventory_parser.spell_report import SpellRuneReport
from inventory_parser.spell_runes import load_rune_config
from inventory_parser.useful_spells import (
    RACCOO_USEFUL_SPELLS_CREDIT_TEXT,
    RACCOO_USEFUL_SPELLS_URL,
    MissingUsefulSpellsReport,
)
from inventory_parser.excel_theme import (
    EVOLVER_FILL,
    FILL_HEADER,
    FILL_ITEM_EMPTY,
    FILL_LABEL,
    FILL_SHEET,
    FILL_SPELL_COUNT,
    FILL_SPELL_DETAIL,
    FILL_SPELL_DETAIL_ALT,
    FONT_BLOCK_HEADER,
    FONT_BLOCK_SUB,
    FONT_BODY,
    FONT_COUNT,
    FONT_HEADER,
    FONT_LEGEND,
    FONT_LEGEND_TITLE,
    FONT_LINK,
    FONT_SECTION,
    SHEET_BACKGROUND_COLS,
    SHEET_BACKGROUND_ROWS,
    apply_workbook_dark_mode,
    spell_block_header_fill,
    spell_tier_fill,
    tier_code_fill,
    tier_bucket_legend_rows,
)
from inventory_parser.evolver import EVOLVER_GAP_LABEL
from inventory_parser.items import EquippedItem
from inventory_parser.slots import SlotFilter, slot_visibility, slots_for_export
from inventory_parser.excel_theme import GEAR_SET_FILLS
from inventory_parser.gear_tiers import SOR_GAP_LEGEND_ROWS, UNKNOWN_TIER_LABEL
from inventory_parser.sor_tier import equipped_tier_label
from inventory_parser.rune_inventory import RuneInventoryReport
from inventory_parser.unmade_gear import UnmadeGearEntry, build_unmade_gear_report

_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
_SPELL_BORDER_COLOR = "404048"
_SPELL_BORDER = Border(
    left=Side(style="thin", color=_SPELL_BORDER_COLOR),
    right=Side(style="thin", color=_SPELL_BORDER_COLOR),
    top=Side(style="thin", color=_SPELL_BORDER_COLOR),
    bottom=Side(style="thin", color=_SPELL_BORDER_COLOR),
)
_COL_SLOT_WIDTH = 13.0
_COL_VISIBILITY_WIDTH = 14.0
_COL_ITEM_WIDTH = 48.0
_COL_SPELL_CHAR = 16.0
_COL_SPELL_BLOCK = 12.0
_COL_SPELL_LEVEL = 8.0
_COL_SPELL_RUNE = 12.0
_COL_SPELL_EXPANSION = 28.0
_COL_SPELL_NAME = _COL_ITEM_WIDTH
_COL_MATRIX_CHAR_MIN = 8.0
_COL_MATRIX_CHAR_MAX = 40.0
_FIRST_CHAR_COL = 3
_LEGEND_TITLE_ROW = 25
_LEGEND_FIRST_ROW = 26
_LEGEND_LAST_ROW = 30
_LEGEND_LABEL_MERGE_COLS = 5  # B through F
_SOR_LEGEND_TITLE_ROW = 25
_SOR_LEGEND_FIRST_ROW = 26

GEAR_T_LEVEL_SHEET_NAME = "Gear T-Level"
UNMADE_GEAR_SHEET_NAME = "Unmade Gear"
MISSING_COLLECTIONS_SHEET_NAME = "Missing Collections"
ACHIEVEMENT_SUMMARY_SHEET_NAME = "Achievement Summary"
QUESTS_SHEET_NAME = "Quests"
RAID_ACHIEVEMENTS_SHEET_NAME = "Raid Achievements"
RUNE_INVENTORY_SHEET_NAME = "Rune Inventory"
MISSING_SPELLS_SHEET_NAME = "Missing Spells"
MISSING_USEFUL_SPELLS_SHEET_NAME = "Missing Useful Spells"


def write_team_workbook(
    report: TeamGearReport,
    output_path: Path,
    *,
    slot_filter: SlotFilter = "all",
    spell_report: SpellRuneReport | None = None,
    missing_useful_report: MissingUsefulSpellsReport | None = None,
    rune_inventory_report: RuneInventoryReport | None = None,
    achievement_report: AchievementReport | None = None,
    unmade_entries: list[UnmadeGearEntry] | None = None,
    slot2=None,
    type5=None,
    type18=None,
    raid_bis=None,
) -> Path:
    """Write team gear workbook with item sheet and SOR gap tracking sheet."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    base_slots = slots_for_export(slot_filter)
    slots_tab1 = list(base_slots)
    slots_tab2 = _slots_for_sor_sheet(report, base_slots)

    wb = Workbook()

    ws_gear = wb.active
    ws_gear.title = "Team gear"
    _write_team_gear_sheet(ws_gear, report, slots_tab1)

    ws_sor = wb.create_sheet(GEAR_T_LEVEL_SHEET_NAME)
    _write_sor_gaps_sheet(ws_sor, report, slots_tab2)

    if spell_report is not None:
        ws_runes = wb.create_sheet("Missing Runes")
        _write_missing_runes_sheet(ws_runes, report, spell_report)
        ws_list = wb.create_sheet(MISSING_SPELLS_SHEET_NAME)
        _write_spell_list_sheet(ws_list, report, spell_report)

    if missing_useful_report is not None and missing_useful_report.entries:
        ws_useful = wb.create_sheet(MISSING_USEFUL_SPELLS_SHEET_NAME)
        _write_missing_useful_spells_sheet(ws_useful, missing_useful_report)

    if rune_inventory_report is not None:
        ws_inventory = wb.create_sheet(RUNE_INVENTORY_SHEET_NAME)
        _write_rune_inventory_sheet(ws_inventory, rune_inventory_report)

    entries = (
        unmade_entries
        if unmade_entries is not None
        else build_unmade_gear_report(report)
    )
    if entries:
        ws_unmade = wb.create_sheet(UNMADE_GEAR_SHEET_NAME)
        _write_unmade_gear_sheet(ws_unmade, entries)

    if achievement_report is not None:
        if achievement_report.missing_collections:
            ws_missing = wb.create_sheet(MISSING_COLLECTIONS_SHEET_NAME)
            _write_missing_collections_sheet(ws_missing, achievement_report)
        if achievement_report.quests:
            ws_quests = wb.create_sheet(QUESTS_SHEET_NAME)
            _write_quests_sheet(ws_quests, achievement_report)
        if achievement_report.summaries:
            ws_summary = wb.create_sheet(ACHIEVEMENT_SUMMARY_SHEET_NAME)
            _write_achievement_summary_sheet(ws_summary, achievement_report)
        if achievement_report.raid_achievements:
            ws_raids = wb.create_sheet(RAID_ACHIEVEMENTS_SHEET_NAME)
            _write_raid_achievements_sheet(ws_raids, achievement_report)

    if slot2 is not None:
        from inventory_parser.slot2_augs.excel import append_slot2_sheets

        append_slot2_sheets(wb, slot2)

    if type5 is not None:
        from inventory_parser.type5_augs.excel import append_type5_sheet

        append_type5_sheet(wb, type5)

    if type18 is not None:
        from inventory_parser.type18_augs.excel import append_type18_sheet

        append_type18_sheet(wb, type18)

    if raid_bis is not None:
        from inventory_parser.raid_bis.excel import append_raid_bis_sheet

        append_raid_bis_sheet(wb, raid_bis)

    apply_workbook_dark_mode(wb)
    return _save_with_fallback(wb, output_path)


def _slots_for_sor_sheet(report: TeamGearReport, base_slots: tuple[str, ...]) -> tuple[str, ...]:
    """Drop Secondary unless at least one character had it equipped on the gear sheet."""
    slots = list(base_slots)
    if "Secondary" in slots and not any("Secondary" in c.slots for c in report.characters):
        slots.remove("Secondary")
    return tuple(slots)


def _write_team_gear_sheet(ws: Worksheet, report: TeamGearReport, slots: tuple[str, ...]) -> None:
    ws.sheet_properties.tabColor = "000000"
    num_cols = _write_sheet_header(ws, report)
    for row_idx, slot in enumerate(slots, start=2):
        _write_slot_label_cells(ws, row_idx, slot)
        for col in range(_FIRST_CHAR_COL, num_cols + 1):
            char_row = report.characters[col - _FIRST_CHAR_COL]
            item = char_row.slots.get(slot)
            if item is None:
                ws.cell(row_idx, col).fill = FILL_ITEM_EMPTY
                continue
            _write_item_cell(ws, row_idx, col, item)
    if slots:
        _apply_auto_filter(ws, num_cols, len(slots))
    _write_gear_legend_on_sheet(ws)
    content_last_col = max(num_cols, _LEGEND_LABEL_MERGE_COLS)
    _fill_legend_row_overflow(ws, _LEGEND_TITLE_ROW, _LEGEND_LAST_ROW, content_last_col)
    data_last_row = 1 + len(slots)
    gap_start = data_last_row + 1
    _fill_sheet_padding(
        ws,
        content_last_row=_LEGEND_LAST_ROW,
        content_last_col=content_last_col,
        gap_rows=range(gap_start, _LEGEND_TITLE_ROW) if gap_start < _LEGEND_TITLE_ROW else None,
    )


def _write_sor_gaps_sheet(ws: Worksheet, report: TeamGearReport, slots: tuple[str, ...]) -> None:
    ws.sheet_properties.tabColor = "542A35"
    num_cols = _write_sheet_header(ws, report)
    for row_idx, slot in enumerate(slots, start=2):
        _write_slot_label_cells(ws, row_idx, slot)
        for col in range(_FIRST_CHAR_COL, num_cols + 1):
            char_row = report.characters[col - _FIRST_CHAR_COL]
            item = char_row.slots.get(slot)
            label = equipped_tier_label(item)
            if label:
                cell = ws.cell(row_idx, col, label)
                cell.font = FONT_BODY
                cell.alignment = _ALIGN
                cell.fill = _fill_for_tier_code(label)
            else:
                ws.cell(row_idx, col).fill = FILL_ITEM_EMPTY
    if slots:
        _apply_auto_filter(ws, num_cols, len(slots))
    _write_sor_legend_on_sheet(ws)
    sor_legend_last = _SOR_LEGEND_FIRST_ROW + len(SOR_GAP_LEGEND_ROWS) - 1
    content_last_col = max(num_cols, _LEGEND_LABEL_MERGE_COLS)
    _fill_legend_row_overflow(ws, _SOR_LEGEND_TITLE_ROW, sor_legend_last, content_last_col)
    data_last_row = 1 + len(slots)
    gap_start = data_last_row + 1
    _fill_sheet_padding(
        ws,
        content_last_row=sor_legend_last,
        content_last_col=content_last_col,
        pad_rows=max(SHEET_BACKGROUND_ROWS, sor_legend_last),
        gap_rows=range(gap_start, _SOR_LEGEND_TITLE_ROW) if gap_start < _SOR_LEGEND_TITLE_ROW else None,
    )


def _write_sheet_header(ws: Worksheet, report: TeamGearReport) -> int:
    headers = ("Slot", "Visibility", *(r.display_name for r in report.characters))
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(1, col, title)
        cell.font = FONT_HEADER
        cell.alignment = _ALIGN_HEADER
        cell.fill = FILL_HEADER
    ws.column_dimensions["A"].width = _COL_SLOT_WIDTH
    ws.column_dimensions["B"].width = _COL_VISIBILITY_WIDTH
    for col in range(_FIRST_CHAR_COL, _FIRST_CHAR_COL + len(report.characters)):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = _COL_ITEM_WIDTH
    return len(headers)


def _write_slot_label_cells(ws: Worksheet, row_idx: int, slot: str) -> None:
    slot_cell = ws.cell(row_idx, 1, slot)
    slot_cell.font = FONT_BODY
    slot_cell.alignment = _ALIGN
    slot_cell.fill = FILL_LABEL
    vis_cell = ws.cell(row_idx, 2, slot_visibility(slot))
    vis_cell.font = FONT_BODY
    vis_cell.alignment = _ALIGN
    vis_cell.fill = FILL_LABEL


def _fill_for_tier_code(code: str):
    if code == EVOLVER_GAP_LABEL:
        return EVOLVER_FILL
    if code == UNKNOWN_TIER_LABEL or code:
        return tier_code_fill(code)
    return FILL_ITEM_EMPTY


def _fill_legend_row_overflow(
    ws: Worksheet,
    first_row: int,
    last_row: int,
    content_last_col: int,
) -> None:
    """Black-fill legend rows beyond the merged label columns up to content width."""
    if content_last_col <= _LEGEND_LABEL_MERGE_COLS:
        return
    for row in range(first_row, last_row + 1):
        for col in range(_LEGEND_LABEL_MERGE_COLS + 1, content_last_col + 1):
            ws.cell(row, col).fill = FILL_SHEET


def _fill_sheet_padding(
    ws: Worksheet,
    *,
    content_last_row: int,
    content_last_col: int,
    pad_rows: int = SHEET_BACKGROUND_ROWS,
    pad_cols: int = SHEET_BACKGROUND_COLS,
    gap_rows: range | None = None,
) -> None:
    """Paint unused chrome black without re-styling already filled content cells."""
    for row in range(1, min(content_last_row, pad_rows) + 1):
        for col in range(1, content_last_col + 1):
            cell = ws.cell(row, col)
            if cell.fill.fill_type is None:
                cell.fill = FILL_SHEET
    for row in range(1, pad_rows + 1):
        for col in range(content_last_col + 1, pad_cols + 1):
            ws.cell(row, col).fill = FILL_SHEET
    for row in range(content_last_row + 1, pad_rows + 1):
        for col in range(1, content_last_col + 1):
            ws.cell(row, col).fill = FILL_SHEET
    if gap_rows is not None:
        for row in gap_rows:
            for col in range(1, content_last_col + 1):
                ws.cell(row, col).fill = FILL_SHEET


def _spell_list_last_row(entry_count: int) -> int:
    """Last row on Missing Spells: banner, header, entries, spacer, footer note."""
    return 6 + entry_count


def _missing_useful_last_row(entry_count: int) -> int:
    """Last row on Missing Useful Spells: banner, header, entries, spacer, notes."""
    return 7 + entry_count


def _fill_for_equipped_item(item: EquippedItem):
    label = equipped_tier_label(item)
    if label:
        return _fill_for_tier_code(label)
    return FILL_ITEM_EMPTY


def _write_item_cell(ws: Worksheet, row_idx: int, col: int, item: EquippedItem) -> None:
    cell = ws.cell(row_idx, col, item.name)
    cell.alignment = _ALIGN
    cell.fill = _fill_for_equipped_item(item)
    url = item.eqresource_url
    cell.font = FONT_LINK if url else FONT_BODY
    if url:
        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url)


def _write_gear_legend_on_sheet(ws: Worksheet) -> None:
    title = ws.cell(_LEGEND_TITLE_ROW, 1, "Gear tier colors")
    title.font = FONT_LEGEND_TITLE
    title.fill = FILL_HEADER
    title.alignment = _ALIGN
    ws.merge_cells(
        start_row=_LEGEND_TITLE_ROW,
        start_column=1,
        end_row=_LEGEND_TITLE_ROW,
        end_column=_LEGEND_LABEL_MERGE_COLS,
    )
    for col in range(1, _LEGEND_LABEL_MERGE_COLS + 1):
        ws.cell(_LEGEND_TITLE_ROW, col).fill = FILL_HEADER

    legend_rows = tier_bucket_legend_rows()
    expected = _LEGEND_LAST_ROW - _LEGEND_FIRST_ROW + 1
    if len(legend_rows) != expected:
        raise ValueError(
            f"Expected {expected} legend rows for A{_LEGEND_FIRST_ROW}:A{_LEGEND_LAST_ROW}, "
            f"got {len(legend_rows)}"
        )

    for offset, (fill, label_text) in enumerate(legend_rows):
        row = _LEGEND_FIRST_ROW + offset
        ws.cell(row, 1).fill = fill
        label = ws.cell(row, 2, label_text)
        label.font = FONT_LEGEND
        label.fill = FILL_LABEL
        label.alignment = _ALIGN
        ws.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=_LEGEND_LABEL_MERGE_COLS,
        )
        for col in range(2, _LEGEND_LABEL_MERGE_COLS + 1):
            ws.cell(row, col).fill = FILL_LABEL


def _write_sor_legend_on_sheet(ws: Worksheet) -> None:
    title = ws.cell(_SOR_LEGEND_TITLE_ROW, 1, "Gear tier codes")
    title.font = FONT_LEGEND_TITLE
    title.fill = FILL_HEADER
    title.alignment = _ALIGN
    ws.merge_cells(
        start_row=_SOR_LEGEND_TITLE_ROW,
        start_column=1,
        end_row=_SOR_LEGEND_TITLE_ROW,
        end_column=_LEGEND_LABEL_MERGE_COLS,
    )
    for col in range(1, _LEGEND_LABEL_MERGE_COLS + 1):
        ws.cell(_SOR_LEGEND_TITLE_ROW, col).fill = FILL_HEADER

    for offset, (code, desc) in enumerate(SOR_GAP_LEGEND_ROWS):
        row = _SOR_LEGEND_FIRST_ROW + offset
        code_cell = ws.cell(row, 1, code or "")
        code_cell.font = FONT_BODY
        code_cell.alignment = _ALIGN
        code_cell.fill = _fill_for_tier_code(code) if code else FILL_ITEM_EMPTY
        label = ws.cell(row, 2, desc)
        label.font = FONT_LEGEND
        label.fill = FILL_LABEL
        label.alignment = _ALIGN
        ws.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=_LEGEND_LABEL_MERGE_COLS,
        )
        for col in range(2, _LEGEND_LABEL_MERGE_COLS + 1):
            ws.cell(row, col).fill = FILL_LABEL


def _apply_auto_filter(ws: Worksheet, num_cols: int, num_slot_rows: int) -> None:
    last_col = get_column_letter(num_cols)
    last_row = 1 + num_slot_rows
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"


def _apply_spell_cell_border(ws: Worksheet, row: int, col: int) -> None:
    ws.cell(row, col).border = _SPELL_BORDER


def _apply_spell_table_borders(
    ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int
) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            _apply_spell_cell_border(ws, r, c)


def _merge_spell_row(
    ws: Worksheet, row: int, min_col: int, max_col: int, fill: PatternFill | None = None
) -> None:
    if max_col <= min_col:
        return
    ws.merge_cells(
        start_row=row,
        start_column=min_col,
        end_row=row,
        end_column=max_col,
    )
    if fill is None:
        return
    for col in range(min_col, max_col + 1):
        ws.cell(row, col).fill = fill


def _write_rune_tier_matrix_section(
    ws: Worksheet,
    *,
    start_row: int,
    title: str,
    subtitle: str,
    header_key: str,
    tiers: tuple[str, ...],
    characters: list[CharacterGear],
    count_for,
) -> int:
    """Tier × character count matrix; returns next free row."""
    summary_cols = max(2, 1 + len(characters))
    last_col = summary_cols
    block_fill = spell_block_header_fill(header_key)

    title_cell = ws.cell(start_row, 1, title)
    title_cell.font = FONT_BLOCK_HEADER
    title_cell.fill = block_fill
    title_cell.alignment = _ALIGN_WRAP
    sub = ws.cell(start_row + 1, 1, subtitle)
    sub.font = FONT_BLOCK_SUB
    sub.fill = block_fill
    sub.alignment = _ALIGN_WRAP
    _merge_spell_row(ws, start_row, 1, last_col, block_fill)
    _merge_spell_row(ws, start_row + 1, 1, last_col, block_fill)
    ws.row_dimensions[start_row].height = 20
    ws.row_dimensions[start_row + 1].height = 18

    header_row = start_row + 2
    label_cell = ws.cell(header_row, 1, "Rune tier")
    label_cell.font = FONT_HEADER
    label_cell.fill = FILL_HEADER
    label_cell.alignment = _ALIGN_HEADER
    for col, char in enumerate(characters, start=2):
        c = ws.cell(header_row, col, char.display_name)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = _ALIGN_HEADER
    _apply_spell_table_borders(ws, header_row, header_row, 1, last_col)

    row = header_row + 1
    for tier in tiers:
        tier_fill = spell_tier_fill(tier)
        tier_cell = ws.cell(row, 1, tier)
        tier_cell.font = FONT_BODY
        tier_cell.fill = tier_fill
        tier_cell.alignment = _ALIGN
        for col, char in enumerate(characters, start=2):
            count = count_for(char, tier)
            cell = ws.cell(row, col, count if count else "")
            cell.alignment = _ALIGN_CENTER
            if count:
                cell.font = FONT_COUNT
                cell.fill = FILL_SPELL_COUNT
            else:
                cell.font = FONT_BODY
                cell.fill = FILL_ITEM_EMPTY
        _apply_spell_table_borders(ws, row, row, 1, last_col)
        row += 1

    return row + 1


def _write_spell_summary_section(
    ws: Worksheet,
    *,
    start_row: int,
    group,
    tiers: tuple[str, ...],
    characters: list[CharacterGear],
    spell_report: SpellRuneReport,
) -> int:
    """Rune count matrix for one spell expansion; returns next free row."""
    return _write_rune_tier_matrix_section(
        ws,
        start_row=start_row,
        title=group.label,
        subtitle=group.turn_in_theme,
        header_key=group.key,
        tiers=tiers,
        characters=characters,
        count_for=lambda char, tier: spell_report.counts_by_persona.get(char.persona_key, {})
        .get(group.label, {})
        .get(tier, 0),
    )


def _spell_characters(team: TeamGearReport, spell_report: SpellRuneReport) -> list[CharacterGear]:
    personas = team.spell_characters if team.spell_characters else team.characters
    if spell_report.persona_keys:
        by_persona = {c.persona_key: c for c in personas}
        return [by_persona[pk] for pk in spell_report.persona_keys if pk in by_persona]
    return list(personas)


def _write_spell_sheet_banner(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str,
    merge_cols: int,
) -> int:
    """Title + subtitle rows; returns the next free row."""
    row = 1
    title_cell = ws.cell(row, 1, title)
    title_cell.font = FONT_SECTION
    title_cell.fill = FILL_HEADER
    title_cell.alignment = _ALIGN
    sub_cell = ws.cell(row + 1, 1, subtitle)
    sub_cell.font = FONT_LEGEND
    sub_cell.fill = FILL_HEADER
    sub_cell.alignment = _ALIGN
    _merge_spell_row(ws, row, 1, merge_cols, FILL_HEADER)
    _merge_spell_row(ws, row + 1, 1, merge_cols, FILL_HEADER)
    ws.row_dimensions[row].height = 24
    return row + 3


def _matrix_character_col_width(label: str) -> float:
    """Excel column width (character units) to fit a matrix header label."""
    return max(_COL_MATRIX_CHAR_MIN, min(_COL_MATRIX_CHAR_MAX, len(label) + 2.0))


def _apply_rune_matrix_column_widths(
    ws: Worksheet,
    characters: list[CharacterGear],
    *,
    tier_col_width: float = 14.0,
) -> None:
    ws.column_dimensions["A"].width = tier_col_width
    for col, char in enumerate(characters, start=2):
        ws.column_dimensions[get_column_letter(col)].width = _matrix_character_col_width(
            char.display_name
        )


def _write_rune_inventory_sheet(
    ws: Worksheet,
    report: RuneInventoryReport,
) -> None:
    ws.sheet_properties.tabColor = "2A6B5C"
    characters = list(report.characters)
    summary_cols = max(2, 1 + len(characters))

    row = _write_spell_sheet_banner(
        ws,
        title="Rune Inventory",
        subtitle="Raid spell runes on hand · General, Bank, and Shared Bank",
        merge_cols=summary_cols,
    )

    for family in report.families:
        row = _write_rune_tier_matrix_section(
            ws,
            start_row=row,
            title=family.label,
            subtitle=f"{family.label} · {family.item_pattern}",
            header_key=family.id,
            tiers=family.tiers,
            characters=characters,
            count_for=lambda char, tier: family.counts.get(char.persona_key, {}).get(tier, 0),
        )

    note = ws.cell(
        row,
        1,
        "Counts include stacked items in General bags, personal Bank, and Shared Bank. "
        "Inert and Covariant Engrams are excluded.",
    )
    note.font = FONT_LEGEND
    note.fill = FILL_SHEET
    note.alignment = _ALIGN_WRAP
    _merge_spell_row(ws, row, 1, summary_cols)

    _apply_rune_matrix_column_widths(ws, characters)
    _fill_sheet_padding(
        ws,
        content_last_row=row,
        content_last_col=summary_cols,
        pad_rows=max(SHEET_BACKGROUND_ROWS, row),
    )


def _write_missing_runes_sheet(
    ws: Worksheet,
    team: TeamGearReport,
    spell_report: SpellRuneReport,
) -> None:
    ws.sheet_properties.tabColor = "5C4688"
    config = load_rune_config()
    tiers = config.tiers
    characters = _spell_characters(team, spell_report)
    summary_cols = max(2, 1 + len(characters))

    row = _write_spell_sheet_banner(
        ws,
        title="Missing Runes",
        subtitle="Rank III runes still needed per character · counts by spell expansion",
        merge_cols=summary_cols,
    )

    for group in spell_report.expansion_groups:
        row = _write_spell_summary_section(
            ws,
            start_row=row,
            group=group,
            tiers=tiers,
            characters=characters,
            spell_report=spell_report,
        )

    note = ws.cell(
        row,
        1,
        "Counts are missing Rk. III spells only. Rune tier = turn-in type (Minor → Glowing). "
        "See Missing Spells tab for individual spells.",
    )
    note.font = FONT_LEGEND
    note.fill = FILL_SHEET
    note.alignment = _ALIGN_WRAP
    _merge_spell_row(ws, row, 1, summary_cols)

    ws.column_dimensions["A"].width = 14.0
    for col in range(2, 2 + len(characters)):
        ws.column_dimensions[get_column_letter(col)].width = 11.0
    _fill_sheet_padding(
        ws,
        content_last_row=row,
        content_last_col=summary_cols,
        pad_rows=max(SHEET_BACKGROUND_ROWS, row),
    )


def _write_spell_list_sheet(
    ws: Worksheet,
    team: TeamGearReport,
    spell_report: SpellRuneReport,
) -> None:
    ws.sheet_properties.tabColor = "3D4A6E"
    last_row = _spell_list_last_row(len(spell_report.entries))

    row = _write_spell_sheet_banner(
        ws,
        title=MISSING_SPELLS_SHEET_NAME,
        subtitle="Every missing Rank III spell from /outputfile missingspells",
        merge_cols=5,
    )

    detail_headers = ("Character", "Level", "Rune", "Expansion", "Spell")
    for col, header in enumerate(detail_headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = _ALIGN_HEADER
    detail_header_row = row
    _apply_spell_table_borders(ws, detail_header_row, detail_header_row, 1, 5)
    row += 1

    prev_display_name: str | None = None
    stripe = 0
    for entry in spell_report.entries:
        if entry.display_name != prev_display_name:
            stripe = 1 - stripe
            prev_display_name = entry.display_name
        row_fill = FILL_SPELL_DETAIL_ALT if stripe else FILL_SPELL_DETAIL

        char_cell = ws.cell(row, 1, entry.display_name)
        char_cell.font = FONT_BODY
        char_cell.alignment = _ALIGN
        char_cell.fill = row_fill

        level_cell = ws.cell(row, 2, entry.level)
        level_cell.font = FONT_BODY
        level_cell.alignment = _ALIGN_CENTER
        level_cell.fill = row_fill

        rune_cell = ws.cell(row, 3, entry.rune_tier)
        rune_cell.font = FONT_BODY
        rune_cell.alignment = _ALIGN_CENTER
        rune_cell.fill = spell_tier_fill(entry.rune_tier)

        expansion_cell = ws.cell(row, 4, entry.expansion or None)
        expansion_cell.font = FONT_BODY
        expansion_cell.alignment = _ALIGN
        expansion_cell.fill = row_fill

        spell_cell = ws.cell(row, 5, entry.spell_name)
        spell_cell.font = FONT_BODY
        spell_cell.alignment = _ALIGN
        spell_cell.fill = row_fill

        _apply_spell_table_borders(ws, row, row, 1, 5)
        row += 1

    if spell_report.entries:
        last_detail_row = row - 1
        ws.auto_filter.ref = f"A{detail_header_row}:E{last_detail_row}"
        ws.freeze_panes = ws.cell(detail_header_row + 1, 1).coordinate

    row += 1
    note = ws.cell(
        row,
        1,
        "Only Rank III lines are listed. Rune tier matches spell level within each band. "
        "Add blocks in spell_rune_bands.json for future level caps.",
    )
    note.font = FONT_LEGEND
    note.fill = FILL_SHEET
    note.alignment = _ALIGN_WRAP
    _merge_spell_row(ws, row, 1, 5, FILL_SHEET)

    ws.column_dimensions["A"].width = _COL_SPELL_CHAR
    ws.column_dimensions["B"].width = _COL_SPELL_LEVEL
    ws.column_dimensions["C"].width = _COL_SPELL_RUNE
    ws.column_dimensions["D"].width = _COL_SPELL_EXPANSION
    ws.column_dimensions["E"].width = _COL_SPELL_NAME
    _fill_sheet_padding(
        ws,
        content_last_row=max(row, last_row),
        content_last_col=5,
        pad_rows=max(row, last_row),
    )


def _write_missing_useful_spells_sheet(
    ws: Worksheet,
    report: MissingUsefulSpellsReport,
) -> None:
    ws.sheet_properties.tabColor = "5A4A6E"
    last_row = _missing_useful_last_row(len(report.entries))
    col_count = 6

    row = _write_spell_sheet_banner(
        ws,
        title=MISSING_USEFUL_SPELLS_SHEET_NAME,
        subtitle="Curated useful spells that still appear in /outputfile missingspells (all levels)",
        merge_cols=col_count,
    )

    detail_headers = (
        "Character",
        "Level",
        "Expansion",
        "Spell",
        "Highest RK",
        "Comments",
    )
    for col, header in enumerate(detail_headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = _ALIGN_HEADER
    detail_header_row = row
    _apply_spell_table_borders(ws, detail_header_row, detail_header_row, 1, col_count)
    row += 1

    prev_display_name: str | None = None
    stripe = 0
    for entry in report.entries:
        if entry.display_name != prev_display_name:
            stripe = 1 - stripe
            prev_display_name = entry.display_name
        row_fill = FILL_SPELL_DETAIL_ALT if stripe else FILL_SPELL_DETAIL

        values = (
            entry.display_name,
            entry.level,
            entry.expansion or None,
            entry.spell_name,
            entry.highest_rk or None,
            entry.comments or None,
        )
        alignments = (
            _ALIGN,
            _ALIGN_CENTER,
            _ALIGN,
            _ALIGN,
            _ALIGN_CENTER,
            _ALIGN_WRAP,
        )
        for col, (value, alignment) in enumerate(zip(values, alignments), start=1):
            cell = ws.cell(row, col, value)
            cell.font = FONT_BODY
            cell.alignment = alignment
            cell.fill = row_fill

        _apply_spell_table_borders(ws, row, row, 1, col_count)
        row += 1

    if report.entries:
        last_detail_row = row - 1
        ws.auto_filter.ref = f"A{detail_header_row}:F{last_detail_row}"
        ws.freeze_panes = ws.cell(detail_header_row + 1, 1).coordinate

    row += 1
    credit = ws.cell(row, 1, RACCOO_USEFUL_SPELLS_CREDIT_TEXT)
    credit.font = FONT_LINK
    credit.fill = FILL_SHEET
    credit.alignment = _ALIGN_WRAP
    credit.hyperlink = Hyperlink(ref=credit.coordinate, target=RACCOO_USEFUL_SPELLS_URL)
    _merge_spell_row(ws, row, 1, col_count, FILL_SHEET)

    row += 1
    note = ws.cell(
        row,
        1,
        "Intersection of the curated useful-spell list with each character's MissingSpells file. "
        "Filter the Character column to focus on one persona.",
    )
    note.font = FONT_LEGEND
    note.fill = FILL_SHEET
    note.alignment = _ALIGN_WRAP
    _merge_spell_row(ws, row, 1, col_count, FILL_SHEET)

    ws.column_dimensions["A"].width = _COL_SPELL_CHAR
    ws.column_dimensions["B"].width = _COL_SPELL_LEVEL
    ws.column_dimensions["C"].width = _COL_SPELL_EXPANSION
    ws.column_dimensions["D"].width = _COL_SPELL_NAME
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 36
    _fill_sheet_padding(
        ws,
        content_last_row=max(row, last_row),
        content_last_col=col_count,
        pad_rows=max(row, last_row),
    )


def _write_unmade_gear_sheet(ws: Worksheet, entries: list[UnmadeGearEntry]) -> None:
    ws.sheet_properties.tabColor = "4A5A38"
    last_row = 1 + len(entries)

    headers = (
        "Character",
        "Item",
        "Count",
        "Bag Location",
        "Expansion",
        "Material",
        "Target Slot",
        "Equipped Tier",
        "Notes",
    )
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = _ALIGN_HEADER

    for row_idx, entry in enumerate(entries, start=2):
        values = (
            entry.display_name,
            entry.item_name,
            entry.count,
            entry.bag_location,
            entry.expansion,
            entry.material,
            entry.target_slot or "",
            entry.equipped_tier,
            entry.notes,
        )
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.font = FONT_BODY
            cell.fill = FILL_SPELL_DETAIL if row_idx % 2 == 0 else FILL_SPELL_DETAIL_ALT
            cell.alignment = _ALIGN if col != 3 else _ALIGN_CENTER

        item_cell = ws.cell(row_idx, 2)
        url = (
            f"https://items.eqresource.com/items.php?id={entry.item_id}"
            if entry.item_id > 0
            else None
        )
        if url:
            item_cell.font = FONT_LINK
            item_cell.hyperlink = Hyperlink(ref=item_cell.coordinate, target=url)

        tier_cell = ws.cell(row_idx, 8)
        if entry.equipped_tier:
            tier_cell.fill = _fill_for_tier_code(entry.equipped_tier)

    if entries:
        ws.auto_filter.ref = f"A1:I{1 + len(entries)}"
        ws.freeze_panes = ws.cell(2, 1).coordinate

    ws.column_dimensions["A"].width = _COL_SPELL_CHAR
    ws.column_dimensions["B"].width = _COL_ITEM_WIDTH
    ws.column_dimensions["C"].width = 8.0
    ws.column_dimensions["D"].width = 18.0
    ws.column_dimensions["E"].width = 10.0
    ws.column_dimensions["F"].width = 10.0
    ws.column_dimensions["G"].width = 14.0
    ws.column_dimensions["H"].width = 14.0
    ws.column_dimensions["I"].width = 24.0
    _fill_sheet_padding(
        ws,
        content_last_row=last_row,
        content_last_col=len(headers),
        pad_rows=max(last_row, SHEET_BACKGROUND_ROWS),
    )


def _write_missing_collections_sheet(ws: Worksheet, report: AchievementReport) -> None:
    ws.sheet_properties.tabColor = "5A4A38"
    entries = report.missing_collections
    last_row = 1 + len(entries)

    headers = (
        "Character",
        "Expansion",
        "Zone",
        "Collection",
        "Missing Item",
        "Progress",
        "Char Has",
        "Total",
    )
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = _ALIGN_HEADER

    for row_idx, entry in enumerate(entries, start=2):
        values = (
            entry.character,
            format_expansion_label(entry.expansion),
            entry.zone,
            entry.collection,
            entry.missing_item,
            entry.progress,
            entry.char_has,
            entry.total,
        )
        row_fill = FILL_SPELL_DETAIL if row_idx % 2 == 0 else FILL_SPELL_DETAIL_ALT
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.font = FONT_BODY
            cell.fill = row_fill
            cell.alignment = _ALIGN if col not in (6, 8) else _ALIGN_CENTER

    if entries:
        ws.auto_filter.ref = f"A1:H{1 + len(entries)}"
        ws.freeze_panes = ws.cell(2, 1).coordinate

    ws.column_dimensions["A"].width = _COL_SPELL_CHAR
    ws.column_dimensions["B"].width = 18.0
    ws.column_dimensions["C"].width = 22.0
    ws.column_dimensions["D"].width = 28.0
    ws.column_dimensions["E"].width = _COL_ITEM_WIDTH
    ws.column_dimensions["F"].width = 10.0
    ws.column_dimensions["G"].width = 8.0
    ws.column_dimensions["H"].width = 8.0
    _fill_sheet_padding(
        ws,
        content_last_row=last_row,
        content_last_col=len(headers),
        pad_rows=max(last_row, SHEET_BACKGROUND_ROWS),
    )


def _write_achievement_summary_sheet(ws: Worksheet, report: AchievementReport) -> None:
    ws.sheet_properties.tabColor = "384A5A"
    entries = report.summaries
    last_row = 1 + len(entries)

    headers = (
        "Character",
        "Section",
        "Completed",
        "Incomplete",
        "Total",
        "Completion %",
    )
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = _ALIGN_HEADER

    for row_idx, entry in enumerate(entries, start=2):
        values = (
            entry.character,
            format_expansion_label(entry.section),
            entry.completed,
            entry.incomplete,
            entry.total,
            entry.completion_pct,
        )
        row_fill = FILL_SPELL_DETAIL if row_idx % 2 == 0 else FILL_SPELL_DETAIL_ALT
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.font = FONT_BODY
            cell.fill = row_fill
            cell.alignment = _ALIGN if col <= 2 else _ALIGN_CENTER

    if entries:
        ws.auto_filter.ref = f"A1:F{1 + len(entries)}"
        ws.freeze_panes = ws.cell(2, 1).coordinate

    ws.column_dimensions["A"].width = _COL_SPELL_CHAR
    ws.column_dimensions["B"].width = 22.0
    ws.column_dimensions["C"].width = 12.0
    ws.column_dimensions["D"].width = 12.0
    ws.column_dimensions["E"].width = 10.0
    ws.column_dimensions["F"].width = 14.0
    _fill_sheet_padding(
        ws,
        content_last_row=last_row,
        content_last_col=len(headers),
        pad_rows=max(last_row, SHEET_BACKGROUND_ROWS),
    )


def _write_quests_sheet(ws: Worksheet, report: AchievementReport) -> None:
    ws.sheet_properties.tabColor = "5A384A"
    entries = report.quests
    last_row = 1 + len(entries)

    headers = (
        "Character",
        "Expansion",
        "Zone",
        "Type",
        "Quest",
        "Status",
    )
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = _ALIGN_HEADER

    for row_idx, entry in enumerate(entries, start=2):
        values = (
            entry.character,
            format_expansion_label(entry.expansion),
            entry.zone,
            entry.quest_type,
            entry.quest,
            entry.status,
        )
        row_fill = FILL_SPELL_DETAIL if row_idx % 2 == 0 else FILL_SPELL_DETAIL_ALT
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.font = FONT_BODY
            cell.fill = row_fill
            cell.alignment = _ALIGN

    if entries:
        ws.auto_filter.ref = f"A1:F{1 + len(entries)}"
        ws.freeze_panes = ws.cell(2, 1).coordinate

    ws.column_dimensions["A"].width = _COL_SPELL_CHAR
    ws.column_dimensions["B"].width = 22.0
    ws.column_dimensions["C"].width = _COL_ITEM_WIDTH
    ws.column_dimensions["D"].width = 12.0
    ws.column_dimensions["E"].width = _COL_ITEM_WIDTH
    ws.column_dimensions["F"].width = 12.0
    _fill_sheet_padding(
        ws,
        content_last_row=last_row,
        content_last_col=len(headers),
        pad_rows=max(last_row, SHEET_BACKGROUND_ROWS),
    )


def _write_raid_achievements_sheet(ws: Worksheet, report: AchievementReport) -> None:
    ws.sheet_properties.tabColor = "5A384A"
    entries = report.raid_achievements
    last_row = 1 + len(entries)

    headers = (
        "Character",
        "Expansion",
        "Raid",
        "Event",
        "Objective",
        "Status",
    )
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = _ALIGN_HEADER

    for row_idx, entry in enumerate(entries, start=2):
        values = (
            entry.character,
            format_expansion_label(entry.expansion),
            entry.raid,
            entry.event,
            entry.objective,
            entry.status,
        )
        row_fill = FILL_SPELL_DETAIL if row_idx % 2 == 0 else FILL_SPELL_DETAIL_ALT
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.font = FONT_BODY
            cell.fill = row_fill
            cell.alignment = _ALIGN

    if entries:
        ws.auto_filter.ref = f"A1:F{1 + len(entries)}"
        ws.freeze_panes = ws.cell(2, 1).coordinate

    ws.column_dimensions["A"].width = _COL_SPELL_CHAR
    ws.column_dimensions["B"].width = 22.0
    ws.column_dimensions["C"].width = _COL_ITEM_WIDTH
    ws.column_dimensions["D"].width = 28.0
    ws.column_dimensions["E"].width = _COL_ITEM_WIDTH
    ws.column_dimensions["F"].width = 12.0
    _fill_sheet_padding(
        ws,
        content_last_row=last_row,
        content_last_col=len(headers),
        pad_rows=max(last_row, SHEET_BACKGROUND_ROWS),
    )


def _save_with_fallback(wb: Workbook, target: Path) -> Path:
    try:
        try:
            wb.save(target)
            return target
        except PermissionError:
            stem = target.stem
            suffix = target.suffix
            parent = target.parent
            for n in range(1, 100):
                alt = parent / f"{stem}_{n}{suffix}"
                try:
                    wb.save(alt)
                    return alt
                except PermissionError:
                    continue
            raise
    finally:
        wb.close()
