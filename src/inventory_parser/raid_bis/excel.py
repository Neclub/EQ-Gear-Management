"""Excel sheet for Raid BiS comparisons."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from inventory_parser.items import EQRESOURCE_ITEM_URL
from inventory_parser.raid_bis.build import RaidBisExport
from inventory_parser.raid_bis.compare import format_stat_deltas
from inventory_parser.slot2_augs.html import format_catalog_fetched_at

SHEET_NAME = "Raid BiS"

STATUS_FILLS = {
    "bis": PatternFill("solid", fgColor="166534"),
    "upgrade": PatternFill("solid", fgColor="854D0E"),
    "empty": PatternFill("solid", fgColor="7F1D1D"),
    "unknown": PatternFill("solid", fgColor="1E3A5F"),
    "weapon": PatternFill("solid", fgColor="374151"),
}
HEADER_FILL = PatternFill("solid", fgColor="1E2430")
HEADER_FONT = Font(color="EEF0F4", bold=True)
WHITE_FONT = Font(color="F9FAFB")
LINK_FONT = Font(color="8CB4FF", underline="single")
LINK_ON_FILL = Font(color="F9FAFB", underline="single")


def append_raid_bis_sheet(wb, bundle: RaidBisExport) -> None:
    ws = wb.create_sheet(SHEET_NAME)
    headers = [
        "Character",
        "Class",
        "Slot",
        "Status",
        "Current",
        "Best in slot",
        "Tier",
        "Vendor cost",
        "Vendor item",
        "Stat changes",
        "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for ch in bundle.characters:
        for slot in ch.slots:
            fill = STATUS_FILLS.get(slot.status)
            ws.cell(row, 1, ch.display_name)
            ws.cell(row, 2, ch.class_abbr or "")
            ws.cell(row, 3, slot.gear_slot)
            status_cell = ws.cell(row, 4, slot.status)
            if fill:
                status_cell.fill = fill
                status_cell.font = WHITE_FONT
            _item_cell(ws.cell(row, 5), slot.current_name, slot.current_id)
            _item_cell(ws.cell(row, 6), slot.recommended_name, slot.recommended_id)
            ws.cell(row, 7, slot.recommended_tier)
            if slot.vendor_cost is not None:
                ws.cell(row, 8, slot.vendor_cost)
            _item_cell(ws.cell(row, 9), slot.vendor_item_name, slot.vendor_item_id)
            ws.cell(row, 10, format_stat_deltas(slot.deltas, class_abbr=ch.class_abbr))
            ws.cell(row, 11, slot.note)
            row += 1

        if ch.total_deltas:
            ws.cell(row, 1, ch.display_name)
            ws.cell(row, 3, "TOTAL")
            ws.cell(row, 10, format_stat_deltas(ch.total_deltas, class_abbr=ch.class_abbr))
            row += 1

    widths = [22, 8, 12, 12, 42, 42, 8, 14, 36, 48, 36]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    fetched = format_catalog_fetched_at(bundle.catalog.fetched_at)
    cache_note = " (cache)" if bundle.catalog.from_cache else ""
    ws.cell(row + 1, 1, f"Catalog fetched {fetched}{cache_note}")


def _item_cell(cell, name: str | None, item_id: int | None) -> None:
    if not name:
        cell.value = ""
        return
    cell.value = name
    if item_id and item_id > 0:
        cell.hyperlink = EQRESOURCE_ITEM_URL.format(item_id=item_id)
        cell.font = LINK_FONT
