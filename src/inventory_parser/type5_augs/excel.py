"""Excel export for Type 5 aug display."""

from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from inventory_parser.excel_theme import (
    FILL_HEADER,
    FONT_HEADER,
    FONT_LINK,
    STATUS_FILL_EMPTY,
)
from inventory_parser.items import EQRESOURCE_ITEM_URL
from inventory_parser.team_report import format_character_display_name
from inventory_parser.type5_augs.build import Type5Export

EMPTY_FONT = Font(name="Calibri", size=11, color="FCA5A5")

SHEET_NAME = "Type 5 Augs"

_HEADERS = (
    "Character",
    "Slot",
    "Expansion",
    "Type 5 Aug",
    "HStr",
    "HSta",
    "HInt",
    "HWis",
    "HAgi",
    "HDex",
    "HCha",
)

_AUG_COL = 4
_EXPANSION_COL = 3
_AUG_WIDTH_MIN = 12.0
_AUG_WIDTH_MAX = 36.0
_EXPANSION_WIDTH_MIN = 12.0
_EXPANSION_WIDTH_MAX = 28.0
_FIXED_WIDTHS = {
    1: 22.0,
    2: 12.0,
    5: 8.0,
    6: 8.0,
    7: 8.0,
    8: 8.0,
    9: 8.0,
    10: 8.0,
    11: 8.0,
}


def _aug_column_width(names: list[str]) -> float:
    """Excel width for the Type 5 Aug column from header + longest name."""
    longest = max((len(n) for n in names), default=0)
    longest = max(longest, len("Type 5 Aug"))
    return max(_AUG_WIDTH_MIN, min(_AUG_WIDTH_MAX, longest + 2.0))


def _expansion_column_width(names: list[str]) -> float:
    longest = max((len(n) for n in names), default=0)
    longest = max(longest, len("Expansion"))
    return max(_EXPANSION_WIDTH_MIN, min(_EXPANSION_WIDTH_MAX, longest + 2.0))


def append_type5_sheet(wb, bundle: Type5Export) -> None:
    """Append a Type 5 Augs sheet onto an existing workbook."""
    ws = wb.create_sheet(SHEET_NAME)
    ws.sheet_properties.tabColor = "2D4A38"
    for col, h in enumerate(_HEADERS, start=1):
        cell = ws.cell(1, col, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    show_server = bundle.show_server_in_columns
    aug_names: list[str] = []
    expansion_names: list[str] = []
    row_idx = 2
    for i, ch in enumerate(bundle.characters):
        server = ch.server
        if bundle.roster and i < len(bundle.roster):
            server = bundle.roster[i].server or server
        label = (
            format_character_display_name(ch.character, ch.class_abbr)
        )
        if show_server and server:
            label = f"{label} ({server})"
        for slot in ch.slots:
            empty = slot.name is None or slot.item_id is None
            stats = slot.stats or {}
            ws.cell(row_idx, 1, label)
            ws.cell(row_idx, 2, slot.gear_slot)
            expansion = "" if empty else (slot.expansion or "")
            expansion_names.append(expansion)
            exp_cell = ws.cell(row_idx, 3, expansion)
            if empty:
                exp_cell.fill = STATUS_FILL_EMPTY
            elif slot.expansion_url:
                exp_cell.hyperlink = slot.expansion_url
                exp_cell.font = FONT_LINK
            aug_label = "Empty" if empty else (slot.name or "")
            aug_names.append(aug_label)
            aug_cell = ws.cell(row_idx, 4, aug_label)
            if empty:
                aug_cell.fill = STATUS_FILL_EMPTY
                aug_cell.font = EMPTY_FONT
            elif slot.item_id and slot.item_id > 0:
                aug_cell.hyperlink = EQRESOURCE_ITEM_URL.format(item_id=slot.item_id)
                aug_cell.font = FONT_LINK
            values = [
                int(stats.get("hstr", 0)) if not empty else "",
                int(stats.get("hsta", 0)) if not empty else "",
                int(stats.get("hint", 0)) if not empty else "",
                int(stats.get("hwis", 0)) if not empty else "",
                int(stats.get("hagi", 0)) if not empty else "",
                int(stats.get("hdex", 0)) if not empty else "",
                int(stats.get("hcha", 0)) if not empty else "",
            ]
            for col, val in enumerate(values, start=5):
                cell = ws.cell(row_idx, col, val)
                if empty:
                    cell.fill = STATUS_FILL_EMPTY
            row_idx += 1

    if row_idx == 2:
        ws.cell(2, 1, "No type 5 holes found on equipped gear.")

    # Catalog note
    note_row = row_idx + 1
    ws.cell(note_row, 1, "Type 5 list (EQ Resource):")
    link = ws.cell(note_row, 2, bundle.catalog_url)
    link.hyperlink = bundle.catalog_url
    link.font = FONT_LINK

    for col, w in _FIXED_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.column_dimensions[get_column_letter(_EXPANSION_COL)].width = (
        _expansion_column_width(expansion_names)
    )
    ws.column_dimensions[get_column_letter(_AUG_COL)].width = _aug_column_width(aug_names)
    ws.row_dimensions[1].height = 20
