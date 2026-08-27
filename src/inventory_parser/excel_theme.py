"""Dark workbook styling for team gear Excel export."""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Base surfaces (sheet background)
FILL_SHEET = PatternFill("solid", fgColor="000000")
SHEET_BACKGROUND_ROWS = 50
SHEET_BACKGROUND_COLS = 26  # column Z
FILL_HEADER = PatternFill("solid", fgColor="2D2D32")
FILL_LABEL = PatternFill("solid", fgColor="252528")
FILL_ITEM_EMPTY = PatternFill("solid", fgColor="222226")

# Text
COLOR_TEXT = "E4E6EB"
COLOR_TEXT_MUTED = "A8ADB8"
COLOR_LINK = "8CB4FF"

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT)
FONT_BODY = Font(name="Calibri", size=11, color=COLOR_TEXT)
FONT_LINK = Font(name="Calibri", size=11, color=COLOR_LINK, underline="single")
FONT_ON_STATUS = Font(name="Calibri", size=11, color=COLOR_TEXT)
FONT_LINK_ON_STATUS = Font(
    name="Calibri", size=11, color=COLOR_TEXT, underline="single"
)
FONT_MUTED = Font(name="Calibri", size=11, italic=True, color=COLOR_TEXT_MUTED)
FONT_LEGEND_TITLE = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT)
FONT_LEGEND = Font(name="Calibri", size=10, color=COLOR_TEXT_MUTED)

# Gear-set fills (newest first) — muted tones for light text
GEAR_SET_FILLS: dict[str, str] = {
    "fracture": "3A3350",
    "shattered_dominion": "453040",
    "rebellion": "542A35",
    "bound": "4D3828",
    "eternal_reverie": "283850",
    "heroic_reflections": "264845",
    "spectral_luclinite": "2C4530",
    "spectral_luminosity": "3A4228",
    "luclinite_coagulated": "3A3935",
    "evolver": "5C4688",
}

EVOLVER_FILL = PatternFill("solid", fgColor=GEAR_SET_FILLS["evolver"])
_GEAR_SET_FILL_CACHE: dict[str, PatternFill] = {
    key: PatternFill("solid", fgColor=color) for key, color in GEAR_SET_FILLS.items()
}
_GEAR_SET_FILL_CACHE["evolver"] = EVOLVER_FILL

# Gear T-Level tier code buckets (muted — readable with COLOR_TEXT)
TIER_COLOR_GREEN = "2D4A38"
TIER_COLOR_YELLOW = "4A4528"
TIER_COLOR_ORANGE = "4A3520"
TIER_COLOR_RED = "4A2830"

DEFAULT_TIER_BUCKET_COLORS: dict[str, str] = {
    "green": TIER_COLOR_GREEN,
    "yellow": TIER_COLOR_YELLOW,
    "orange": TIER_COLOR_ORANGE,
    "red": TIER_COLOR_RED,
    "evolver": GEAR_SET_FILLS["evolver"],
}

_TIER_BUCKET_FILLS: dict[str, PatternFill] = {
    TIER_COLOR_GREEN: PatternFill("solid", fgColor=TIER_COLOR_GREEN),
    TIER_COLOR_YELLOW: PatternFill("solid", fgColor=TIER_COLOR_YELLOW),
    TIER_COLOR_ORANGE: PatternFill("solid", fgColor=TIER_COLOR_ORANGE),
    TIER_COLOR_RED: PatternFill("solid", fgColor=TIER_COLOR_RED),
}

# Shared status fills for Type 7/8 / Raid BiS (muted dark-mode palette)
STATUS_FILL_BIS = PatternFill("solid", fgColor=TIER_COLOR_GREEN)
STATUS_FILL_UPGRADE = PatternFill("solid", fgColor=TIER_COLOR_YELLOW)
STATUS_FILL_EMPTY = PatternFill("solid", fgColor=TIER_COLOR_RED)
STATUS_FILL_UNKNOWN = PatternFill("solid", fgColor="283850")
STATUS_FILL_NEUTRAL = PatternFill("solid", fgColor="374151")
STATUS_FILLS: dict[str, PatternFill] = {
    "bis": STATUS_FILL_BIS,
    "upgrade": STATUS_FILL_UPGRADE,
    "empty": STATUS_FILL_EMPTY,
    "unknown": STATUS_FILL_UNKNOWN,
    "weapon": STATUS_FILL_NEUTRAL,
    "no_fit": STATUS_FILL_NEUTRAL,
}

FILL_ANNIVERSARY = PatternFill("solid", fgColor="3A3350")
FILL_OWNED = PatternFill("solid", fgColor=TIER_COLOR_GREEN)

# Missing Spells sheet
SPELL_BLOCK_HEADER_COLORS: dict[str, str] = {
    "121-125": "283850",
    "126-130": "3A3350",
    "nos": "2C4530",
    "ls": "283850",
    "tob": "3A4228",
    "sor": "3A3350",
}
SPELL_TIER_COLORS: dict[str, str] = {
    "Minor": "2A3545",
    "Lesser": "2E3A4C",
    "Median": "343C52",
    "Greater": "3A3850",
    "Glowing": "453848",
}
_SPELL_BLOCK_HEADER_DEFAULT = "2D2D32"
_SPELL_TIER_DEFAULT = "252528"
_SPELL_BLOCK_HEADER_FILL_CACHE: dict[str, PatternFill] = {
    key: PatternFill("solid", fgColor=color)
    for key, color in SPELL_BLOCK_HEADER_COLORS.items()
}
_SPELL_BLOCK_HEADER_FILL_CACHE[_SPELL_BLOCK_HEADER_DEFAULT] = PatternFill(
    "solid", fgColor=_SPELL_BLOCK_HEADER_DEFAULT
)
_SPELL_TIER_FILL_CACHE: dict[str, PatternFill] = {
    key: PatternFill("solid", fgColor=color) for key, color in SPELL_TIER_COLORS.items()
}
_SPELL_TIER_FILL_CACHE[_SPELL_TIER_DEFAULT] = PatternFill("solid", fgColor=_SPELL_TIER_DEFAULT)

FILL_SPELL_DETAIL = PatternFill("solid", fgColor="1A1A1E")
FILL_SPELL_DETAIL_ALT = PatternFill("solid", fgColor="222228")
FILL_SPELL_COUNT = PatternFill("solid", fgColor="2E3340")
COLOR_TEXT_ACCENT = "F0F2F8"
FONT_COUNT = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_ACCENT)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=COLOR_TEXT)
FONT_BLOCK_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT)
FONT_BLOCK_SUB = Font(name="Calibri", size=10, color=COLOR_TEXT_MUTED)

_LIGHT_TEXT_HEX = frozenset(
    {
        COLOR_TEXT.upper(),
        COLOR_TEXT_MUTED.upper(),
        COLOR_LINK.upper(),
        COLOR_TEXT_ACCENT.upper(),
        "EEF0F4",
        "F9FAFB",
        "FCA5A5",
        "8FD4A8",
        "E8C97A",
    }
)


def gear_set_fill(key: str) -> PatternFill:
    return _GEAR_SET_FILL_CACHE[key]


def spell_block_header_fill(block_label: str) -> PatternFill:
    return (
        _SPELL_BLOCK_HEADER_FILL_CACHE.get(block_label)
        or _SPELL_BLOCK_HEADER_FILL_CACHE[_SPELL_BLOCK_HEADER_DEFAULT]
    )


def spell_tier_fill(tier: str) -> PatternFill:
    return _SPELL_TIER_FILL_CACHE.get(tier) or _SPELL_TIER_FILL_CACHE[_SPELL_TIER_DEFAULT]


def resolved_tier_bucket_colors() -> dict[str, str]:
    """Five semantic bucket hex colors, with optional user overrides from settings."""
    from inventory_parser.character_column_order import load_tier_color_overrides

    colors = dict(DEFAULT_TIER_BUCKET_COLORS)
    colors.update(load_tier_color_overrides())
    return colors


def _pattern_fill_for_hex(hex_color: str) -> PatternFill:
    key = hex_color.upper()
    fill = _TIER_BUCKET_FILLS.get(key)
    if fill is None:
        fill = PatternFill("solid", fgColor=key)
        _TIER_BUCKET_FILLS[key] = fill
    return fill


def evolver_fill() -> PatternFill:
    """PatternFill for Evolver cells using the resolved bucket color."""
    return _pattern_fill_for_hex(resolved_tier_bucket_colors()["evolver"])


def tier_code_fill_color(code: str) -> str:
    """Semantic Gear T-Level background color for a tier code."""
    colors = resolved_tier_bucket_colors()
    if code == "SOR-R2":
        return colors["green"]
    if code in ("SOR-R1", "ANI27"):
        return colors["yellow"]
    if code.startswith("TOB-"):
        return colors["orange"]
    return colors["red"]


def tier_code_fill(code: str) -> PatternFill:
    return _pattern_fill_for_hex(tier_code_fill_color(code))


def build_tier_code_colors() -> dict[str, str]:
    """Map every known tier code (plus Evolver / ???) to HTML theme hex colors."""
    from inventory_parser.evolver import EVOLVER_GAP_LABEL
    from inventory_parser.gear_tiers import GEAR_TIERS_NEWEST_FIRST, UNKNOWN_TIER_LABEL

    colors = {tier.code: tier_code_fill_color(tier.code) for tier in GEAR_TIERS_NEWEST_FIRST}
    colors[EVOLVER_GAP_LABEL] = resolved_tier_bucket_colors()["evolver"]
    colors[UNKNOWN_TIER_LABEL] = tier_code_fill_color(UNKNOWN_TIER_LABEL)
    return colors


_TIER_LEGEND_LABELS: dict[str, str] = {
    "green": "SOR-R2 (current SoR raid)",
    "yellow": "SOR-R1, ANI27",
    "orange": "All TOB tiers",
    "red": "LS, NoS, SOR group, ???, other",
    "evolver": "Evolver",
}


def tier_bucket_legend_rows() -> tuple[tuple[PatternFill, str], ...]:
    """Legend swatches for Team Gear / HTML footer (semantic tier buckets)."""
    colors = resolved_tier_bucket_colors()
    return (
        (_pattern_fill_for_hex(colors["green"]), _TIER_LEGEND_LABELS["green"]),
        (_pattern_fill_for_hex(colors["yellow"]), _TIER_LEGEND_LABELS["yellow"]),
        (_pattern_fill_for_hex(colors["orange"]), _TIER_LEGEND_LABELS["orange"]),
        (_pattern_fill_for_hex(colors["red"]), _TIER_LEGEND_LABELS["red"]),
        (_pattern_fill_for_hex(colors["evolver"]), _TIER_LEGEND_LABELS["evolver"]),
    )


def build_gear_legend() -> list[dict[str, str]]:
    """HTML footer legend entries matching tier bucket colors."""
    colors = resolved_tier_bucket_colors()
    return [
        {"key": key, "label": _TIER_LEGEND_LABELS[key], "color": colors[key]}
        for key in ("green", "yellow", "orange", "red", "evolver")
    ]


def tier_legend_entries() -> list[dict[str, str | bool]]:
    """Help-dialog rows: key, live color, default color, label, and isCustom flag."""
    colors = resolved_tier_bucket_colors()
    defaults = DEFAULT_TIER_BUCKET_COLORS
    is_custom = any(
        colors[key].upper() != defaults[key].upper()
        for key in ("green", "yellow", "orange", "red", "evolver")
    )
    return [
        {
            "key": key,
            "color": colors[key],
            "defaultColor": defaults[key],
            "label": _TIER_LEGEND_LABELS[key],
            "isCustom": is_custom,
        }
        for key in ("green", "yellow", "orange", "red", "evolver")
    ]


def _font_hex(font: Font | None) -> str | None:
    if font is None or font.color is None:
        return None
    color = font.color
    if getattr(color, "type", None) == "theme":
        return None
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return None
    hex_val = str(rgb).upper()
    if len(hex_val) == 8 and hex_val.startswith("FF"):
        hex_val = hex_val[2:]
    if hex_val in ("000000", "00000000"):
        return None
    return hex_val[-6:] if len(hex_val) >= 6 else hex_val


def _needs_light_text(font: Font | None) -> bool:
    """True when the cell still uses Excel's default/dark text on a dark sheet."""
    hex_val = _font_hex(font)
    if hex_val is None:
        return True
    return hex_val not in _LIGHT_TEXT_HEX


def finalize_dark_sheet(
    ws: Worksheet,
    *,
    pad_rows: int = SHEET_BACKGROUND_ROWS,
    pad_cols: int = SHEET_BACKGROUND_COLS,
) -> None:
    """Apply uniform dark-mode fills/fonts to a sheet's used range and pad chrome."""
    content_last_row = max(ws.max_row or 1, 1)
    content_last_col = max(ws.max_column or 1, 1)
    pad_rows = max(pad_rows, content_last_row + 2)
    pad_cols = max(pad_cols, content_last_col)

    for row in ws.iter_rows(
        min_row=1,
        max_row=content_last_row,
        min_col=1,
        max_col=content_last_col,
    ):
        for cell in row:
            if cell.fill is None or cell.fill.fill_type is None:
                cell.fill = FILL_ITEM_EMPTY if cell.value is not None else FILL_SHEET
            if cell.value is None:
                continue
            if not _needs_light_text(cell.font):
                continue
            underline = bool(cell.font and cell.font.underline)
            bold = bool(cell.font and cell.font.bold)
            italic = bool(cell.font and cell.font.italic)
            if cell.hyperlink or underline:
                cell.font = FONT_LINK
            elif bold:
                cell.font = FONT_HEADER
            elif italic:
                cell.font = FONT_MUTED
            else:
                cell.font = FONT_BODY

    for row_idx in range(1, pad_rows + 1):
        for col_idx in range(1, pad_cols + 1):
            if row_idx <= content_last_row and col_idx <= content_last_col:
                continue
            cell = ws.cell(row_idx, col_idx)
            if cell.fill is None or cell.fill.fill_type is None:
                cell.fill = FILL_SHEET


def apply_workbook_dark_mode(wb: Workbook) -> None:
    """Ensure every worksheet uses the same dark-mode chrome."""
    for ws in wb.worksheets:
        finalize_dark_sheet(ws)
