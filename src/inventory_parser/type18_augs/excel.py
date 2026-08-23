"""Excel export for Type 18/19 aug suggestions + catalog."""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from inventory_parser.items import EQRESOURCE_ITEM_URL
from inventory_parser.type18_augs.build import Type18Export
from inventory_parser.type18_augs.categories import HEROIC_STAT_KEYS
from inventory_parser.type18_augs.suggestions import is_caster_class

HEADER_FILL = PatternFill("solid", fgColor="1E2430")
HEADER_FONT = Font(color="EEF0F4", bold=True)
LINK_FONT = Font(color="8CB4FF", underline="single")
ANNIV_FILL = PatternFill("solid", fgColor="3B2F1A")
OWNED_FILL = PatternFill("solid", fgColor="1A3B2A")

SHEET_NAME = "Type 18-19 Augs"
CATALOG_SHEET_NAME = "Type 18-19 Catalog"

_SUGGEST_HEADERS = (
    "Character",
    "Class",
    "Class abbr",
    "Priority",
    "#",
    "Guide name",
    "Suggested",
    "Type",
    "Category",
    "Anniversary",
    "Alternative (non-anniv)",
    "Alt type",
    "Owned",
    "AC / Mana",
    "HP / Spell Dmg",
    "Mana",
    "Spell Damage",
    "End",
    "HStr",
    "HSta",
    "HInt",
    "HWis",
    "HAgi",
    "HDex",
    "HCha",
)

_CATALOG_HEADERS = (
    "Type",
    "Category",
    "Lore group",
    "Item Lore",
    "Anniversary",
    "Name",
    "AC",
    "HP",
    "Mana",
    "Spell Damage",
    "End",
    "HStr",
    "HSta",
    "HInt",
    "HWis",
    "HAgi",
    "HDex",
    "HCha",
)


def _link_name(cell, name: str, item_id: int | None) -> None:
    cell.value = name
    if item_id and item_id > 0:
        cell.hyperlink = EQRESOURCE_ITEM_URL.format(item_id=item_id)
        cell.font = LINK_FONT


def _autosize_columns(ws: Worksheet, *, max_width: float = 48.0) -> None:
    """Set column widths from cell text length (fit content)."""
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_width, max(6.0, longest + 2.0))


def _write_suggest_stats(ws, row_idx: int, stats: dict, *, caster: bool) -> None:
    ac = int(stats.get("ac", 0) or 0)
    hp = int(stats.get("hp", 0) or 0)
    mana = int(stats.get("mana", 0) or 0)
    spell_dmg = int(stats.get("spell_damage", 0) or 0)
    # Dual-purpose columns: casters see Mana / Spell Damage in the lead pair.
    ws.cell(row_idx, 14, mana if caster else ac)
    ws.cell(row_idx, 15, spell_dmg if caster else hp)
    ws.cell(row_idx, 16, mana)
    ws.cell(row_idx, 17, spell_dmg)
    ws.cell(row_idx, 18, int(stats.get("endurance", 0) or 0))
    for col, key in enumerate(HEROIC_STAT_KEYS, start=19):
        ws.cell(row_idx, col, int(stats.get(key, 0) or 0))


def _suggestion_owned(sug, *, owned_ids: set[int], owned_names: set[str]) -> bool:
    if sug is None:
        return False
    if sug.item_id > 0 and sug.item_id in owned_ids:
        return True
    name = (sug.name or "").casefold()
    return bool(name) and name in owned_names


def _write_catalog_stats(ws, row_idx: int, start_col: int, stats: dict) -> None:
    values = [
        int(stats.get("ac", 0) or 0),
        int(stats.get("hp", 0) or 0),
        int(stats.get("mana", 0) or 0),
        int(stats.get("spell_damage", 0) or 0),
        int(stats.get("endurance", 0) or 0),
    ]
    values.extend(int(stats.get(k, 0) or 0) for k in HEROIC_STAT_KEYS)
    for col, val in enumerate(values, start=start_col):
        ws.cell(row_idx, col, val)


def _append_suggestions_sheet(wb, bundle: Type18Export) -> None:
    ws = wb.create_sheet(SHEET_NAME)
    for col, h in enumerate(_SUGGEST_HEADERS, start=1):
        cell = ws.cell(1, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    by_class = {b.class_abbr: b for b in bundle.suggestions}
    row_idx = 2

    def write_block(character_label: str, block, *, owned_ids: set[int], owned_names: set[str]) -> None:
        nonlocal row_idx
        caster = bool(block.caster_stats) or is_caster_class(block.class_abbr)
        for row in (*block.primary, *block.optional, *block.fortification):
            sug = row.suggested
            alt = row.alternative
            ws.cell(row_idx, 1, character_label)
            ws.cell(row_idx, 2, block.class_name)
            ws.cell(row_idx, 3, block.class_abbr)
            ws.cell(row_idx, 4, row.priority)
            ws.cell(row_idx, 5, row.rank)
            ws.cell(row_idx, 6, row.guide_name)
            if sug is not None:
                _link_name(ws.cell(row_idx, 7), sug.name, sug.item_id)
                ws.cell(row_idx, 8, sug.type_label)
                ws.cell(row_idx, 9, sug.category)
                ann = ws.cell(row_idx, 10, "Yes" if sug.anniversary else "")
                if sug.anniversary:
                    ann.fill = ANNIV_FILL
                _write_suggest_stats(ws, row_idx, sug.stats, caster=caster)
            else:
                ws.cell(row_idx, 7, row.guide_name)
                ann = ws.cell(
                    row_idx,
                    10,
                    "Yes"
                    if "enduring harmony" in row.guide_name.casefold()
                    or "jubilation" in row.guide_name.casefold()
                    or "selenelion" in row.guide_name.casefold()
                    else "",
                )
                if ann.value == "Yes":
                    ann.fill = ANNIV_FILL
            if alt is not None:
                _link_name(ws.cell(row_idx, 11), alt.name, alt.item_id)
                ws.cell(row_idx, 12, alt.type_label)
            owned = _suggestion_owned(
                sug, owned_ids=owned_ids, owned_names=owned_names
            )
            owned_cell = ws.cell(row_idx, 13, "Yes" if owned else "")
            if owned:
                owned_cell.fill = OWNED_FILL
            row_idx += 1

    if bundle.characters:
        for ch in bundle.characters:
            block = by_class.get(ch.class_abbr)
            if block is None:
                continue
            write_block(
                ch.display_name or ch.name,
                block,
                owned_ids=ch.owned_ids,
                owned_names=ch.owned_names,
            )
    else:
        for block in bundle.suggestions:
            write_block(
                "",
                block,
                owned_ids=set(),
                owned_names=set(),
            )

    if row_idx == 2:
        ws.cell(2, 1, "No Type 18/19 class suggestions available.")

    note_row = row_idx + 1
    if bundle.cheat_sheet_url:
        ws.cell(note_row, 1, "Class cheat sheet:")
        link = ws.cell(note_row, 2, bundle.cheat_sheet_url)
        link.hyperlink = bundle.cheat_sheet_url
        link.font = LINK_FONT
        note_row += 1
    ws.cell(
        note_row,
        1,
        "Owned is Yes when that character’s inventory contains the suggested aug.",
    )
    note_row += 1
    ws.cell(
        note_row,
        1,
        "AC/Mana and HP/Spell Dmg columns use Mana + Spell Damage for caster classes.",
    )
    note_row += 1
    ws.cell(
        note_row,
        1,
        "Anniversary suggestions always include a non-anniversary Alternative.",
    )

    ws.row_dimensions[1].height = 20
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_SUGGEST_HEADERS))}{max(row_idx - 1, 1)}"
    )
    _autosize_columns(ws)


def _append_catalog_sheet(wb, bundle: Type18Export) -> None:
    ws = wb.create_sheet(CATALOG_SHEET_NAME)
    for col, h in enumerate(_CATALOG_HEADERS, start=1):
        cell = ws.cell(1, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    row_idx = 2
    for e in bundle.entries:
        stats = e.stats or {}
        ws.cell(
            row_idx,
            1,
            e.type_label or ("18/19" if e.aug_type == 18 else str(e.aug_type)),
        )
        ws.cell(row_idx, 2, e.category)
        ws.cell(row_idx, 3, e.lore_group or "")
        ws.cell(row_idx, 4, e.item_lore or "")
        anniv_cell = ws.cell(row_idx, 5, "Yes" if e.anniversary else "")
        if e.anniversary:
            anniv_cell.fill = ANNIV_FILL
        _link_name(ws.cell(row_idx, 6), e.name or "", e.item_id)
        _write_catalog_stats(ws, row_idx, 7, stats)
        row_idx += 1

    if row_idx == 2:
        ws.cell(2, 1, "No Type 18/19 augs found in the catalog.")

    ws.row_dimensions[1].height = 20
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_CATALOG_HEADERS))}{max(row_idx - 1, 1)}"
    )
    _autosize_columns(ws)


def append_type18_sheet(wb, bundle: Type18Export) -> None:
    """Append suggestion + catalog sheets for Type 18/19 augs."""
    _append_suggestions_sheet(wb, bundle)
    _append_catalog_sheet(wb, bundle)
