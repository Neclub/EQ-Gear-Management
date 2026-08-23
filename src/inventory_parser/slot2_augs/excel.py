"""Excel export for Slot2 aug reports."""

from __future__ import annotations

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from inventory_parser import APP_NAME, __version__
from inventory_parser.excel_theme import (
    FILL_HEADER,
    FONT_HEADER,
    FONT_LINK,
    FONT_LINK_ON_STATUS,
    FONT_MUTED,
    FONT_ON_STATUS,
    STATUS_FILLS,
)
from inventory_parser.slot2_augs.compare import NEEDS_UPGRADE_STATUSES, REPORT_ROW_STATUSES
from inventory_parser.slot2_augs.build import Slot2Export
from inventory_parser.slot2_augs.html import format_catalog_fetched_at
from inventory_parser.slot2_augs.profiles import PROFILE_FOCUS_LABEL
from inventory_parser.items import EQRESOURCE_ITEM_URL


def _set_item_cell(cell, name: str | None, item_id: int | None, *, on_fill: bool) -> None:
    label = name or "(empty)"
    cell.value = label
    if name and item_id and item_id > 0:
        cell.hyperlink = EQRESOURCE_ITEM_URL.format(item_id=item_id)
        cell.font = FONT_LINK_ON_STATUS if on_fill else FONT_LINK
    elif on_fill:
        cell.font = FONT_ON_STATUS


def append_slot2_sheets(wb, bundle: Slot2Export) -> None:
    """Append Slot2 sheets onto an existing EQ Gear Management workbook."""
    show_server = bundle.show_server_in_columns

    # --- Sheet 1: Stat summary (totals if suggested augs equipped) ---
    ws_sum = wb.create_sheet("Stat Summary")
    ws_sum.sheet_properties.tabColor = "2D4A38"
    sum_headers = [
        "Character",
        "Slots changed",
        "Focus",
        "AC",
        "HP",
        "ATK",
        "Heal Amount",
        "Spell Damage",
        "Clairvoyance",
    ]
    for col, h in enumerate(sum_headers, start=1):
        cell = ws_sum.cell(1, col, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    sum_row = 2
    for i, ch in enumerate(bundle.characters):
        server = ch.server
        if bundle.roster and i < len(bundle.roster):
            server = bundle.roster[i].server or server
        label = f"{ch.character} ({server})" if show_server and server else ch.character
        focus_label = PROFILE_FOCUS_LABEL.get(ch.profile, "HDex")
        summary = ch.stat_summary or {}
        values = [
            label,
            ch.slots_changed,
            f"{int(summary.get('focus', 0)):+d} {focus_label}",
            int(summary.get("ac", 0)),
            int(summary.get("hp", 0)),
            int(summary.get("atk", 0)),
            int(summary.get("heal_amount", 0)),
            int(summary.get("spell_damage", 0)),
            int(summary.get("clairvoyance", 0)),
        ]
        for col, val in enumerate(values, start=1):
            ws_sum.cell(sum_row, col, val)
        sum_row += 1

    if sum_row == 2:
        ws_sum.cell(2, 1, "No suggested upgrades to summarize.")

    for i, w in enumerate([22, 14, 14, 8, 8, 8, 12, 12, 12], start=1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.row_dimensions[1].height = 20

    # --- Sheet 2: All current type 7/8 augs ---
    ws = wb.create_sheet("Augs")
    ws.sheet_properties.tabColor = "283850"

    headers = [
        "Character",
        "Slot",
        "Current",
        "Upgrade to",
        "Owned?",
        "Expansion",
        "Status",
        "Note",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    row_i = 2
    for i, ch in enumerate(bundle.characters):
        server = ch.server
        if bundle.roster and i < len(bundle.roster):
            server = bundle.roster[i].server or server
        label = f"{ch.character} ({server})" if show_server and server else ch.character
        for cmp_ in ch.comparisons:
            if cmp_.status not in REPORT_ROW_STATUSES:
                continue
            show_upgrade = cmp_.status in NEEDS_UPGRADE_STATUSES
            fill = STATUS_FILLS.get(cmp_.status)
            on_fill = fill is not None
            ws.cell(row_i, 1, label)
            ws.cell(row_i, 2, cmp_.gear_slot)
            _set_item_cell(
                ws.cell(row_i, 3),
                cmp_.current_name,
                cmp_.current_id,
                on_fill=on_fill,
            )
            rec_cell = ws.cell(row_i, 4)
            if show_upgrade and cmp_.recommended_name:
                _set_item_cell(
                    rec_cell,
                    cmp_.recommended_name,
                    cmp_.recommended_id,
                    on_fill=on_fill,
                )
            else:
                rec_cell.value = ""
                if on_fill:
                    rec_cell.font = FONT_ON_STATUS

            owned_cell = ws.cell(row_i, 5)
            if show_upgrade and cmp_.move_from_slot:
                owned_cell.value = f"Move from {cmp_.move_from_slot}"
            elif show_upgrade and cmp_.recommended_owned is not None:
                if cmp_.recommended_owned:
                    owned_cell.value = "Owned"
                elif cmp_.craft_component_owned and cmp_.craft_component_name:
                    owned_cell.value = (
                        f"Need to farm (have {cmp_.craft_component_name})"
                    )
                else:
                    owned_cell.value = "Need to farm"
            else:
                owned_cell.value = ""

            exp_cell = ws.cell(row_i, 6)
            exp_cell.value = (
                cmp_.recommended_expansion
                if show_upgrade and cmp_.recommended_expansion
                else ""
            )

            status_cell = ws.cell(row_i, 7, cmp_.status)
            note_cell = ws.cell(row_i, 8, cmp_.note)
            if fill:
                for col in range(1, 9):
                    cell = ws.cell(row_i, col)
                    cell.fill = fill
                    if col not in (3, 4):
                        cell.font = FONT_ON_STATUS
                _set_item_cell(
                    ws.cell(row_i, 3),
                    cmp_.current_name,
                    cmp_.current_id,
                    on_fill=True,
                )
                if show_upgrade and cmp_.recommended_name:
                    _set_item_cell(
                        ws.cell(row_i, 4),
                        cmp_.recommended_name,
                        cmp_.recommended_id,
                        on_fill=True,
                    )
                else:
                    ws.cell(row_i, 4).value = ""
                    ws.cell(row_i, 4).font = FONT_ON_STATUS
                rec_cell.alignment = Alignment(wrap_text=True, vertical="top")
                status_cell.alignment = Alignment(vertical="top")
                note_cell.font = FONT_ON_STATUS
            row_i += 1

    if row_i == 2:
        ws.cell(2, 1, "No graded type 7/8 slots found.")

    widths = [18, 12, 36, 36, 14, 20, 10, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20

    # --- Sheet 3: Need to Farm ---
    ws_farm = wb.create_sheet("Need to Farm")
    ws_farm.sheet_properties.tabColor = "4A2830"
    farm_headers = ["Character", "Slot", "Aug", "Expansion"]
    for col, h in enumerate(farm_headers, start=1):
        cell = ws_farm.cell(1, col, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    farm_row = 2
    for entry in bundle.farm_list:
        label = (
            f"{entry.character} ({entry.server})"
            if show_server and entry.server
            else entry.character
        )
        ws_farm.cell(farm_row, 1, label)
        ws_farm.cell(farm_row, 2, entry.gear_slot)
        aug_label = entry.name
        if entry.craft_component_name:
            tag = "Have" if entry.craft_component_owned else "Need"
            aug_label = f"{entry.name}  [{tag} {entry.craft_component_name}]"
        _set_item_cell(ws_farm.cell(farm_row, 3), aug_label, entry.item_id, on_fill=False)
        ws_farm.cell(farm_row, 4, entry.expansion or "")
        farm_row += 1

    if farm_row == 2:
        ws_farm.cell(2, 1, "No recommended upgrades to farm.")

    for i, w in enumerate([18, 12, 52, 22], start=1):
        ws_farm.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 4: Ranked reference ---
    ws2 = wb.create_sheet("Ranked Augs")
    ws2.sheet_properties.tabColor = "252528"
    meta = [
        f"Profile: {bundle.profile_label}",
        f"Artisan's Prize owned: {'Yes' if bundle.artisans_prize_owned else 'No'} (from inventory)",
        f"Anniversary augs: {'Included' if bundle.include_anniversary else 'Excluded'}",
        f"Catalog fetched: {format_catalog_fetched_at(bundle.catalog.fetched_at)}"
        + (" (cache)" if bundle.catalog.from_cache else ""),
        f"{APP_NAME} {__version__}",
    ]
    if bundle.warnings:
        meta.extend(bundle.warnings)
    ws2.cell(1, 1, " | ".join(meta)).font = FONT_MUTED

    focus_header = "Focus"
    ref_headers = [
        "#",
        "Name",
        "ID",
        "Profile",
        focus_header,
        "AC",
        "HP",
        "ATK",
        "Heal",
        "Spell Dmg",
        "Clairvoyance",
        "Slots",
        "Source",
    ]
    for col, h in enumerate(ref_headers, start=1):
        cell = ws2.cell(3, col, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    for i, aug in enumerate(bundle.ranked_augs, start=1):
        row = i + 3
        slot_desc = aug.slot_text or (
            "Ear only"
            if aug.ear_only
            else (
                f"All except {', '.join(sorted(aug.excluded_bases))}"
                if aug.excluded_bases
                else "All"
            )
        )
        stats = aug.effective_stats()
        values = [
            i,
            aug.name,
            aug.item_id,
            PROFILE_FOCUS_LABEL.get(aug.profile, aug.profile),
            aug.focus_heroic,
            aug.ac,
            aug.hp,
            aug.atk or int(stats.get("atk", 0)),
            int(stats.get("heal_amount", 0)),
            int(stats.get("spell_damage", 0)),
            int(stats.get("clairvoyance", 0)),
            slot_desc,
            aug.source,
        ]
        for col, val in enumerate(values, start=1):
            ws2.cell(row, col, val)
        name_cell = ws2.cell(row, 2)
        name_cell.hyperlink = EQRESOURCE_ITEM_URL.format(item_id=aug.item_id)
        name_cell.font = FONT_LINK

    widths2 = [4, 40, 10, 8, 8, 6, 8, 6, 6, 8, 10, 40, 40]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 5: Legend ---
    ws3 = wb.create_sheet("Aug Legend")
    ws3.sheet_properties.tabColor = "2D2D32"
    ws3.cell(1, 1, "Status").font = FONT_HEADER
    ws3.cell(1, 2, "Meaning").font = FONT_HEADER
    legend = [
        ("upgrade", "Better type 7/8 aug available"),
        ("empty", "Type 7/8 hole is empty"),
        ("unknown", "Current aug not in raidloot or EQ Resource"),
        ("bis", "Already BiS — Upgrade to left blank"),
        ("no_fit", "Ignored / no catalog fit — omitted from Augs sheet"),
    ]
    for i, (status, meaning) in enumerate(legend, start=2):
        cell = ws3.cell(i, 1, status)
        cell.fill = STATUS_FILLS[status]
        cell.font = FONT_ON_STATUS
        ws3.cell(i, 2, meaning)
    ws3.cell(8, 1, "Owned?").font = FONT_HEADER
    ws3.cell(8, 2, "Recommended upgrade present in inventory, or move from another slot")
    ws3.cell(9, 1, "Need to farm").font = FONT_HEADER
    ws3.cell(9, 2, "Recommended upgrade missing from inventory (see Need to Farm sheet)")
    ws3.cell(10, 1, "Have Focus/Ore").font = FONT_HEADER
    ws3.cell(
        10,
        2,
        "Need to farm, but the matching Focus of Fortitude (or Ensanguined ore) is in bags/bank",
    )
    ws3.cell(10, 1, "Move from").font = FONT_HEADER
    ws3.cell(
        10,
        2,
        "Recommended aug is already equipped in another slot (e.g. Range → Head)",
    )
    ws3.cell(11, 1, "Stat Summary").font = FONT_HEADER
    ws3.cell(
        11,
        2,
        "Totals if Upgrade/Empty slots use the suggested augs (BiS/unknown excluded)",
    )
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 72
