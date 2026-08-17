from pathlib import Path

from inventory_parser.achievement_files import (
    achievement_character_key,
    collect_achievement_paths,
    discover_achievements_files,
    is_achievements_file,
    parse_achievements_filename,
)
from inventory_parser.achievement_parser import (
    EVERQUEST_BASE_LABEL,
    clean_quest_name,
    clean_raid_objective,
    format_expansion_label,
    parse_achievements_file,
    parse_quest_parent,
    parse_raid_parent,
    raid_header_name,
    split_collection_name,
)
from inventory_parser.achievement_report import build_achievement_report
from inventory_parser.team_report import build_team_report
from inventory_parser.excel_export import (
    ACHIEVEMENT_SUMMARY_SHEET_NAME,
    MISSING_COLLECTIONS_SHEET_NAME,
    QUESTS_SHEET_NAME,
    RAID_ACHIEVEMENTS_SHEET_NAME,
    write_team_workbook,
)
from inventory_parser.missing_spells import split_input_paths

EXAMPLES = Path(__file__).resolve().parents[1] / "Examples"
ACHIEVEMENTS = EXAMPLES / "Achievements"
SHAMLUB_ACH = ACHIEVEMENTS / "Shamlub_xegony-Achievements.txt"


def test_is_achievements_file() -> None:
    assert is_achievements_file(SHAMLUB_ACH)
    assert not is_achievements_file(EXAMPLES / "Shamlub_bristle-Inventory.txt")


def test_parse_achievements_filename() -> None:
    assert parse_achievements_filename(SHAMLUB_ACH) == ("Shamlub", "xegony")
    assert achievement_character_key(SHAMLUB_ACH) == "Shamlub_xegony"


def test_split_input_paths_includes_achievements() -> None:
    inv = EXAMPLES / "Shamlub_bristle-Inventory.txt"
    inventory_paths, spell_paths, achievement_paths = split_input_paths([inv, SHAMLUB_ACH])
    assert inventory_paths == [inv]
    assert spell_paths == []
    assert achievement_paths == [SHAMLUB_ACH]


def test_split_collection_name() -> None:
    collection, zone = split_collection_name("The Depths of Fear (Rain of Fear)")
    assert collection == "The Depths of Fear"
    assert zone == "Rain of Fear"


def test_parse_missing_collection_items() -> None:
    parsed = parse_achievements_file(SHAMLUB_ACH)
    missing = parsed.missing_collections
    assert missing
    strange = next(item for item in missing if item.item == "Strange Black Rock")
    assert strange.section == "Rain of Fear"
    assert strange.zone == "Rain of Fear"
    assert strange.collection == "The Depths of Fear"
    assert strange.owned == 0
    assert strange.total == 1


def test_section_summary_counts_top_level_achievements() -> None:
    parsed = parse_achievements_file(SHAMLUB_ACH)
    general = next(s for s in parsed.section_summaries if s.section == "General")
    assert general.completed > 0
    assert general.total == general.completed + general.incomplete


def test_discover_achievements_files() -> None:
    files = discover_achievements_files(ACHIEVEMENTS)
    assert SHAMLUB_ACH.resolve() in [p.resolve() for p in files]


def test_parse_raid_parent_and_clean_raid_objective() -> None:
    assert parse_raid_parent("Conqueror of Labyrinth of Spite: Echo of Hate") == (
        "Labyrinth of Spite",
        "Echo of Hate",
    )
    assert parse_raid_parent("Vanquisher of Labyrinth of Spite: Echo of Hate") == (
        "Labyrinth of Spite",
        "Echo of Hate",
    )
    assert parse_raid_parent("Conqueror of The Plane of Sky") == ("The Plane of Sky", "")
    assert parse_raid_parent("Echo of Hate: Give in to Greed") is None
    assert raid_header_name("Labyrinth of Spite", "Echo of Hate") == (
        "Conqueror of Labyrinth of Spite: Echo of Hate"
    )
    assert clean_raid_objective("Echo of Hate: Give in to Greed", "Echo of Hate") == (
        "Give in to Greed"
    )
    assert clean_raid_objective("Echo of Hate: Enraged", "Echo of Hate") == "Enraged"
    parsed = parse_achievements_file(SHAMLUB_ACH)
    rain_fear = [
        raid for raid in parsed.missing_raid_achievements if raid.section == "Rain of Fear"
    ]
    assert rain_fear
    assert any(
        raid.raid == "Conqueror of Vulak`Aerr" and raid.objective == "None Shall Pass"
        for raid in rain_fear
    )
    assert all(raid.objective for raid in parsed.missing_raid_achievements)
    assert all(raid.section for raid in parsed.missing_raid_achievements)
    assert any(not raid.complete for raid in rain_fear)
    assert any(raid.complete for raid in parsed.missing_raid_achievements)


def test_raid_achievements_exclude_completed() -> None:
    parsed = parse_achievements_file(SHAMLUB_ACH)
    completed = {
        (raid.raid, raid.objective)
        for raid in parsed.missing_raid_achievements
        if raid.section == "Shattering of Ro"
        and raid.raid == "Conqueror of Candlemaker's Workshop: Waxwork Abolishion"
    }
    assert ("Conqueror of Candlemaker's Workshop: Waxwork Abolishion", "") not in completed


def test_raid_report_keeps_partial_objectives() -> None:
    inv = EXAMPLES / "Shamlub_bristle-Inventory.txt"
    report = build_team_report([inv])
    ach_report = build_achievement_report(
        report,
        achievement_paths={"shamlub_bristle": SHAMLUB_ACH},
    )
    assert ach_report is not None
    vanquisher = [
        row
        for row in ach_report.raid_achievements
        if row.raid == "Conqueror of Candlemaker's Workshop: Waxwork Abolishion"
    ]
    statuses = {row.objective: row.status for row in vanquisher}
    assert statuses["Wax Free Zone"] == "Done"
    assert statuses["Waxing Intensity"] == "Missing"
    assert "Waxwork Abolishion" not in statuses

    echo = [
        row
        for row in ach_report.raid_achievements
        if row.raid == "Conqueror of Labyrinth of Spite: Echo of Hate"
    ]
    assert {row.objective for row in echo} == {
        "Give in to Greed",
        "What It Wants",
        "Unfocused",
        "Enraged",
    }
    assert all(row.event == "Echo of Hate" for row in echo)
    assert all(row.status == "Missing" for row in echo)
    assert not any(
        row.raid.startswith("Vanquisher of") for row in ach_report.raid_achievements
    )


def test_format_expansion_label() -> None:
    assert format_expansion_label("Shattering of Ro") == "Shattering of Ro (2025)"
    assert format_expansion_label("Rain of Fear") == "Rain of Fear (2012)"
    assert format_expansion_label("EverQuest") == EVERQUEST_BASE_LABEL
    assert format_expansion_label("General") == "General"


def test_build_achievement_report_includes_sorted_raid_rows() -> None:
    inv = EXAMPLES / "Shamlub_bristle-Inventory.txt"
    report = build_team_report([inv])
    ach_report = build_achievement_report(
        report,
        achievement_paths={"shamlub_bristle": SHAMLUB_ACH},
    )
    assert ach_report is not None
    assert ach_report.raid_achievements
    expansions = [format_expansion_label(row.expansion) for row in ach_report.raid_achievements]
    assert expansions.index("Shattering of Ro (2025)") < expansions.index("Rain of Fear (2012)")
    assert expansions.index("Rain of Fear (2012)") < expansions.index(EVERQUEST_BASE_LABEL)


def test_build_achievement_report_with_explicit_path() -> None:
    inv = EXAMPLES / "Shamlub_bristle-Inventory.txt"
    report = build_team_report([inv])
    ach_report = build_achievement_report(
        report,
        achievement_paths={"shamlub_bristle": SHAMLUB_ACH},
    )
    assert ach_report is not None
    strange_row = next(
        row for row in ach_report.missing_collections if row.missing_item == "Strange Black Rock"
    )
    assert strange_row.char_has == ""
    assert any(row.section == "General" for row in ach_report.summaries)
    assert any(row.expansion == "Rain of Fear" for row in ach_report.raid_achievements)


def test_missing_collections_char_has_from_team_inventory(tmp_path: Path) -> None:
    holder_inv = tmp_path / "Tanklub_bristle-Inventory.txt"
    holder_inv.write_text(
        "Location\tName\tID\tCount\tSlots\n"
        "General1-Slot1\tStrange Black Rock\t12345\t1\t6\n",
        encoding="utf-8",
    )
    missing_inv = EXAMPLES / "Shamlub_bristle-Inventory.txt"
    report = build_team_report([missing_inv, holder_inv])
    ach_report = build_achievement_report(
        report,
        achievement_paths={"shamlub_bristle": SHAMLUB_ACH},
    )
    assert ach_report is not None
    strange_row = next(
        row for row in ach_report.missing_collections if row.missing_item == "Strange Black Rock"
    )
    assert strange_row.char_has == "Tanklub"


def test_achievement_report_once_per_character_across_personas(tmp_path: Path) -> None:
    """Personas share achievements/collections — emit one row set per character."""
    pal_inv = tmp_path / "Shamlub_bristle-PAL-Inventory.txt"
    shd_inv = tmp_path / "Shamlub_bristle-SHD-Inventory.txt"
    pal_inv.write_text(
        "Location\tName\tID\tCount\tSlots\nHead\tPAL Helm\t1\t1\t0\n",
        encoding="utf-8",
    )
    shd_inv.write_text(
        "Location\tName\tID\tCount\tSlots\n"
        "Head\tSHD Helm\t2\t1\t0\n"
        "General1-Slot1\tStrange Black Rock\t12345\t1\t6\n",
        encoding="utf-8",
    )
    report = build_team_report([pal_inv, shd_inv])
    assert len(report.characters) == 2

    parsed = parse_achievements_file(SHAMLUB_ACH)
    ach_report = build_achievement_report(
        report,
        achievement_paths={"shamlub_bristle": SHAMLUB_ACH},
    )
    assert ach_report is not None
    assert len(ach_report.missing_collections) == len(parsed.missing_collections)
    assert len(ach_report.summaries) == len(
        [s for s in parsed.section_summaries if s.total > 0]
    )
    assert {row.character for row in ach_report.missing_collections} == {"Shamlub"}
    assert {row.character for row in ach_report.summaries} == {"Shamlub"}
    assert {row.character for row in ach_report.raid_achievements} == {"Shamlub"}
    assert ach_report.raid_achievements
    assert ach_report.quests
    assert {row.character for row in ach_report.quests} == {"Shamlub"}

    strange_row = next(
        row for row in ach_report.missing_collections if row.missing_item == "Strange Black Rock"
    )
    # Char Has lists the character once even when multiple persona dumps hold the item.
    assert strange_row.char_has == "Shamlub"


def test_char_has_dedupes_personas_of_holder(tmp_path: Path) -> None:
    holder_pal = tmp_path / "Tanklub_bristle-PAL-Inventory.txt"
    holder_shd = tmp_path / "Tanklub_bristle-SHD-Inventory.txt"
    for path in (holder_pal, holder_shd):
        path.write_text(
            "Location\tName\tID\tCount\tSlots\n"
            "General1-Slot1\tStrange Black Rock\t12345\t1\t6\n",
            encoding="utf-8",
        )
    missing_inv = EXAMPLES / "Shamlub_bristle-Inventory.txt"
    report = build_team_report([missing_inv, holder_pal, holder_shd])
    ach_report = build_achievement_report(
        report,
        achievement_paths={"shamlub_bristle": SHAMLUB_ACH},
    )
    assert ach_report is not None
    strange_row = next(
        row for row in ach_report.missing_collections if row.missing_item == "Strange Black Rock"
    )
    assert strange_row.char_has == "Tanklub"


def test_collect_achievement_paths_from_achievementdata(tmp_path: Path) -> None:
    inv = EXAMPLES / "Shamlub_bristle-Inventory.txt"
    data_dir = tmp_path / "AchievementData"
    data_dir.mkdir()
    copied = data_dir / "Shamlub_bristle-Achievements.txt"
    copied.write_text(SHAMLUB_ACH.read_text(encoding="utf-8"), encoding="utf-8")
    work_inv = tmp_path / inv.name
    work_inv.write_text(inv.read_text(encoding="utf-8"), encoding="utf-8")

    discovered = collect_achievement_paths([work_inv])
    assert discovered["shamlub_bristle"].resolve() == copied.resolve()


def test_achievement_sheets_in_workbook(tmp_path: Path) -> None:
    inv = EXAMPLES / "Shamlub_bristle-Inventory.txt"
    report = build_team_report([inv])
    ach_report = build_achievement_report(
        report,
        achievement_paths={"shamlub_bristle": SHAMLUB_ACH},
    )
    assert ach_report is not None
    out = tmp_path / "crew.xlsx"
    write_team_workbook(report, out, achievement_report=ach_report)

    from openpyxl import load_workbook

    wb = load_workbook(out, data_only=True)
    assert MISSING_COLLECTIONS_SHEET_NAME in wb.sheetnames
    assert ACHIEVEMENT_SUMMARY_SHEET_NAME in wb.sheetnames
    assert QUESTS_SHEET_NAME in wb.sheetnames
    assert RAID_ACHIEVEMENTS_SHEET_NAME in wb.sheetnames
    missing_ws = wb[MISSING_COLLECTIONS_SHEET_NAME]
    assert missing_ws.cell(1, 5).value == "Missing Item"
    assert missing_ws.cell(1, 7).value == "Char Has"
    quest_ws = wb[QUESTS_SHEET_NAME]
    assert [quest_ws.cell(1, col).value for col in range(1, 7)] == [
        "Character",
        "Expansion",
        "Zone",
        "Type",
        "Quest",
        "Status",
    ]
    assert quest_ws.auto_filter.ref is not None
    assert any(
        quest_ws.cell(row, 3).value == "Arcstone, Shattered Isles"
        and quest_ws.cell(row, 5).value == "Quench the Fire"
        for row in range(2, quest_ws.max_row + 1)
    )
    raid_ws = wb[RAID_ACHIEVEMENTS_SHEET_NAME]
    assert [raid_ws.cell(1, col).value for col in range(1, 7)] == [
        "Character",
        "Expansion",
        "Raid",
        "Event",
        "Objective",
        "Status",
    ]
    assert any(
        raid_ws.cell(row, 4).value == "Echo of Hate"
        and raid_ws.cell(row, 5).value == "Give in to Greed"
        for row in range(2, raid_ws.max_row + 1)
    )
    assert raid_ws.cell(2, 2).value == EVERQUEST_BASE_LABEL or " (20" in str(raid_ws.cell(2, 2).value)
    assert raid_ws.auto_filter.ref is not None
    assert any(
        missing_ws.cell(row, 5).value == "Strange Black Rock"
        for row in range(2, missing_ws.max_row + 1)
    )


def test_parse_quest_parent_and_clean_quest_name() -> None:
    assert parse_quest_parent("Mercenary of Arcstone, Shattered Isles") == (
        "Mercenary",
        "Arcstone, Shattered Isles",
    )
    assert parse_quest_parent("Partisan of Scarred Grove") == ("Partisan", "Scarred Grove")
    assert parse_quest_parent("Hero of Candlemaker's Workshop: Waxwork Abolishion") is None
    assert (
        clean_quest_name(
            "Quench the Fire - from Archivist Kavros in Arcstone, Shattered Isles"
        )
        == "Quench the Fire"
    )
    assert (
        clean_quest_name("Sergeant Malachi in Hodstock Hills - Scaled Invaders")
        == "Scaled Invaders"
    )
    assert clean_quest_name("Garvin Windrunner - Reduce the Risk") == "Reduce the Risk"


def test_parse_mercenary_partisan_quest_children() -> None:
    parsed = parse_achievements_file(SHAMLUB_ACH)
    arcstone = [
        item
        for item in parsed.quest_achievements
        if item.section == "Shattering of Ro"
        and item.quest_type == "Mercenary"
        and item.zone == "Arcstone, Shattered Isles"
    ]
    assert {item.quest for item in arcstone} == {
        "Quench the Fire",
        "Fungal Outbreak",
        "Cleanse the Corruption",
    }
    assert all(not item.complete for item in arcstone)

    scarred_partisan = [
        item
        for item in parsed.quest_achievements
        if item.zone == "Scarred Grove" and item.quest_type == "Partisan"
    ]
    assert any(item.quest == "Missing Scout" and item.complete for item in scarred_partisan)
    assert any(item.quest == "Storm Chasing" and not item.complete for item in scarred_partisan)

    labyrinth_merc = [
        item
        for item in parsed.quest_achievements
        if item.zone == "Labyrinth of Spite" and item.quest_type == "Mercenary"
    ]
    assert labyrinth_merc
    assert all(item.complete for item in labyrinth_merc)

    assert not any(
        item.quest.casefold().startswith("complete either")
        for item in parsed.quest_achievements
    )
    assert not any(
        item.quest.casefold().startswith("complete the achievement")
        for item in parsed.quest_achievements
    )
    assert not any("Waxwork Abolishion" in item.quest for item in parsed.quest_achievements)
    assert not any(
        item.quest == "Mercenary of The Plane of War" for item in parsed.quest_achievements
    )


def test_quest_report_omits_complete_lines_and_keeps_partial() -> None:
    inv = EXAMPLES / "Shamlub_bristle-Inventory.txt"
    report = build_team_report([inv])
    ach_report = build_achievement_report(
        report,
        achievement_paths={"shamlub_bristle": SHAMLUB_ACH},
    )
    assert ach_report is not None
    quests = ach_report.quests

    arcstone = [
        row
        for row in quests
        if row.zone == "Arcstone, Shattered Isles" and row.quest_type == "Mercenary"
    ]
    assert {row.quest for row in arcstone} == {
        "Quench the Fire",
        "Fungal Outbreak",
        "Cleanse the Corruption",
    }
    assert all(row.status == "Missing" for row in arcstone)

    scarred_partisan = [
        row for row in quests if row.zone == "Scarred Grove" and row.quest_type == "Partisan"
    ]
    statuses = {row.quest: row.status for row in scarred_partisan}
    assert statuses["Missing Scout"] == "Done"
    assert statuses["Storm Chasing"] == "Missing"

    scarred_merc = [
        row for row in quests if row.zone == "Scarred Grove" and row.quest_type == "Mercenary"
    ]
    assert any(row.quest == "Smash to Mulch!" and row.status == "Missing" for row in scarred_merc)
    assert any(
        row.quest == "Construct Deconstruction" and row.status == "Done"
        for row in scarred_merc
    )
    assert not any("Complete either" in row.quest for row in scarred_merc)

    assert not any(
        row.zone == "Labyrinth of Spite" and row.quest_type == "Mercenary" for row in quests
    )
    assert any(
        row.zone == "Labyrinth of Spite" and row.quest_type == "Partisan" and row.status == "Missing"
        for row in quests
    )
    assert not any(row.zone == "Hodstock Hills" for row in quests)
