from pathlib import Path

from inventory_parser.export_bundle import build_export_bundle
from inventory_parser.html_export import serialize_report
from inventory_parser.excel_export import write_team_workbook
from openpyxl import load_workbook

EXAMPLES = Path(__file__).resolve().parents[1] / "Examples"
FIXTURE = Path(__file__).resolve().parent / "fixtures" / "raidloot_dex_sample.html"


def test_include_slot2_false_omits_section_and_sheets(tmp_path: Path) -> None:
    inv = EXAMPLES / "Deflub_bristle-Inventory.txt"
    bundle = build_export_bundle(
        [inv],
        include_spells=False,
        include_achievements=False,
        include_slot2=False,
    )
    assert bundle.slot2 is None
    payload = serialize_report(bundle)
    assert all(s["type"] != "slot2_augs" for s in payload["sections"])
    out = tmp_path / "no_slot2.xlsx"
    write_team_workbook(bundle.team, out, slot2=None)
    names = load_workbook(out).sheetnames
    assert "Augs" not in names
    assert "Need to Farm" not in names
    assert "Stat Summary" not in names


def test_include_slot2_offline_adds_section(tmp_path: Path) -> None:
    inv = EXAMPLES / "Deflub_bristle-Inventory.txt"
    html = FIXTURE.read_text(encoding="utf-8")
    bundle = build_export_bundle(
        [inv],
        include_spells=False,
        include_achievements=False,
        include_slot2=True,
        catalog_html=html,
        fetch_eqr_augs=False,
        fetch_chest_class=False,
        fetch_expansions=False,
        type78_slot_by_parent_id={},
    )
    assert bundle.slot2 is not None
    payload = serialize_report(bundle)
    section = next(s for s in payload["sections"] if s["type"] == "slot2_augs")
    assert section["id"] == "slot2_augs"
    assert section["title"] == "Type 7/8 Augs"
    assert "upgrades" in section["data"]
    out = tmp_path / "with_slot2.xlsx"
    write_team_workbook(bundle.team, out, slot2=bundle.slot2)
    names = load_workbook(out).sheetnames
    assert "Augs" in names
    assert "Need to Farm" in names
    assert "Stat Summary" in names
    assert "Ranked Augs" in names
    assert "Aug Legend" in names
