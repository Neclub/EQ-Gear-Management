from pathlib import Path

from openpyxl import load_workbook

from inventory_parser.team_report import build_team_report
from inventory_parser.excel_export import UNMADE_GEAR_SHEET_NAME, write_team_workbook
from inventory_parser.gear_tiers import tier_rank
from inventory_parser.unmade_gear import (
    build_unmade_gear_report,
    is_bag_location,
    is_tier_below_target,
    parse_unmade_material,
)

EXAMPLES = Path(__file__).resolve().parents[1] / "Examples"


def test_is_bag_location() -> None:
    assert is_bag_location("General 1-Slot25")
    assert is_bag_location("General 12")
    assert not is_bag_location("Bank9-Slot12")
    assert not is_bag_location("SharedBank6-Slot29")
    assert not is_bag_location("Head")


def test_parse_fractured_sor_mat() -> None:
    mat = parse_unmade_material("Fractured Mask Fastener")
    assert mat is not None
    assert mat.expansion == "SoR"
    assert mat.material == "T2"
    assert mat.target_tier == "SOR-R2"
    assert mat.target_slot == "Face"


def test_parse_fractured_idol_polishing_cloth_targets_range() -> None:
    mat = parse_unmade_material("Fractured Idol Polishing Cloth")
    assert mat is not None
    assert mat.target_slot == "Range"


def test_parse_fractured_charm_polishing_cloth_targets_charm() -> None:
    mat = parse_unmade_material("Fractured Charm Polishing Cloth")
    assert mat is not None
    assert mat.target_slot == "Charm"


def test_parse_fractured_ring_polishing_cloth_targets_fingers() -> None:
    mat = parse_unmade_material("Fractured Ring Polishing Cloth")
    assert mat is not None
    assert mat.expansion == "SoR"
    assert mat.material == "T2"
    assert mat.target_tier == "SOR-R2"
    assert mat.target_slot == "Fingers-1"


def test_parse_rejects_ore() -> None:
    assert parse_unmade_material("Riven Arcana Ore") is None


def test_parse_rebellion_tradeskill_mat() -> None:
    mat = parse_unmade_material("Necklace Clasp of Rebellion")
    assert mat is not None
    assert mat.expansion == "ToB"
    assert mat.material == "T2"
    assert mat.target_tier == "TOB-R2"
    assert mat.target_slot == "Neck"


def test_parse_rejects_finished_rebellion_gear() -> None:
    assert parse_unmade_material("Defender's Charm of Rebellion") is None
    assert parse_unmade_material("Legionnaire Helm of Rebellion") is None


def test_parse_diminished_shattered_container() -> None:
    mat = parse_unmade_material("Diminished Shattered Dominion Head Armor")
    assert mat is not None
    assert mat.expansion == "SoR"
    assert mat.material == "T1"
    assert mat.target_tier == "SOR-R1"
    assert mat.target_slot == "Head"


def test_parse_obscured_bound_container() -> None:
    mat = parse_unmade_material("Obscured Wrist Armor of the Bound")
    assert mat is not None
    assert mat.expansion == "ToB"
    assert mat.material == "T1"
    assert mat.target_tier == "TOB-R1"
    assert mat.target_slot == "Wrist-1"


def test_parse_rejects_finished_bound_gear() -> None:
    assert parse_unmade_material("Legionnaire Helm of the Bound") is None
    assert parse_unmade_material("Loremaster Bracer of the Bound") is None


def test_tier_rank_newest_is_highest() -> None:
    assert tier_rank("SOR-R2") > tier_rank("SOR-R1")
    assert tier_rank("TOB-R2") > tier_rank("TOB-R1")


def test_is_tier_below_target() -> None:
    assert is_tier_below_target(None, "SOR-R2")
    assert is_tier_below_target("???", "TOB-R2")
    assert is_tier_below_target("TOB-R1", "TOB-R2")
    assert not is_tier_below_target("TOB-R2", "TOB-R2")
    assert not is_tier_below_target("SOR-R2", "SOR-R2")
    assert not is_tier_below_target("Evolver", "SOR-R2", is_evolver=True)


def test_rebellion_mat_listed_when_neck_empty(tmp_path: Path) -> None:
    inv = tmp_path / "Test_bristle-Inventory.txt"
    inv.write_text(
        "Location\tName\tID\tCount\tSlots\n"
        "General 1-Slot1\tNecklace Clasp of Rebellion\t170443\t1\t6\n",
        encoding="utf-8",
    )
    report = build_team_report([inv])
    entries = build_unmade_gear_report(report)
    assert [row.item_name for row in entries] == ["Necklace Clasp of Rebellion"]


def test_ring_mat_listed_when_fingers1_evolver_and_fingers2_below(tmp_path: Path) -> None:
    inv = tmp_path / "Songlub_xegony-Inventory.txt"
    inv.write_text(
        "Location\tName\tID\tCount\tSlots\n"
        "Fingers\tHarbinger's Fine Harasser Ring of Malice\t168066\t1\t6\n"
        "Fingers-Slot6\tDevotee's Enhancement of Enduring Harmony\t138895\t1\t6\n"
        "Fingers\tRing of Roaring Skies\t175732\t1\t6\n"
        "General 1-Slot43\tFractured Ring Polishing Cloth\t170807\t1\t6\n",
        encoding="utf-8",
    )
    report = build_team_report([inv])
    entries = build_unmade_gear_report(report)
    assert [row.item_name for row in entries] == ["Fractured Ring Polishing Cloth"]
    assert entries[0].target_slot == "Fingers-2"
    assert entries[0].equipped_tier == "SOR-R1"


def test_ring_mat_listed_when_both_fingers_at_sor_r2(tmp_path: Path) -> None:
    inv = tmp_path / "Test_bristle-Inventory.txt"
    inv.write_text(
        "Location\tName\tID\tCount\tSlots\n"
        "Fingers\tRing of Resonant Fracture\t175900\t1\t6\n"
        "Fingers\tBand of Resonant Fracture\t175901\t1\t6\n"
        "General 1-Slot1\tFractured Ring Polishing Cloth\t170807\t1\t6\n",
        encoding="utf-8",
    )
    report = build_team_report([inv])
    entries = build_unmade_gear_report(report)
    assert [row.item_name for row in entries] == ["Fractured Ring Polishing Cloth"]
    assert entries[0].target_slot == "Fingers-1"
    assert entries[0].equipped_tier == "SOR-R2"


def test_rebellion_mat_listed_when_face_at_tob_r2(tmp_path: Path) -> None:
    inv = tmp_path / "Test_bristle-Inventory.txt"
    inv.write_text(
        "Location\tName\tID\tCount\tSlots\n"
        "Face\tAcrobat's Mask of Rebellion\t173919\t1\t6\n"
        "General 1-Slot1\tMask Fastener of Rebellion\t170442\t1\t6\n",
        encoding="utf-8",
    )
    report = build_team_report([inv])
    entries = build_unmade_gear_report(report)
    assert [row.item_name for row in entries] == ["Mask Fastener of Rebellion"]
    assert entries[0].target_slot == "Face"
    assert entries[0].equipped_tier == "TOB-R2"


def test_songlub_lists_diminished_head_and_weapon_core() -> None:
    report = build_team_report([EXAMPLES / "Songlub_bristle-Inventory.txt"])
    entries = build_unmade_gear_report(report)
    names = {row.item_name for row in entries}
    assert "Diminished Shattered Dominion Head Armor" in names
    assert "Finesse Core of Rebellion" in names


def test_deflub_excludes_ore_and_finished_bound_gear() -> None:
    report = build_team_report([EXAMPLES / "Deflub_bristle-Inventory.txt"])
    entries = build_unmade_gear_report(report)
    names = {row.item_name for row in entries}
    assert "Riven Arcana Ore" not in names
    assert "Legionnaire Helm of the Bound" not in names
    assert "Maul of Rebellion" not in names


def test_healub_lists_obscured_bound_container() -> None:
    report = build_team_report([EXAMPLES / "Healub_bristle-Inventory.txt"])
    entries = build_unmade_gear_report(report)
    names = {row.item_name for row in entries}
    assert "Obscured Wrist Armor of the Bound" in names


def test_deflub_lists_fractured_string_serving() -> None:
    report = build_team_report([EXAMPLES / "Deflub_bristle-Inventory.txt"])
    entries = build_unmade_gear_report(report)
    names = {row.item_name for row in entries}
    assert "Fractured String Serving" in names


def test_excel_export_includes_unmade_gear_tab(tmp_path: Path) -> None:
    paths = [
        EXAMPLES / "Songlub_bristle-Inventory.txt",
        EXAMPLES / "Deflub_bristle-Inventory.txt",
    ]
    report = build_team_report(paths)
    out = tmp_path / "unmade.xlsx"
    write_team_workbook(report, out)

    wb = load_workbook(out, data_only=True)
    assert UNMADE_GEAR_SHEET_NAME in wb.sheetnames
    ws = wb[UNMADE_GEAR_SHEET_NAME]
    assert ws.cell(1, 1).value == "Character"
    assert ws.cell(1, 2).value == "Item"
    item_names = {
        ws.cell(row, 2).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 2).value
    }
    assert "Diminished Shattered Dominion Head Armor" in item_names
    assert "Fractured String Serving" in item_names
    assert "Legionnaire Helm of the Bound" not in item_names
