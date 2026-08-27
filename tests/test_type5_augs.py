"""Tests for Type 5 aug display export."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from inventory_parser.excel_export import write_team_workbook
from inventory_parser.export_bundle import build_export_bundle
from inventory_parser.html_export import serialize_report
from inventory_parser.type5_augs.build import TYPE5_CATALOG_URL
from inventory_parser.type5_augs.excel import SHEET_NAME
from inventory_parser.type5_augs.html import serialize_type5_section
from inventory_parser.type5_augs.vanquisher import (
    VANQUISHER_AUGS,
    lookup_vanquisher_aug,
    vanquisher_label,
)

EXAMPLES = Path(__file__).resolve().parents[1] / "Examples"
FIXTURES = Path(__file__).resolve().parent / "fixtures"


def test_include_type5_false_omits_section_and_sheet(tmp_path: Path) -> None:
    inv = EXAMPLES / "Deflub_bristle-Inventory.txt"
    bundle = build_export_bundle(
        [inv],
        include_spells=False,
        include_achievements=False,
        include_slot2=False,
        include_type5=False,
    )
    assert bundle.type5 is None
    payload = serialize_report(bundle)
    assert all(s["type"] != "type5_augs" for s in payload["sections"])
    out = tmp_path / "no_type5.xlsx"
    write_team_workbook(bundle.team, out, type5=None)
    assert SHEET_NAME not in load_workbook(out).sheetnames


def test_include_type5_offline_adds_section(tmp_path: Path) -> None:
    inv = EXAMPLES / "Deflub_bristle-Inventory.txt"
    aug_html = (FIXTURES / "eqresource_aug_173378.html").read_text(encoding="utf-8")
    bundle = build_export_bundle(
        [inv],
        include_spells=False,
        include_achievements=False,
        include_slot2=False,
        include_type5=True,
        type5_slot_by_parent_id={},
        type5_fetch_eqr_augs=False,
        type5_eqr_aug_html_by_id={173378: aug_html},
        fetch_chest_class=False,
    )
    assert bundle.type5 is not None
    payload = serialize_report(bundle)
    section = next(s for s in payload["sections"] if s["type"] == "type5_augs")
    assert section["title"] == "Type 5 Augs"
    assert section["data"]["catalogUrl"] == TYPE5_CATALOG_URL

    data = serialize_type5_section(bundle.type5)
    assert "rows" in data
    assert data["eqResourceItemUrl"]

    out = tmp_path / "with_type5.xlsx"
    write_team_workbook(bundle.team, out, type5=bundle.type5)
    assert SHEET_NAME in load_workbook(out).sheetnames


def test_type5_html_template_has_per_character_cards() -> None:
    from inventory_parser.package_data import read_data_text

    html = read_data_text("team_report.html")
    assert "function updateType5AugsContent" in html
    assert "raidBisCharTitle(ch)" in html
    assert "type5PersonaKey" in html
    assert "empty of" in html
    assert ".type5-augs .card" in html
    assert ".type5-augs .card-head" in html
    assert "raid-bis-char-plate" in html
    assert "expansionUrl" in html
    assert "expansionTitle" in html


def test_type5_serialize_empty_and_stats() -> None:
    from inventory_parser.parser import InventoryData, InventoryItem
    from inventory_parser.team_report import CharacterGear, TeamGearReport
    from inventory_parser.type5_augs.build import build_type5_export

    data = InventoryData(
        character="Tester",
        server="bristle",
        filepath="x",
        class_abbr="WAR",
        items=[
            InventoryItem("Head", "Helm of Tests", 175860, 1, 3),
            InventoryItem("Head-Slot1", "Immovable Green Gem", 173378, 1, 0),
            InventoryItem("Ear", "Earring A", 175914, 1, 4),
            InventoryItem("Ear-Slot2", "Empty", 0, 0, 0),
        ],
    )
    char = CharacterGear(
        character="Tester",
        server="bristle",
        filepath="x",
        class_abbr="WAR",
        inventory_data=data,
    )
    team = TeamGearReport(characters=[char], warnings=[])
    aug_html = (FIXTURES / "eqresource_aug_173378.html").read_text(encoding="utf-8")
    export = build_type5_export(
        team,
        type5_slot_by_parent_id={175860: 1, 175914: 2},
        eqr_aug_html_by_id={173378: aug_html},
        fetch_eqr_augs=False,
        fetch_expansions=False,
    )
    payload = serialize_type5_section(export)
    by_slot = {r["gearSlot"]: r for r in payload["rows"]}
    assert by_slot["Head"]["name"] == "Immovable Green Gem"
    assert by_slot["Head"]["empty"] is False
    assert by_slot["Head"]["expansion"] == "The Outer Brood"
    assert by_slot["Head"]["expansionUrl"] is None
    assert by_slot["Head"]["hsta"] == 135
    assert by_slot["Head"]["hstr"] == 41
    assert by_slot["Ear-1"]["empty"] is True
    assert by_slot["Ear-1"]["name"] is None
    assert by_slot["Ear-1"]["expansion"] is None
    assert payload["catalogUrl"] == TYPE5_CATALOG_URL


def test_vanquisher_catalog_labels() -> None:
    assert len(VANQUISHER_AUGS) == 5
    expected = {
        163995: ("Vanq ToL", "Vanquisher of Terror of Luclin"),
        164196: ("Vanq NoS", "Vanquisher of Night of Shadows"),
        151234: ("Vanq LS", "Vanquisher of Laurion's Song"),
        151793: ("Vanq ToB", "Vanquisher of The Outer Brood"),
        153972: ("Vanq SoR", "Vanquisher of Shattering of Ro"),
    }
    for item_id, (label, title) in expected.items():
        hit = vanquisher_label(item_id)
        assert hit is not None
        assert hit[0] == label
        assert hit[2] == title
        assert "achievements.php?id=" in hit[1]

    by_name = vanquisher_label(None, "divine medallion")
    assert by_name is not None
    assert by_name[0] == "Vanq NoS"

    assert lookup_vanquisher_aug(999999, "Immovable Green Gem") is None
    assert vanquisher_label(173378, "Immovable Green Gem") is None


def test_type5_vanquisher_expansion_override(tmp_path: Path) -> None:
    from inventory_parser.parser import InventoryData, InventoryItem
    from inventory_parser.team_report import CharacterGear, TeamGearReport
    from inventory_parser.type5_augs.build import build_type5_export

    data = InventoryData(
        character="Tester",
        server="bristle",
        filepath="x",
        class_abbr="WAR",
        items=[
            InventoryItem("Head", "Helm of Tests", 175860, 1, 3),
            InventoryItem("Head-Slot1", "Divine Medallion", 164196, 1, 0),
            InventoryItem("Ear", "Earring A", 175914, 1, 4),
            InventoryItem("Ear-Slot2", "Immovable Green Gem", 173378, 1, 0),
        ],
    )
    char = CharacterGear(
        character="Tester",
        server="bristle",
        filepath="x",
        class_abbr="WAR",
        inventory_data=data,
    )
    team = TeamGearReport(characters=[char], warnings=[])
    aug_html = (FIXTURES / "eqresource_aug_173378.html").read_text(encoding="utf-8")
    export = build_type5_export(
        team,
        type5_slot_by_parent_id={175860: 1, 175914: 2},
        eqr_aug_html_by_id={173378: aug_html},
        fetch_eqr_augs=False,
        fetch_expansions=False,
    )
    payload = serialize_type5_section(export)
    by_slot = {r["gearSlot"]: r for r in payload["rows"]}

    vanq = by_slot["Head"]
    assert vanq["name"] == "Divine Medallion"
    assert vanq["expansion"] == "Vanq NoS"
    assert vanq["expansionTitle"] == "Vanquisher of Night of Shadows"
    assert vanq["expansionUrl"] == (
        "https://achievements.eqresource.com/achievements.php?id=3001009"
    )

    other = by_slot["Ear-1"]
    assert other["name"] == "Immovable Green Gem"
    assert other["expansion"] == "The Outer Brood"
    assert other["expansionUrl"] is None

    out = tmp_path / "vanq_type5.xlsx"
    write_team_workbook(team, out, type5=export)
    ws = load_workbook(out)[SHEET_NAME]
    rows = {ws.cell(r, 2).value: r for r in range(2, ws.max_row + 1)}
    head_row = rows["Head"]
    assert ws.cell(head_row, 3).value == "Vanq NoS"
    assert "achievements.php?id=3001009" in str(ws.cell(head_row, 3).hyperlink.target)
    ear_row = rows["Ear-1"]
    assert ws.cell(ear_row, 3).value == "The Outer Brood"
    assert ws.cell(ear_row, 3).hyperlink is None
