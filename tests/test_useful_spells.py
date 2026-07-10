"""Tests for curated useful-spell intersection with MissingSpells dumps."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from inventory_parser.excel_export import (
    GEAR_T_LEVEL_SHEET_NAME,
    MISSING_SPELLS_SHEET_NAME,
    MISSING_USEFUL_SPELLS_SHEET_NAME,
    RUNE_INVENTORY_SHEET_NAME,
    write_team_workbook,
)
from inventory_parser.export_bundle import build_export_bundle
from inventory_parser.html_export import serialize_report
from inventory_parser.team_report import build_team_report
from inventory_parser.useful_spells import (
    UsefulSpell,
    build_missing_useful_spells_report,
    load_useful_spells,
    useful_matches_missing,
)

EXAMPLES = Path(__file__).resolve().parents[1] / "Examples"


def test_load_useful_spells_has_shd_not_shk() -> None:
    catalog = load_useful_spells()
    assert "SHD" in catalog
    assert "SHK" not in catalog
    assert len(catalog["PAL"]) >= 40
    assert any(s.name == "Brilliant Expurgation" for s in catalog["PAL"])


def test_useful_matches_rk_iii_and_unranked() -> None:
    ranked = UsefulSpell(name="Brilliant Expurgation", level=130, highest_rk="III")
    assert useful_matches_missing(ranked, "Brilliant Expurgation Rk. III")
    assert useful_matches_missing(ranked, "Brilliant Expurgation Rk. II")
    assert not useful_matches_missing(ranked, "Some Other Spell Rk. III")

    unranked = UsefulSpell(name="Force of Revocation", level=130, highest_rk="n/a")
    assert useful_matches_missing(unranked, "Force of Revocation")


def test_useful_matches_numeric_rank() -> None:
    useful = UsefulSpell(name="Dichotomic Fury", level=101, highest_rk="6")
    assert useful_matches_missing(useful, "Dichotomic Fury VI")
    assert useful_matches_missing(useful, "Dichotomic Fury 6")
    assert not useful_matches_missing(useful, "Dichotomic Something Else VI")


def test_build_missing_useful_for_deflub_pal() -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    team = build_team_report(paths)
    report = build_missing_useful_spells_report(team, inventory_paths=paths)
    assert report is not None
    assert report.entries
    assert all(e.display_name == "Deflub ( PAL )" for e in report.entries)
    names = {e.spell_name for e in report.entries}
    assert any("Brilliant Expurgation" in n for n in names)
    # Lower-level useful spells from the dump are included (not 121–130 only)
    assert any(e.level < 121 for e in report.entries)


def test_excel_missing_useful_tab(tmp_path: Path) -> None:
    paths = sorted(EXAMPLES.glob("*-Inventory.txt"))
    bundle = build_export_bundle(paths)
    assert bundle.missing_useful_report is not None
    assert bundle.missing_useful_report.entries

    out = tmp_path / "useful.xlsx"
    write_team_workbook(
        bundle.team,
        out,
        spell_report=bundle.spell_report,
        missing_useful_report=bundle.missing_useful_report,
        rune_inventory_report=bundle.rune_inventory_report,
        unmade_entries=bundle.unmade_entries,
    )
    wb = load_workbook(out, data_only=True)
    assert MISSING_USEFUL_SPELLS_SHEET_NAME in wb.sheetnames
    # Sheet order: after Missing Spells, before Rune Inventory
    names = wb.sheetnames
    assert names.index(MISSING_SPELLS_SHEET_NAME) < names.index(MISSING_USEFUL_SPELLS_SHEET_NAME)
    assert names.index(MISSING_USEFUL_SPELLS_SHEET_NAME) < names.index(RUNE_INVENTORY_SHEET_NAME)

    ws = wb[MISSING_USEFUL_SPELLS_SHEET_NAME]
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Character"
    )
    assert [ws.cell(header_row, c).value for c in range(1, 7)] == [
        "Character",
        "Level",
        "Expansion",
        "Spell",
        "Highest RK",
        "Comments",
    ]
    assert ws.auto_filter.ref
    assert any(
        ws.cell(r, 1).value == "Deflub ( PAL )"
        for r in range(header_row + 1, ws.max_row + 1)
    )


def test_html_missing_useful_has_character_filter() -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    bundle = build_export_bundle(paths)
    payload = serialize_report(bundle)
    section = next(s for s in payload["sections"] if s["id"] == "missing_useful_spells")
    assert section["title"] == MISSING_USEFUL_SPELLS_SHEET_NAME
    assert section["data"]["characterColumn"] == 0
    assert section["data"]["rows"]
    assert all(row[0] == "Deflub ( PAL )" for row in section["data"]["rows"])
    credit = section["data"]["credit"]
    assert credit["text"] == 'Based on "SOR - Raccoo\'s list of useful spells"'
    assert "docs.google.com/spreadsheets" in credit["url"]


def test_excel_missing_useful_credit_hyperlink(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    bundle = build_export_bundle(paths)
    out = tmp_path / "useful_credit.xlsx"
    write_team_workbook(
        bundle.team,
        out,
        missing_useful_report=bundle.missing_useful_report,
    )
    wb = load_workbook(out, data_only=False)
    ws = wb[MISSING_USEFUL_SPELLS_SHEET_NAME]
    credit_row = next(
        r
        for r in range(1, ws.max_row + 1)
        if ws.cell(r, 1).value == 'Based on "SOR - Raccoo\'s list of useful spells"'
    )
    cell = ws.cell(credit_row, 1)
    assert cell.hyperlink is not None
    assert "docs.google.com/spreadsheets" in cell.hyperlink.target


def test_no_useful_tab_without_entries(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Stablub_bristle-Inventory.txt"]
    team = build_team_report(paths)
    out = tmp_path / "no_useful.xlsx"
    write_team_workbook(team, out)
    wb = load_workbook(out, data_only=True)
    assert MISSING_USEFUL_SPELLS_SHEET_NAME not in wb.sheetnames
    assert GEAR_T_LEVEL_SHEET_NAME in wb.sheetnames
