from pathlib import Path

from openpyxl import load_workbook

from inventory_parser.team_report import build_team_report
from inventory_parser.excel_export import (
    GEAR_T_LEVEL_SHEET_NAME,
    MISSING_SPELLS_SHEET_NAME,
    MISSING_USEFUL_SPELLS_SHEET_NAME,
    RUNE_INVENTORY_SHEET_NAME,
    UNMADE_GEAR_SHEET_NAME,
    _matrix_character_col_width,
    write_team_workbook,
)
from inventory_parser.excel_theme import TIER_COLOR_ORANGE
from inventory_parser.rune_inventory import build_rune_inventory_report
from inventory_parser.excel_theme import SHEET_BACKGROUND_COLS, SHEET_BACKGROUND_ROWS
from inventory_parser.spell_report import build_spell_rune_report
from inventory_parser.parser import extract_equipped_items, parse_inventory_file
from inventory_parser.slots import TEAM_GEAR_SLOTS, VISIBLE_SLOTS

EXAMPLES = Path(__file__).resolve().parents[1] / "Examples"
_FIRST_CHAR_COL = 3


def _read_matrix(path: Path) -> dict[str, dict[str, str | None]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Team gear"]
    chars = [ws.cell(1, c).value for c in range(_FIRST_CHAR_COL, ws.max_column + 1)]
    out: dict[str, dict[str, str | None]] = {ch: {} for ch in chars if ch}
    for r in range(2, ws.max_row + 1):
        slot = ws.cell(r, 1).value
        if not slot:
            continue
        for c, ch in enumerate(chars, start=_FIRST_CHAR_COL):
            if not ch:
                continue
            out[ch][slot] = ws.cell(r, c).value
    return out


def test_slot_rows_visible_first(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "crew.xlsx"
    write_team_workbook(report, out)

    wb = load_workbook(out, data_only=False)
    ws = wb["Team gear"]
    slots = [ws.cell(r, 1).value for r in range(2, 2 + len(TEAM_GEAR_SLOTS))]
    assert slots == list(TEAM_GEAR_SLOTS)
    assert slots[: len(VISIBLE_SLOTS)] == list(VISIBLE_SLOTS)
    assert ws.cell(2, 2).value == "Visible"
    assert ws.cell(2 + len(VISIBLE_SLOTS), 2).value == "Non-visible"


def test_deflub_arms_item_name(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "crew.xlsx"
    write_team_workbook(report, out)

    matrix = _read_matrix(out)
    assert matrix["Deflub ( PAL )"]["Arms"] == "Exarch Vambraces of Eternal Reverie"


def test_item_hyperlink_to_eqresource(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "crew.xlsx"
    write_team_workbook(report, out)

    wb = load_workbook(out, data_only=False)
    ws = wb["Team gear"]
    charm_row = 2 + list(TEAM_GEAR_SLOTS).index("Charm")
    cell = ws.cell(charm_row, _FIRST_CHAR_COL)
    assert cell.value == "Defender's Charm of Rebellion"
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == "https://items.eqresource.com/items.php?id=173940"


def test_empty_secondary_left_blank(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Monklub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "crew.xlsx"
    write_team_workbook(report, out)

    matrix = _read_matrix(out)
    assert matrix["Monklub ( MNK )"].get("Secondary") in (None, "")


def test_matches_extract_equipped(tmp_path: Path) -> None:
    path = EXAMPLES / "Stablub_bristle-Inventory.txt"
    data = parse_inventory_file(path)
    assert data is not None
    equipped, _evolver_keys = extract_equipped_items(data)

    report = build_team_report([path])
    out = tmp_path / "one.xlsx"
    write_team_workbook(report, out)

    matrix = _read_matrix(out)
    for slot in TEAM_GEAR_SLOTS:
        if slot in equipped:
            assert matrix["Stablub"].get(slot) == equipped[slot].name
            assert report.characters[0].slots[slot].item_id == equipped[slot].item_id
        else:
            assert matrix["Stablub"].get(slot) in (None, "")


def test_visible_only_export(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "visible.xlsx"
    write_team_workbook(report, out, slot_filter="visible")

    wb = load_workbook(out, data_only=True)
    ws = wb["Team gear"]
    slots = [ws.cell(r, 1).value for r in range(2, 2 + len(VISIBLE_SLOTS))]
    assert slots == list(VISIBLE_SLOTS)
    assert "Charm" not in slots


def test_includes_all_example_characters() -> None:
    paths = sorted(EXAMPLES.glob("*_bristle-Inventory.txt"))
    report = build_team_report(paths)
    names = {r.character for r in report.characters}
    assert names == {
        "Deflub",
        "Healub",
        "Magelub",
        "Monklub",
        "Shamlub",
        "Songlub",
        "Stablub",
    }


def test_gear_legend_on_a26_a30(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "crew.xlsx"
    write_team_workbook(report, out)

    wb = load_workbook(out, data_only=False)
    ws = wb["Team gear"]
    assert ws.cell(25, 1).value == "Gear tier colors"
    assert ws.cell(26, 2).value == "Green — SOR-R2 (current SoR raid)"
    assert ws.cell(30, 2).value == "Purple — Evolver"


def test_black_background_to_row_50_column_z(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "crew.xlsx"
    write_team_workbook(report, out)

    wb = load_workbook(out, data_only=False)
    ws = wb["Team gear"]
    corner = ws.cell(50, 26)
    assert corner.fill.start_color.rgb in ("00000000", "FF000000")
    assert ws.cell(25, 15).fill.start_color.rgb in ("00000000", "FF000000")


def test_no_excel_table_only_autofilter(tmp_path: Path) -> None:
    """Excel Table + worksheet autoFilter duplicates XML and triggers repair dialog."""
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "crew.xlsx"
    write_team_workbook(report, out)

    wb = load_workbook(out, data_only=False)
    ws = wb["Team gear"]
    assert list(ws.tables.keys()) == []
    assert ws.auto_filter.ref is not None


def test_rebellion_cell_has_fill(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "crew.xlsx"
    write_team_workbook(report, out)

    wb = load_workbook(out, data_only=False)
    ws = wb["Team gear"]
    charm_row = 2 + list(TEAM_GEAR_SLOTS).index("Charm")
    charm_cell = ws.cell(charm_row, _FIRST_CHAR_COL)
    assert charm_cell.value == "Defender's Charm of Rebellion"
    assert charm_cell.fill.start_color.rgb in (
        f"00{TIER_COLOR_ORANGE}",
        f"FF{TIER_COLOR_ORANGE}",
    )

    assert wb.sheetnames == ["Team gear", GEAR_T_LEVEL_SHEET_NAME, UNMADE_GEAR_SHEET_NAME]


def _sor_slot_names(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb[GEAR_T_LEVEL_SHEET_NAME]
    return [
        ws.cell(r, 1).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 1).value and r < 25
    ]


def _sor_slot_row(path: Path, slot: str) -> int:
    wb = load_workbook(path, data_only=True)
    ws = wb[GEAR_T_LEVEL_SHEET_NAME]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == slot:
            return row
    raise AssertionError(f"Slot {slot!r} not found on Gear T-Level sheet")


def test_sor_gaps_sheet_and_secondary_row(tmp_path: Path) -> None:
    report = build_team_report([EXAMPLES / "Stablub_bristle-Inventory.txt"])
    out = tmp_path / "sor.xlsx"
    write_team_workbook(report, out)
    assert GEAR_T_LEVEL_SHEET_NAME in load_workbook(out, data_only=True).sheetnames
    assert "Secondary" in _sor_slot_names(out)

    report2 = build_team_report([EXAMPLES / "Monklub_bristle-Inventory.txt"])
    out2 = tmp_path / "sor2.xlsx"
    write_team_workbook(report2, out2)
    assert "Secondary" not in _sor_slot_names(out2)


def test_sor_gaps_markers_on_deflub(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    out = tmp_path / "sor.xlsx"
    write_team_workbook(build_team_report(paths), out)

    wb = load_workbook(out, data_only=True)
    ws = wb[GEAR_T_LEVEL_SHEET_NAME]
    charm_row = _sor_slot_row(out, "Charm")
    assert ws.cell(charm_row, _FIRST_CHAR_COL).value == "TOB-R2"


def test_gear_t_level_hyperlink_and_item_name_hover(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    out = tmp_path / "sor.xlsx"
    write_team_workbook(build_team_report(paths), out)

    wb = load_workbook(out, data_only=False)
    ws = wb[GEAR_T_LEVEL_SHEET_NAME]
    charm_row = _sor_slot_row(out, "Charm")
    cell = ws.cell(charm_row, _FIRST_CHAR_COL)
    assert cell.value == "TOB-R2"
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == "https://items.eqresource.com/items.php?id=173940"
    assert cell.hyperlink.tooltip == "Defender's Charm of Rebellion"
    assert cell.comment is not None
    assert cell.comment.text == "Defender's Charm of Rebellion"


def test_gear_t_level_sor_r2_on_deflub_fingers(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    out = tmp_path / "gear_t_level.xlsx"
    write_team_workbook(build_team_report(paths), out)

    wb = load_workbook(out, data_only=True)
    ws = wb[GEAR_T_LEVEL_SHEET_NAME]
    fingers_row = _sor_slot_row(out, "Fingers-2")
    assert ws.cell(fingers_row, _FIRST_CHAR_COL).value == "SOR-R2"


def test_sor_gaps_evolver_label(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    out = tmp_path / "sor_evolver.xlsx"
    write_team_workbook(build_team_report(paths), out)

    wb = load_workbook(out, data_only=True)
    ws = wb[GEAR_T_LEVEL_SHEET_NAME]
    ear2_row = _sor_slot_row(out, "Ear-2")
    assert ws.cell(ear2_row, _FIRST_CHAR_COL).value == "Evolver"


def test_spell_tabs_all_examples(tmp_path: Path) -> None:
    paths = sorted(EXAMPLES.glob("*-Inventory.txt"))
    team = build_team_report(paths)
    spell = build_spell_rune_report(team, inventory_paths=paths)
    assert spell is not None
    rune = build_rune_inventory_report(team)
    assert rune is not None
    from inventory_parser.useful_spells import build_missing_useful_spells_report

    useful = build_missing_useful_spells_report(team, inventory_paths=paths)
    out = tmp_path / "crew_spells.xlsx"
    write_team_workbook(
        team,
        out,
        spell_report=spell,
        missing_useful_report=useful,
        rune_inventory_report=rune,
    )

    wb = load_workbook(out, data_only=True)
    expected = [
        "Team gear",
        GEAR_T_LEVEL_SHEET_NAME,
        "Missing Runes",
        MISSING_SPELLS_SHEET_NAME,
    ]
    if useful is not None and useful.entries:
        expected.append(MISSING_USEFUL_SPELLS_SHEET_NAME)
    expected.extend([RUNE_INVENTORY_SHEET_NAME, UNMADE_GEAR_SHEET_NAME])
    assert wb.sheetnames == expected
    assert wb["Missing Runes"]["A1"].value == "Missing Runes"
    ws = wb[MISSING_SPELLS_SHEET_NAME]
    detail_row = next(
        r
        for r in range(1, ws.max_row + 1)
        if ws.cell(r, 1).value == "Deflub ( PAL )" and ws.cell(r, 5).value == "Committal Rk. III"
    )
    assert ws.cell(detail_row, 2).value == 126
    assert ws.cell(detail_row, 3).value == "Minor"
    assert ws.cell(detail_row, 4).value == "Shattering of Ro (2025)"


def test_spell_list_black_background_through_z(tmp_path: Path) -> None:
    paths = sorted(EXAMPLES.glob("*-Inventory.txt"))
    team = build_team_report(paths)
    spell = build_spell_rune_report(team, inventory_paths=paths)
    assert spell is not None
    out = tmp_path / "spell_bg.xlsx"
    write_team_workbook(team, out, spell_report=spell)

    wb = load_workbook(out, data_only=False)
    ws = wb[MISSING_SPELLS_SHEET_NAME]
    last_row = max(
        r
        for r in range(1, ws.max_row + 1)
        if any(ws.cell(r, c).value for c in range(1, 6))
    )
    assert last_row > SHEET_BACKGROUND_ROWS

    black = ("00000000", "FF000000")
    assert ws.cell(last_row, SHEET_BACKGROUND_COLS).fill.start_color.rgb in black
    assert ws.cell(last_row, 5).fill.start_color.rgb in black
    assert ws.cell(1, SHEET_BACKGROUND_COLS).fill.start_color.rgb in black


def test_character_class_in_column_headers(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "class_headers.xlsx"
    write_team_workbook(report, out)

    wb = load_workbook(out, data_only=True)
    ws = wb["Team gear"]
    assert ws.cell(1, _FIRST_CHAR_COL).value == "Deflub ( PAL )"


def test_character_without_spell_file_has_plain_name(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Stablub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "no_class.xlsx"
    write_team_workbook(report, out)

    wb = load_workbook(out, data_only=True)
    ws = wb["Team gear"]
    assert ws.cell(1, _FIRST_CHAR_COL).value == "Stablub"


def test_no_spell_tabs_without_data(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    rune = build_rune_inventory_report(report)
    assert rune is not None
    out = tmp_path / "gear_only.xlsx"
    write_team_workbook(report, out, spell_report=None, rune_inventory_report=rune)
    names = load_workbook(out).sheetnames
    assert "Missing Runes" not in names
    assert MISSING_SPELLS_SHEET_NAME not in names
    assert RUNE_INVENTORY_SHEET_NAME in names


def test_rune_inventory_sheet_counts(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    rune_report = build_rune_inventory_report(report)
    assert rune_report is not None
    out = tmp_path / "runes.xlsx"
    write_team_workbook(report, out, rune_inventory_report=rune_report)

    wb = load_workbook(out, data_only=True)
    assert RUNE_INVENTORY_SHEET_NAME in wb.sheetnames
    ws = wb[RUNE_INVENTORY_SHEET_NAME]
    assert ws["A1"].value == "Rune Inventory"
    assert ws["A4"].value == "NoS"
    assert ws["A5"].value == "NoS · {Tier} Symbol of Shar Vahl"
    display_name = report.characters[0].display_name
    assert ws.column_dimensions["B"].width == _matrix_character_col_width(display_name)
    assert ws.column_dimensions["B"].width > 11.0


def test_rune_inventory_sheet_omitted_without_runes(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Stablub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    assert build_rune_inventory_report(report) is None
    out = tmp_path / "no_runes.xlsx"
    write_team_workbook(report, out)
    assert RUNE_INVENTORY_SHEET_NAME not in load_workbook(out).sheetnames


def test_excel_not_purchased_next_to_rk3_name(tmp_path: Path) -> None:
    from inventory_parser.missing_spells import persona_key
    from inventory_parser.team_report import CharacterGear, TeamGearReport

    spell_file = tmp_path / "Zenbane_bristle-CLR-MissingSpells.txt"
    spell_file.write_text("126\tAppeasement\n126\tWord of Wellbeing Rk. II\n", encoding="utf-8")
    char = CharacterGear(
        character="Zenbane",
        server="bristle",
        filepath="",
        class_abbr="CLR",
    )
    team = TeamGearReport(characters=[char], spell_characters=[char])
    spell = build_spell_rune_report(team, spell_paths={persona_key("Zenbane", "bristle", "CLR"): spell_file})
    assert spell is not None
    out = tmp_path / "zenbane.xlsx"
    write_team_workbook(team, out, spell_report=spell)
    ws = load_workbook(out, data_only=True)[MISSING_SPELLS_SHEET_NAME]
    names = {ws.cell(r, 5).value for r in range(1, ws.max_row + 1)}
    assert "Appeasement Rk. III  Not Purchased" in names
    assert "Word of Wellbeing Rk. III" in names
    wb = load_workbook(out, data_only=False)
    ws = wb[MISSING_SPELLS_SHEET_NAME]
    by_name = {ws.cell(r, 5).value: ws.cell(r, 5) for r in range(1, ws.max_row + 1)}
    appease = by_name["Appeasement Rk. III  Not Purchased"]
    assert appease.hyperlink is not None
    assert appease.hyperlink.target == "https://spells.eqresource.com/spells.php?id=71168"
