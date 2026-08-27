"""Tests for Gear T-Level semantic tier color buckets."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from inventory_parser.evolver import EVOLVER_GAP_LABEL
from inventory_parser.excel_export import GEAR_T_LEVEL_SHEET_NAME, write_team_workbook
from inventory_parser.excel_theme import (
    GEAR_SET_FILLS,
    TIER_COLOR_GREEN,
    TIER_COLOR_ORANGE,
    TIER_COLOR_RED,
    TIER_COLOR_YELLOW,
    build_tier_code_colors,
    spell_tier_fill,
    tier_code_fill,
    tier_code_fill_color,
)
from inventory_parser.export_bundle import build_export_bundle
from inventory_parser.gear_tiers import UNKNOWN_TIER_LABEL
from inventory_parser.html_export import extract_report_json, write_team_html
from inventory_parser.team_report import build_team_report

EXAMPLES = Path(__file__).resolve().parent.parent / "Examples"
_FIRST_CHAR_COL = 3


def _rgb_hex(cell) -> str:
    rgb = cell.fill.start_color.rgb
    if rgb is None:
        return ""
    return rgb[-6:].upper()


def test_tier_code_fill_color_rules() -> None:
    assert tier_code_fill_color("SOR-R2") == TIER_COLOR_GREEN
    assert tier_code_fill_color("SOR-R1") == TIER_COLOR_YELLOW
    assert tier_code_fill_color("ANI27") == TIER_COLOR_YELLOW
    assert tier_code_fill_color("TOB-R2") == TIER_COLOR_ORANGE
    assert tier_code_fill_color("TOB-G1") == TIER_COLOR_ORANGE
    assert tier_code_fill_color("LS-R2") == TIER_COLOR_RED
    assert tier_code_fill_color("NoS-R1") == TIER_COLOR_RED
    assert tier_code_fill_color("SOR-G2") == TIER_COLOR_RED
    assert tier_code_fill_color(UNKNOWN_TIER_LABEL) == TIER_COLOR_RED


def test_pattern_fill_helpers_return_cached_singletons() -> None:
    assert tier_code_fill("SOR-R2") is tier_code_fill("SOR-R2")
    assert tier_code_fill("TOB-R1") is tier_code_fill("TOB-G3")
    assert spell_tier_fill("Minor") is spell_tier_fill("Minor")
    assert spell_tier_fill("UnknownTier") is spell_tier_fill("AlsoUnknown")


def test_build_tier_code_colors_includes_evolver() -> None:
    colors = build_tier_code_colors()
    assert colors["SOR-R2"] == TIER_COLOR_GREEN
    assert colors[EVOLVER_GAP_LABEL] == GEAR_SET_FILLS["evolver"]
    assert colors[UNKNOWN_TIER_LABEL] == TIER_COLOR_RED


def test_excel_gear_t_level_tier_bucket_colors(tmp_path: Path) -> None:
    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    out = tmp_path / "tiers.xlsx"
    write_team_workbook(report, out)

    ws = load_workbook(out, data_only=False)[GEAR_T_LEVEL_SHEET_NAME]
    sor_r2_row = next(
        r for r in range(2, ws.max_row + 1) if ws.cell(r, _FIRST_CHAR_COL).value == "SOR-R2"
    )
    assert _rgb_hex(ws.cell(sor_r2_row, _FIRST_CHAR_COL)) == TIER_COLOR_GREEN.upper()

    tob_row = next(
        r for r in range(2, ws.max_row + 1) if ws.cell(r, _FIRST_CHAR_COL).value == "TOB-R2"
    )
    assert _rgb_hex(ws.cell(tob_row, _FIRST_CHAR_COL)) == TIER_COLOR_ORANGE.upper()


def test_html_theme_tier_codes_match_buckets(tmp_path: Path) -> None:
    inv = EXAMPLES / "Deflub_bristle-Inventory.txt"
    bundle = build_export_bundle([inv], include_spells=False, include_achievements=False, include_slot2=False)
    out = tmp_path / "tiers.html"
    write_team_html(bundle, out)
    report = extract_report_json(out.read_text(encoding="utf-8"))
    tier_codes = report["theme"]["tierCodes"]
    assert tier_codes["SOR-R2"] == TIER_COLOR_GREEN
    assert tier_codes["TOB-R2"] == TIER_COLOR_ORANGE
    assert tier_codes["LS-R2"] == TIER_COLOR_RED
    assert tier_codes[EVOLVER_GAP_LABEL] == GEAR_SET_FILLS["evolver"]
    gear = next(s for s in report["sections"] if s["id"] == "team_gear")
    charm_cell = next(
        cell
        for row in gear["data"]["rows"]
        if row["slot"] == "Charm"
        for cell in row["cells"]
        if cell and "Rebellion" in cell.get("name", "")
    )
    assert charm_cell["tierCode"] == "TOB-R2"


def test_custom_tier_colors_apply_to_excel_and_html(tmp_path: Path) -> None:
    from inventory_parser.character_column_order import save_tier_color

    custom_green = "112233"
    custom_evolver = "AABBCC"
    save_tier_color("green", custom_green)
    save_tier_color("evolver", custom_evolver)

    assert tier_code_fill_color("SOR-R2") == custom_green
    colors = build_tier_code_colors()
    assert colors["SOR-R2"] == custom_green
    assert colors[EVOLVER_GAP_LABEL] == custom_evolver

    paths = [EXAMPLES / "Deflub_bristle-Inventory.txt"]
    report = build_team_report(paths)
    xlsx = tmp_path / "custom_tiers.xlsx"
    write_team_workbook(report, xlsx)
    ws = load_workbook(xlsx, data_only=False)[GEAR_T_LEVEL_SHEET_NAME]
    sor_r2_row = next(
        r for r in range(2, ws.max_row + 1) if ws.cell(r, _FIRST_CHAR_COL).value == "SOR-R2"
    )
    assert _rgb_hex(ws.cell(sor_r2_row, _FIRST_CHAR_COL)) == custom_green

    bundle = build_export_bundle(
        paths, include_spells=False, include_achievements=False, include_slot2=False
    )
    html = tmp_path / "custom_tiers.html"
    write_team_html(bundle, html)
    payload = extract_report_json(html.read_text(encoding="utf-8"))
    assert payload["theme"]["tierCodes"]["SOR-R2"] == custom_green
    assert payload["theme"]["tierCodes"][EVOLVER_GAP_LABEL] == custom_evolver
    legend = {entry["key"]: entry["color"] for entry in payload["gearLegend"]}
    assert legend["green"] == custom_green
    assert legend["evolver"] == custom_evolver
