"""Tests for Type 18/19 aug catalog export."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from inventory_parser.excel_export import write_team_workbook
from inventory_parser.export_bundle import build_export_bundle
from inventory_parser.html_export import serialize_report
from inventory_parser.type18_augs.build import build_type18_export
from inventory_parser.type18_augs.catalog import (
    TYPE18_CATALOG_URL,
    _extra_stats_from_item_html,
    _is_ignored_non_aug,
    fetch_type18_catalog,
    parse_item_lore,
    parse_total_results,
    type19_search_payload,
)
from inventory_parser.type18_augs.categories import (
    category_from_name,
    classify_aug_type,
    is_anniversary_aug,
    stats_rank_key,
    type_label,
)
from inventory_parser.type18_augs.excel import CATALOG_SHEET_NAME, SHEET_NAME
from inventory_parser.type18_augs.html import serialize_type18_section
from inventory_parser.type18_augs.suggestions import (
    build_class_suggestions,
    hint_aug_type,
    load_cheat_sheet,
    name_series,
)
from inventory_parser.slot2_augs.raidloot import parse_aug_slot_types

EXAMPLES = Path(__file__).resolve().parents[1] / "Examples"
FIXTURES = Path(__file__).resolve().parent / "fixtures"

_ITEM_HTML = {
    169780: (FIXTURES / "eqresource_aug_type18_attacker.html").read_text(encoding="utf-8"),
    109559: (FIXTURES / "eqresource_aug_type18_selenelion.html").read_text(encoding="utf-8"),
    169781: (FIXTURES / "eqresource_aug_type18_assaulting.html").read_text(encoding="utf-8"),
    169791: (FIXTURES / "eqresource_aug_type19_attacker.html").read_text(encoding="utf-8"),
    138886: (FIXTURES / "eqresource_aug_type19_enduring.html").read_text(encoding="utf-8"),
}


def _offline_catalog_kwargs() -> dict:
    return {
        "type18_html_by_page": {
            1: (FIXTURES / "eqresource_type18_search_snip.html").read_text(encoding="utf-8"),
        },
        "type19_html_overrides": [
            (FIXTURES / "eqresource_type19_search_snip.html").read_text(encoding="utf-8"),
        ],
        "type18_item_html_by_id": _ITEM_HTML,
        "type18_allow_network": False,
    }


def test_classify_aug_type_18_and_19():
    # Dual-slot "18, 19" counts as Type 18; "19" only as Type 19.
    assert classify_aug_type(frozenset({18, 19})) == 18
    assert classify_aug_type(frozenset({19})) == 19
    assert classify_aug_type(frozenset({18})) == 18
    assert classify_aug_type(frozenset({7, 8})) is None
    assert classify_aug_type(frozenset()) is None


def test_type_label_dual_and_raid_only():
    assert type_label(frozenset({18, 19})) == "18/19"
    assert type_label(frozenset({19})) == "19"
    assert type_label(frozenset({18})) == "18"
    assert type_label(frozenset()) == ""


def test_category_from_name_longest_match():
    assert category_from_name("Acolyte's Attacker of the Harbinger") == "Attacker"
    assert category_from_name("Acolyte's Dorsal Defense of Legacies Lost") == "Dorsal Defense"
    assert category_from_name("Acolyte's Assaulting Fire of the Harbinger") == "Assaulting"
    assert category_from_name("Acolyte's Assault of the Harbinger") == "Assault"
    assert category_from_name("Secret Dawn Qua") == "Qua"
    assert category_from_name("Mystery Gem") == "Other"


def test_anniversary_markers():
    assert is_anniversary_aug("Acolyte's Attacker of the Selenelion")
    assert is_anniversary_aug("Silver Attacker of Jubilation")
    assert is_anniversary_aug("Devotee's Attacker of Enduring Harmony")
    assert not is_anniversary_aug("Acolyte's Attacker of the Harbinger")
    assert not is_anniversary_aug(None)


def test_parse_item_lore_and_slot_types():
    html18 = _ITEM_HTML[169780]
    html19 = _ITEM_HTML[169791]
    assert parse_item_lore(html18) == "Acolyte's Attacker of the Harbinger"
    assert parse_aug_slot_types(html18) == frozenset({18, 19})
    assert classify_aug_type(parse_aug_slot_types(html18)) == 18
    assert parse_aug_slot_types(html19) == frozenset({19})
    assert classify_aug_type(parse_aug_slot_types(html19)) == 19


def test_parse_total_results():
    html = (FIXTURES / "eqresource_type18_search_snip.html").read_text(encoding="utf-8")
    assert parse_total_results(html) == 3


def test_type19_search_payload_uses_augtype_not_augslot():
    # EQ Resource "Fits Aug Slot" is augtype; setting augslot=19 with it yields zero rows.
    p = type19_search_payload(name="Devotee's", page=2)
    assert p["augtype"] == "19"
    assert p["augslot"] == ""
    assert p["type"] == "augs"
    assert p["name"] == "Devotee's"
    assert p["page"] == "2"


def test_enduring_harmony_cloaks_are_ignored():
    assert _is_ignored_non_aug("Mantle of Enduring Harmony")
    assert _is_ignored_non_aug("Cloak of Enduring Harmony")
    assert _is_ignored_non_aug("Cape of Enduring Harmony")
    assert not _is_ignored_non_aug("Devotee's Defense of Enduring Harmony")


def test_extra_stats_parses_spell_damage_from_adjacent_td():
    html = (FIXTURES / "eqresource_aug_type18_casting_stats.html").read_text(
        encoding="utf-8"
    )
    extra = _extra_stats_from_item_html(html)
    assert extra.get("spell_damage") == 11
    assert extra.get("heal_amount") == 11
    assert extra.get("clairvoyance") == 16
    assert extra.get("mana") == 360


def test_stats_rank_key_orders_hp_then_ac():
    a = {"hp": 470, "ac": 40, "hstr": 10}
    b = {"hp": 409, "ac": 50, "hstr": 20}
    assert stats_rank_key(a, "A") < stats_rank_key(b, "B")


def test_fetch_catalog_offline_classifies_and_sorts():
    cat = fetch_type18_catalog(
        allow_network=False,
        type18_html_by_page={
            1: (FIXTURES / "eqresource_type18_search_snip.html").read_text(encoding="utf-8"),
        },
        type19_html_overrides=[
            (FIXTURES / "eqresource_type19_search_snip.html").read_text(encoding="utf-8"),
        ],
        item_html_by_id=_ITEM_HTML,
    )
    by_id = {e.item_id: e for e in cat.entries}
    # Dual-slot Acolyte's (18+19) classify as Type 18 with label 18/19.
    assert by_id[169780].aug_type == 18
    assert by_id[169780].type_label == "18/19"
    assert by_id[169780].category == "Attacker"
    assert by_id[169780].lore_group == "Acolyte's Combatant Augmentation Group"
    assert by_id[169791].aug_type == 19
    assert by_id[169791].type_label == "19"
    assert by_id[109559].anniversary is True
    assert by_id[138886].anniversary is True
    assert by_id[169781].category == "Assaulting"

    attackers19 = [e for e in cat.entries if e.category == "Attacker" and e.aug_type == 19]
    assert [e.item_id for e in attackers19] == [138886, 169791]
    attackers18 = [e for e in cat.entries if e.category == "Attacker" and e.aug_type == 18]
    assert [e.item_id for e in attackers18] == [169780, 109559]
    assert all(e.type_label == "18/19" for e in attackers18)


def test_include_type18_false_omits_section_and_sheet(tmp_path: Path) -> None:
    inv = EXAMPLES / "Deflub_bristle-Inventory.txt"
    bundle = build_export_bundle(
        [inv],
        include_spells=False,
        include_achievements=False,
        include_slot2=False,
        include_type5=False,
        include_type18=False,
        include_raid_bis=False,
    )
    assert bundle.type18 is None
    payload = serialize_report(bundle)
    assert all(s["type"] != "type18_augs" for s in payload["sections"])
    out = tmp_path / "no_type18.xlsx"
    write_team_workbook(bundle.team, out, type18=None)
    assert SHEET_NAME not in load_workbook(out).sheetnames


def test_include_type18_offline_adds_section(tmp_path: Path) -> None:
    inv = EXAMPLES / "Deflub_bristle-Inventory.txt"
    bundle = build_export_bundle(
        [inv],
        include_spells=False,
        include_achievements=False,
        include_slot2=False,
        include_type5=False,
        include_type18=True,
        include_raid_bis=False,
        fetch_chest_class=False,
        **_offline_catalog_kwargs(),
    )
    assert bundle.type18 is not None
    assert len(bundle.type18.entries) >= 5
    payload = serialize_report(bundle)
    section = next(s for s in payload["sections"] if s["type"] == "type18_augs")
    assert section["title"] == "Type 18/19 Augs"
    assert section["data"]["type18Url"] == TYPE18_CATALOG_URL

    data = serialize_type18_section(bundle.type18)
    assert "rows" in data
    assert data["eqResourceItemUrl"]
    assert data["suggestions"]
    assert data["characters"]
    assert data["characters"][0]["classAbbr"]
    assert "ownedIds" in data["characters"][0]
    assert any(r["anniversary"] for r in data["rows"])
    assert any(r.get("typeLabel") == "18/19" for r in data["rows"])
    assert any(r.get("typeLabel") == "19" for r in data["rows"])

    out = tmp_path / "with_type18.xlsx"
    write_team_workbook(bundle.team, out, type18=bundle.type18)
    wb = load_workbook(out)
    assert SHEET_NAME in wb.sheetnames
    assert CATALOG_SHEET_NAME in wb.sheetnames
    suggest = wb[SHEET_NAME]
    assert suggest.cell(1, 1).value == "Character"
    assert suggest.cell(2, 1).value  # character name on first data row


def test_cheat_sheet_and_anniversary_alternative():
    from inventory_parser.type18_augs.suggestions import load_cheat_sheet as _load

    _load.cache_clear()
    sheet = load_cheat_sheet()
    assert "WAR" in sheet
    assert any("Enduring Harmony" in n for n in sheet["WAR"].primary)
    assert sheet["WIZ"].primary == sheet["MAG"].primary
    assert hint_aug_type("Devotee's Attacker of Enduring Harmony") == 19
    assert hint_aug_type("Acolyte's Attacker of the Harbinger") == 18
    assert name_series("Devotee's Attacker of Enduring Harmony") == "of enduring harmony"

    cat = fetch_type18_catalog(
        allow_network=False,
        type18_html_by_page={
            1: (FIXTURES / "eqresource_type18_search_snip.html").read_text(encoding="utf-8"),
        },
        type19_html_overrides=[
            (FIXTURES / "eqresource_type19_search_snip.html").read_text(encoding="utf-8"),
        ],
        item_html_by_id=_ITEM_HTML,
    )
    from inventory_parser.type18_augs.catalog import Type18CatalogEntry

    entries = list(cat.entries)
    entries.append(
        Type18CatalogEntry(
            item_id=999001,
            name="Devotee's Attacker of Legacies Lost",
            aug_type=19,
            type_label="19",
            category="Attacker",
            lore_group="Devotee's Combatant",
            item_lore="Devotee's Attacker of Legacies Lost",
            anniversary=False,
            stats={"hp": 400, "ac": 40},
        )
    )
    # Exact guide Fortifications so extras below stay "unused".
    for iid, name, aug_type, label, hp in (
        (999010, "Acolyte's Fortification of the Harbinger", 18, "18/19", 360),
        (999011, "Acolyte's Fortification of Legacies Lost", 18, "18/19", 340),
        (999012, "Devotee's Fortification of the Harbinger", 19, "19", 409),
        (999013, "Devotee's Fortification of Legacies Lost", 19, "19", 400),
    ):
        entries.append(
            Type18CatalogEntry(
                item_id=iid,
                name=name,
                aug_type=aug_type,
                type_label=label,
                category="Fortification",
                lore_group=None,
                item_lore=name,
                anniversary=False,
                stats={"hp": hp, "ac": 40},
            )
        )
    entries.append(
        Type18CatalogEntry(
            item_id=999002,
            name="Devotee's Fortification of the Selenelion",
            aug_type=19,
            type_label="19",
            category="Fortification",
            lore_group=None,
            item_lore="Devotee's Fortification of the Selenelion",
            anniversary=True,
            stats={"hp": 500, "ac": 50, "mana": 500},
        )
    )
    entries.append(
        Type18CatalogEntry(
            item_id=999003,
            name="Acolyte's Fortification of Whispering Midnight",
            aug_type=18,
            type_label="18/19",
            category="Fortification",
            lore_group=None,
            item_lore=None,
            anniversary=False,
            stats={"hp": 200, "ac": 20},
        )
    )
    blocks = build_class_suggestions(entries, class_abbrs=["WAR", "WIZ"])
    war = next(b for b in blocks if b.class_abbr == "WAR")
    assert war.caster_stats is False
    assert all(
        "Defense" not in (r.suggested.category if r.suggested else r.guide_name)
        and "Ventral" not in r.guide_name
        and "Dorsal" not in r.guide_name
        for r in war.primary
    )
    assert any("Ventral Defense" in r.guide_name for r in war.optional)
    enduring = next(
        r
        for r in war.primary
        if "Enduring Harmony" in r.guide_name and "Attacker" in r.guide_name
    )
    assert enduring.suggested is not None
    assert enduring.suggested.anniversary is True
    assert enduring.alternative is not None
    assert enduring.alternative.anniversary is False
    assert "Enduring Harmony" not in enduring.alternative.name
    assert war.fortification
    assert war.fortification[0].suggested is not None
    assert war.fortification[0].suggested.category == "Fortification"
    # Greatest unused Fortification first (999002 hp 500 before 999003 hp 200).
    assert war.fortification[0].suggested.item_id == 999002

    wiz = next(b for b in blocks if b.class_abbr == "WIZ")
    assert wiz.caster_stats is True

    owned_blocks = build_class_suggestions(
        entries,
        class_abbrs=["WAR"],
        owned_ids_by_class={"WAR": {999010}},
    )
    war_owned = next(b for b in owned_blocks if b.class_abbr == "WAR")
    fort_harb = next(
        r
        for r in war_owned.primary
        if r.suggested and r.suggested.item_id == 999010
    )
    assert fort_harb.owned is True
    assert all(
        not getattr(r, "note", None)
        for r in (*war_owned.primary, *war_owned.optional, *war_owned.fortification)
    )


def test_type18_html_template_has_filters() -> None:
    from inventory_parser.package_data import read_data_text

    html = read_data_text("team_report.html")
    assert "function updateType18AugsContent" in html
    assert "type18-augs" in html
    assert "type18Filters" in html
    assert 'type18_augs: "icon-gem"' in html
    assert "renderType18Suggestions" in html
    assert "type18DefaultCharacter" in html
    assert "type18SuggestionOwned" in html
    assert 'section.type === "type18_augs"' in html
    assert "toolbar-char-wrap" in html
    assert "Alternative" in html
    assert "Fortification (unused)" in html
    assert "Spell Dmg" in html
    assert "badge owned" in html
    assert ">Owned<" in html
    assert "18/19" in html


def test_build_type18_export_categories():
    from inventory_parser.type18_augs.build import Type18Character

    export = build_type18_export(
        allow_network=False,
        type18_html_by_page={
            1: (FIXTURES / "eqresource_type18_search_snip.html").read_text(encoding="utf-8"),
        },
        type19_html_overrides=[
            (FIXTURES / "eqresource_type19_search_snip.html").read_text(encoding="utf-8"),
        ],
        item_html_by_id=_ITEM_HTML,
        class_abbrs=["WAR"],
        characters=[
            Type18Character(
                key="war|char",
                name="WarChar",
                display_name="WarChar",
                class_abbr="WAR",
                owned_ids={169780},
                owned_names={"ornate attacker of the harbinger"},
            )
        ],
    )
    assert "Attacker" in export.categories
    assert "Assaulting" in export.categories
    assert export.suggestions
    assert export.suggestions[0].class_abbr == "WAR"
    assert export.cheat_sheet_url
    assert len(export.characters) == 1
    data = serialize_type18_section(export)
    assert data["characters"][0]["classAbbr"] == "WAR"
    assert 169780 in data["characters"][0]["ownedIds"]
    # Owned is applied in HTML/Excel from the character, not baked into class rows.
    assert all(
        not r.get("owned")
        for block in data["suggestions"]
        for r in (*block["primary"], *block["optional"], *block["fortification"])
    )