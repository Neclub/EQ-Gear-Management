from pathlib import Path

from openpyxl import load_workbook

from inventory_parser.team_report import build_team_report
from inventory_parser.excel_export import write_team_workbook
from inventory_parser.missing_spells import (
    discover_persona_bindings,
    persona_key,
)
from inventory_parser.spell_report import build_spell_rune_report

_MINIMAL_INVENTORY = "Location\tName\tID\tCount\tSlots\nHead\tTest Helm\t1\t1\t0\n"
_PAL_INVENTORY = "Location\tName\tID\tCount\tSlots\nHead\tPAL Helm\t1\t1\t0\n"
_SHD_INVENTORY = "Location\tName\tID\tCount\tSlots\nHead\tSHD Helm\t2\t2\t0\n"


def test_shared_inventory_spell_only_no_gear_columns(tmp_path: Path) -> None:
    inv = tmp_path / "Deflub_bristle-Inventory.txt"
    pal_spell = tmp_path / "Deflub_bristle-PAL-MissingSpells.txt"
    shd_spell = tmp_path / "Deflub_bristle-SHD-MissingSpells.txt"
    inv.write_text(_MINIMAL_INVENTORY, encoding="utf-8")
    pal_spell.write_text("126\tCommittal Rk. III\n", encoding="utf-8")
    shd_spell.write_text("126\tBanshee Skin VIII Rk. III\n", encoding="utf-8")

    report = build_team_report([inv])
    assert len(report.characters) == 0
    assert len(report.spell_characters) == 2
    assert any("Skipping team gear" in w for w in report.warnings)
    names = {c.display_name for c in report.spell_characters}
    assert names == {"Deflub ( PAL )", "Deflub ( SHD )"}


def test_explicit_spell_ignores_other_spells_in_folder(tmp_path: Path) -> None:
    inv = tmp_path / "Deflub_bristle-Inventory.txt"
    pal_spell = tmp_path / "Deflub_bristle-PAL-MissingSpells.txt"
    shd_spell = tmp_path / "Deflub_bristle-SHD-MissingSpells.txt"
    inv.write_text(_MINIMAL_INVENTORY, encoding="utf-8")
    pal_spell.write_text("126\tCommittal Rk. III\n", encoding="utf-8")
    shd_spell.write_text("126\tBanshee Skin VIII Rk. III\n", encoding="utf-8")

    report = build_team_report([inv], spell_paths=[pal_spell])
    assert len(report.characters) == 1
    assert report.characters[0].display_name == "Deflub ( PAL )"
    assert len(report.spell_characters) == 1
    assert not any("Skipping team gear" in w for w in report.warnings)


def test_single_spell_with_inventory_gets_gear_column(tmp_path: Path) -> None:
    inv = tmp_path / "Deflub_bristle-Inventory.txt"
    pal_spell = tmp_path / "Deflub_bristle-PAL-MissingSpells.txt"
    inv.write_text(_MINIMAL_INVENTORY, encoding="utf-8")
    pal_spell.write_text("126\tCommittal Rk. III\n", encoding="utf-8")

    report = build_team_report([inv], spell_paths=[pal_spell])
    assert len(report.characters) == 1
    assert report.characters[0].display_name == "Deflub ( PAL )"
    assert not any("Skipping team gear" in w for w in report.warnings)


def test_subfolder_per_persona_inventory(tmp_path: Path) -> None:
    pal_dir = tmp_path / "PAL"
    shd_dir = tmp_path / "SHD"
    pal_dir.mkdir()
    shd_dir.mkdir()
    pal_inv = pal_dir / "Deflub_bristle-Inventory.txt"
    shd_inv = shd_dir / "Deflub_bristle-Inventory.txt"
    pal_spell = pal_dir / "Deflub_bristle-PAL-MissingSpells.txt"
    shd_spell = shd_dir / "Deflub_bristle-SHD-MissingSpells.txt"
    pal_inv.write_text(_PAL_INVENTORY, encoding="utf-8")
    shd_inv.write_text(_SHD_INVENTORY, encoding="utf-8")
    pal_spell.write_text("126\tCommittal Rk. III\n", encoding="utf-8")
    shd_spell.write_text("126\tBanshee Skin VIII Rk. III\n", encoding="utf-8")

    report = build_team_report([pal_inv, shd_inv], spell_paths=[pal_spell, shd_spell])
    assert len(report.characters) == 2
    assert len(report.spell_characters) == 2
    gear_by_class = {c.class_abbr: c.slots["Head"].name for c in report.characters}
    assert gear_by_class["PAL"] == "PAL Helm"
    assert gear_by_class["SHD"] == "SHD Helm"


def test_inventory_only_no_class_suffix(tmp_path: Path) -> None:
    inv = tmp_path / "Stablub_bristle-Inventory.txt"
    inv.write_text(_MINIMAL_INVENTORY, encoding="utf-8")
    report = build_team_report([inv])
    assert len(report.characters) == 1
    assert len(report.spell_characters) == 0
    assert report.characters[0].display_name == "Stablub"
    assert report.characters[0].class_abbr is None


def test_spell_without_inventory_warns(tmp_path: Path) -> None:
    spell = tmp_path / "Deflub_bristle-PAL-MissingSpells.txt"
    spell.write_text("126\tCommittal Rk. III\n", encoding="utf-8")
    result = discover_persona_bindings([], spell_paths=[spell])
    assert not result.bindings
    assert len(result.warnings) == 1
    assert "No inventory file" in result.warnings[0]


def test_multiple_spells_shared_inventory_bindings(tmp_path: Path) -> None:
    inv = tmp_path / "Deflub_bristle-Inventory.txt"
    inv.write_text(_MINIMAL_INVENTORY, encoding="utf-8")
    (tmp_path / "Deflub_bristle-PAL-MissingSpells.txt").write_text(
        "126\tCommittal Rk. III\n", encoding="utf-8"
    )
    (tmp_path / "Deflub_bristle-SHD-MissingSpells.txt").write_text(
        "126\tHarm Touch Rk. III\n", encoding="utf-8"
    )
    result = discover_persona_bindings([inv])
    assert len(result.bindings) == 2
    assert all(not b.include_gear for b in result.bindings)


def test_spell_report_persona_columns(tmp_path: Path) -> None:
    inv = tmp_path / "Deflub_bristle-Inventory.txt"
    pal_spell = tmp_path / "Deflub_bristle-PAL-MissingSpells.txt"
    shd_spell = tmp_path / "Deflub_bristle-SHD-MissingSpells.txt"
    inv.write_text(_MINIMAL_INVENTORY, encoding="utf-8")
    pal_spell.write_text("126\tCommittal Rk. III\n", encoding="utf-8")
    shd_spell.write_text("126\tBanshee Skin VIII Rk. III\n", encoding="utf-8")

    team = build_team_report([inv], spell_paths=[pal_spell, shd_spell])
    spell_report = build_spell_rune_report(
        team, inventory_paths=[inv], extra_spell_paths=[pal_spell, shd_spell]
    )
    assert spell_report is not None
    pal_pk = persona_key("Deflub", "bristle", "PAL")
    shd_pk = persona_key("Deflub", "bristle", "SHD")
    assert spell_report.counts_by_persona[pal_pk]["Shattering of Ro (2025)"]["Minor"] == 1
    assert spell_report.counts_by_persona[shd_pk]["Shattering of Ro (2025)"]["Minor"] == 1


def test_excel_auto_discovered_shared_inventory_spell_tabs_only(tmp_path: Path) -> None:
    inv = tmp_path / "Deflub_bristle-Inventory.txt"
    pal_spell = tmp_path / "Deflub_bristle-PAL-MissingSpells.txt"
    shd_spell = tmp_path / "Deflub_bristle-SHD-MissingSpells.txt"
    inv.write_text(_MINIMAL_INVENTORY, encoding="utf-8")
    pal_spell.write_text("126\tCommittal Rk. III\n", encoding="utf-8")
    shd_spell.write_text("126\tBanshee Skin VIII Rk. III\n", encoding="utf-8")

    team = build_team_report([inv])
    spell_report = build_spell_rune_report(team, inventory_paths=[inv])
    out = tmp_path / "crew.xlsx"
    write_team_workbook(team, out, spell_report=spell_report)

    wb = load_workbook(out, data_only=True)
    assert wb["Team gear"].cell(1, 3).value is None
    rune_headers = [wb["Missing Runes"].cell(6, c).value for c in range(2, 4)]
    assert rune_headers == ["Deflub ( PAL )", "Deflub ( SHD )"]


def test_excel_explicit_spell_gets_gear_column(tmp_path: Path) -> None:
    inv = tmp_path / "Deflub_bristle-Inventory.txt"
    pal_spell = tmp_path / "Deflub_bristle-PAL-MissingSpells.txt"
    shd_spell = tmp_path / "Deflub_bristle-SHD-MissingSpells.txt"
    inv.write_text(_MINIMAL_INVENTORY, encoding="utf-8")
    pal_spell.write_text("126\tCommittal Rk. III\n", encoding="utf-8")
    shd_spell.write_text("126\tBanshee Skin VIII Rk. III\n", encoding="utf-8")

    team = build_team_report([inv], spell_paths=[pal_spell])
    spell_report = build_spell_rune_report(
        team, inventory_paths=[inv], extra_spell_paths=[pal_spell]
    )
    out = tmp_path / "crew.xlsx"
    write_team_workbook(team, out, spell_report=spell_report)

    wb = load_workbook(out, data_only=True)
    assert wb["Team gear"].cell(1, 3).value == "Deflub ( PAL )"


_WAR_INVENTORY = "Location\tName\tID\tCount\tSlots\nHead\tWAR Helm\t3\t1\t0\n"


def test_class_tagged_inventories_ignore_generic(tmp_path: Path) -> None:
    generic = tmp_path / "Deflub_bristle-Inventory.txt"
    war_inv = tmp_path / "Deflub_bristle-WAR-Inventory.txt"
    pal_inv = tmp_path / "Deflub_bristle-PAL-Inventory.txt"
    shd_inv = tmp_path / "Deflub_bristle-SHD-Inventory.txt"
    war_spell = tmp_path / "Deflub_bristle-WAR-MissingSpells.txt"
    pal_spell = tmp_path / "Deflub_bristle-PAL-MissingSpells.txt"
    shd_spell = tmp_path / "Deflub_bristle-SHD-MissingSpells.txt"
    generic.write_text(_MINIMAL_INVENTORY, encoding="utf-8")
    war_inv.write_text(_WAR_INVENTORY, encoding="utf-8")
    pal_inv.write_text(_PAL_INVENTORY, encoding="utf-8")
    shd_inv.write_text(_SHD_INVENTORY, encoding="utf-8")
    war_spell.write_text("126\tWar Cry Rk. III\n", encoding="utf-8")
    pal_spell.write_text("126\tCommittal Rk. III\n", encoding="utf-8")
    shd_spell.write_text("126\tBanshee Skin VIII Rk. III\n", encoding="utf-8")

    inventories = [generic, war_inv, pal_inv, shd_inv]
    report = build_team_report(inventories)
    assert len(report.characters) == 3
    assert len(report.spell_characters) == 3
    assert not any("Skipping team gear" in w for w in report.warnings)
    names = {c.display_name for c in report.characters}
    assert names == {"Deflub ( WAR )", "Deflub ( PAL )", "Deflub ( SHD )"}
    gear_by_class = {c.class_abbr: c.slots["Head"].name for c in report.characters}
    assert gear_by_class["WAR"] == "WAR Helm"
    assert gear_by_class["PAL"] == "PAL Helm"
    assert gear_by_class["SHD"] == "SHD Helm"
    assert all(
        Path(c.filepath).name != "Deflub_bristle-Inventory.txt" for c in report.characters
    )


def test_html_character_filter_uses_persona_names(tmp_path: Path) -> None:
    from inventory_parser.export_bundle import build_export_bundle
    from inventory_parser.html_export import extract_report_json, write_team_html

    war_inv = tmp_path / "Deflub_bristle-WAR-Inventory.txt"
    pal_inv = tmp_path / "Deflub_bristle-PAL-Inventory.txt"
    shd_inv = tmp_path / "Deflub_bristle-SHD-Inventory.txt"
    war_inv.write_text(_WAR_INVENTORY, encoding="utf-8")
    pal_inv.write_text(_PAL_INVENTORY, encoding="utf-8")
    shd_inv.write_text(_SHD_INVENTORY, encoding="utf-8")

    bundle = build_export_bundle(
        [war_inv, pal_inv, shd_inv],
        include_spells=False,
        include_achievements=False,
    )
    out = tmp_path / "crew.html"
    write_team_html(bundle, out)
    text = out.read_text(encoding="utf-8")
    assert "chip.dataset.key = character.name" in text
    assert "function filterChipLabel" in text
    report = extract_report_json(text)
    chars = report["meta"]["characters"]
    assert len(chars) == 3
    assert {c["shortName"] for c in chars} == {"Deflub"}
    assert {c["name"] for c in chars} == {
        "Deflub ( WAR )",
        "Deflub ( PAL )",
        "Deflub ( SHD )",
    }


def test_class_tagged_inventory_without_spells_gets_gear_column(tmp_path: Path) -> None:
    war_inv = tmp_path / "Deflub_bristle-WAR-Inventory.txt"
    war_inv.write_text(_WAR_INVENTORY, encoding="utf-8")
    report = build_team_report([war_inv])
    assert len(report.characters) == 1
    assert len(report.spell_characters) == 0
    assert report.characters[0].display_name == "Deflub ( WAR )"
    assert report.characters[0].class_abbr == "WAR"
    assert report.characters[0].slots["Head"].name == "WAR Helm"


def test_special_naming_examples_folder() -> None:
    special = Path(__file__).resolve().parents[1] / "Examples" / "SpecialNaming"
    inventories = sorted(special.glob("*-Inventory.txt"))
    report = build_team_report(inventories)
    assert len(report.characters) == 3
    assert len(report.spell_characters) == 3
    assert {c.class_abbr for c in report.characters} == {"WAR", "PAL", "SHD"}
    assert not any("Skipping team gear" in w for w in report.warnings)