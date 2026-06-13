from pathlib import Path

from inventory_parser.team_report import build_team_report
from inventory_parser.evolver import (
    equipped_item_is_evolver,
    evolver_augment_slot_number,
    is_evolver_augment_row,
    parent_location_of_evolver_row,
)
from inventory_parser.excel_export import write_team_workbook
from inventory_parser.parser import (
    InventoryData,
    InventoryItem,
    extract_equipped_items,
    parse_inventory_file,
)

EXAMPLES = Path(__file__).resolve().parents[1] / "Examples"
ACHIEVEMENTS = (
    Path(__file__).resolve().parents[2]
    / "Example"
    / "Achievments"
    / "Inventory"
)


def test_evolver_augment_row_detection() -> None:
    assert is_evolver_augment_row("Ear-Slot6") is True
    assert is_evolver_augment_row("Primary-Slot5") is True
    assert is_evolver_augment_row("Secondary-Slot4") is False
    assert is_evolver_augment_row("Primary-Slot6") is False
    assert is_evolver_augment_row("Primary-Slot3") is False
    assert is_evolver_augment_row("General 5-Slot6-Slot1") is False
    assert evolver_augment_slot_number("Primary") == 5
    assert evolver_augment_slot_number("Ear") == 6
    assert parent_location_of_evolver_row("Primary-Slot5") == "Primary"


def test_resonant_fracture_with_final_augment_not_evolver() -> None:
    items = [
        InventoryItem("Ear", "Defender's Earring of Resonant Fracture", 175000, 1, 6),
        InventoryItem("Ear-Slot1", "Empty", 0, 0, 0),
        InventoryItem("Ear-Slot6", "Some Augment", 123, 1, 6),
    ]
    data = InventoryData("Test", "bristle", "test.txt", items=items)
    equipped, evolver_keys = extract_equipped_items(data)
    assert "Ear-1" not in evolver_keys
    assert equipped_item_is_evolver(equipped["Ear-1"].name) is False


def test_evolver_keys_from_dump() -> None:
    data = parse_inventory_file(EXAMPLES / "Deflub_bristle-Inventory.txt")
    assert data is not None
    equipped, evolver_keys = extract_equipped_items(data)
    assert "Ear-2" in evolver_keys
    assert "Ear-1" not in evolver_keys
    assert "Primary" not in evolver_keys
    assert equipped["Ear-2"].name == "Blooded Righteous Protector's Earring of Rallos"


def test_primary_evolver_from_slot5_row() -> None:
    items = [
        InventoryItem("Primary", "Blooded Righteous Protector's Staff of Rallos", 164500, 1, 6),
        InventoryItem("Primary-Slot1", "Empty", 0, 0, 0),
        InventoryItem("Primary-Slot5", "Some Augment", 123, 1, 6),
    ]
    data = InventoryData("Test", "bristle", "test.txt", items=items)
    equipped, evolver_keys = extract_equipped_items(data)
    assert "Primary" in evolver_keys


def test_tier_primary_with_slot5_row_not_evolver() -> None:
    path = ACHIEVEMENTS / "Shamlub_bristle-Inventory.txt"
    if not path.is_file():
        return
    report = build_team_report([path])
    char = next(c for c in report.characters if c.character == "Shamlub")
    assert char.slots["Primary"].is_evolver is False
    assert char.slots["Primary"].name == "Spectral Luclinite Great Censer"


def test_secondary_not_evolver_from_slot4_row() -> None:
    report = build_team_report([EXAMPLES / "Shamlub_bristle-Inventory.txt"])
    char = report.characters[0]
    assert char.slots["Secondary"].is_evolver is False
    assert char.slots["Primary"].is_evolver is False


def test_equipped_evolver_flag_from_dump() -> None:
    report = build_team_report([EXAMPLES / "Shamlub_bristle-Inventory.txt"])
    char = report.characters[0]
    assert char.slots["Face"].is_evolver is True
    assert char.slots["Chest"].is_evolver is False


def test_evolver_fill_in_excel(tmp_path: Path) -> None:
    report = build_team_report([EXAMPLES / "Shamlub_bristle-Inventory.txt"])
    out = tmp_path / "evolver.xlsx"
    write_team_workbook(report, out)

    from openpyxl import load_workbook

    wb = load_workbook(out, data_only=False)
    ws = wb["Team gear"]
    shamlub_col = next(
        c for c in range(3, ws.max_column + 1) if ws.cell(1, c).value == "Shamlub ( SHM )"
    )
    face_row = next(r for r in range(2, 25) if ws.cell(r, 1).value == "Face")
    sec_row = next(r for r in range(2, 25) if ws.cell(r, 1).value == "Secondary")
    chest_row = next(r for r in range(2, 25) if ws.cell(r, 1).value == "Chest")
    purple = ("005C4688", "FF5C4688")
    assert ws.cell(face_row, shamlub_col).fill.start_color.rgb in purple
    assert ws.cell(sec_row, shamlub_col).fill.start_color.rgb not in purple
    assert ws.cell(chest_row, shamlub_col).fill.start_color.rgb not in purple
