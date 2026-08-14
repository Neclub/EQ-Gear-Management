from pathlib import Path

from inventory_parser.character_column_order import (
    apply_character_column_order,
    build_column_roster,
    normalize_output_format,
    order_roster_entries,
    paths_for_roster_removal,
    save_character_column_order,
    save_output_format,
    saved_character_column_order,
    saved_output_format,
)
from inventory_parser.team_report import build_team_report
from inventory_parser.export_bundle import build_export_bundle
from inventory_parser.missing_spells import split_input_paths

EXAMPLES = Path(__file__).resolve().parents[1] / "Examples"


def _example_paths() -> list[str]:
    inv_dir = EXAMPLES / "Inventory"
    spell_dir = EXAMPLES / "SpellData"
    paths: list[str] = []
    for folder in (inv_dir, spell_dir):
        if folder.is_dir():
            paths.extend(str(path.resolve()) for path in folder.glob("*.txt"))
    return paths


def test_order_roster_entries_uses_saved_order() -> None:
    entries = build_column_roster(_example_paths())
    names = [entry.display_name for entry in entries]
    assert len(names) >= 2
    reversed_keys = [entry.persona_key for entry in reversed(entries)]
    ordered = order_roster_entries(entries, reversed_keys)
    assert [entry.display_name for entry in ordered] == list(reversed(names))


def test_apply_character_column_order_reorders_team_report() -> None:
    paths = [Path(p) for p in _example_paths()]
    inventory_paths, spell_paths, _ = split_input_paths(paths)
    report = build_team_report(inventory_paths, spell_paths=spell_paths or None)
    original = [character.display_name for character in report.characters]
    custom_order = [character.persona_key for character in reversed(report.characters)]
    apply_character_column_order(report, custom_order)
    assert [character.display_name for character in report.characters] == list(reversed(original))


def test_build_export_bundle_honors_character_column_order() -> None:
    paths = [Path(p) for p in _example_paths()]
    inventory_paths, spell_paths, _ = split_input_paths(paths)
    base = build_team_report(inventory_paths, spell_paths=spell_paths or None)
    custom_order = [character.persona_key for character in reversed(base.characters)]
    bundle = build_export_bundle(paths, character_column_order=custom_order, include_slot2=False)
    assert [character.persona_key for character in bundle.team.characters] == custom_order
    if bundle.spell_report is not None:
        spell_keys = bundle.spell_report.persona_keys
        assert spell_keys == [key for key in custom_order if key in spell_keys]


def test_paths_for_roster_removal_drops_only_unused_files() -> None:
    roster = build_column_roster(_example_paths())
    assert len(roster) >= 2
    removing = [roster[0]]
    drop_paths = paths_for_roster_removal(removing, roster, _example_paths())
    assert drop_paths
    remaining_paths = [path for path in _example_paths() if path not in drop_paths]
    remaining_roster = [entry for entry in roster if entry not in removing]
    rebuilt = build_column_roster(remaining_paths)
    assert [entry.persona_key for entry in rebuilt] == [entry.persona_key for entry in remaining_roster]


def test_save_and_load_character_column_order(tmp_path, monkeypatch) -> None:
    from inventory_parser import character_column_order as module

    settings_file = tmp_path / "settings.json"
    monkeypatch.setattr(module, "settings_path", lambda: settings_file)
    save_character_column_order(["Tank_bristle", "Heal_bristle_CLR"])
    assert saved_character_column_order() == ["Tank_bristle", "Heal_bristle_CLR"]


def test_save_and_load_output_format(tmp_path, monkeypatch) -> None:
    from inventory_parser import character_column_order as module

    settings_file = tmp_path / "settings.json"
    monkeypatch.setattr(module, "settings_path", lambda: settings_file)
    assert saved_output_format() == "both"
    assert normalize_output_format("nope") == "both"
    assert save_output_format("html") == "html"
    assert saved_output_format() == "html"
    save_character_column_order(["Tank_bristle"])
    assert saved_output_format() == "html"
    assert saved_character_column_order() == ["Tank_bristle"]


def test_excel_unmade_matches_bundle_character_column_order(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from inventory_parser.excel_export import UNMADE_GEAR_SHEET_NAME, write_team_workbook

    paths = [Path(p) for p in _example_paths()]
    inventory_paths, spell_paths, _ = split_input_paths(paths)
    base = build_team_report(inventory_paths, spell_paths=spell_paths or None)
    custom_order = [character.persona_key for character in reversed(base.characters)]
    bundle = build_export_bundle(
        paths,
        character_column_order=custom_order,
        include_spells=False,
        include_achievements=False,
        include_slot2=False,
    )
    if not bundle.unmade_entries:
        return

    out = tmp_path / "unmade_order.xlsx"
    write_team_workbook(
        bundle.team,
        out,
        unmade_entries=bundle.unmade_entries,
    )
    wb = load_workbook(out, data_only=True)
    ws = wb[UNMADE_GEAR_SHEET_NAME]
    excel_names = [
        ws.cell(row, 1).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 1).value
    ]
    expected = [entry.display_name for entry in bundle.unmade_entries]
    assert excel_names == expected


def test_export_bundle_parses_each_inventory_once(monkeypatch) -> None:
    from collections import Counter

    from inventory_parser import achievement_report as achievement_report_module
    from inventory_parser import parser as parser_module
    from inventory_parser import rune_inventory as rune_inventory_module
    from inventory_parser import team_report as team_report_module
    from inventory_parser import unmade_gear as unmade_gear_module

    paths = [Path(p) for p in _example_paths()]
    inventory_paths, _, _ = split_input_paths(paths)
    assert inventory_paths

    calls: list[str] = []
    real_parse = parser_module.parse_inventory_file

    def counting_parse(filepath):
        calls.append(str(Path(filepath).resolve()))
        return real_parse(filepath)

    for module in (
        parser_module,
        team_report_module,
        unmade_gear_module,
        rune_inventory_module,
        achievement_report_module,
    ):
        monkeypatch.setattr(module, "parse_inventory_file", counting_parse)

    build_export_bundle(
        paths,
        include_spells=False,
        include_achievements=True,
        include_slot2=False,
    )

    counts = Counter(calls)
    assert counts
    for path, count in counts.items():
        assert count == 1, f"{path} parsed {count} times"


def test_settings_path_uses_eqgm_appdata(tmp_path: Path, monkeypatch) -> None:
    from inventory_parser.character_column_order import settings_path

    monkeypatch.setattr("inventory_parser.character_column_order.sys.platform", "win32")
    monkeypatch.setenv("LOCALAPPDATA", str(tmp_path))
    path = settings_path()
    assert path == tmp_path / "EQGM" / "settings.json"
