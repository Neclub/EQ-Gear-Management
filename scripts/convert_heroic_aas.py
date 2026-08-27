"""Convert the Fanra Heroic AA spreadsheet into bundled heroic_aas.json."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "Examples" / "Achievements" / "Heroic AA.xlsx"
OUTPUT_FILE = ROOT / "src" / "inventory_parser" / "data" / "heroic_aas.json"

FANRA_URL = "https://everquest.fanra.info/wiki/Hero%27s_Special_AAs"

INTRO = (
    "In the Alternate Advancement (AA) window, the Special tab contains "
    "these three Hero's AAs."
)

ABILITIES = [
    {
        "id": "fortitude",
        "label": "Hero's Fortitude",
        "description": (
            "Increases your armor class, attack power and the maximum amount "
            "of attack you can gain from items."
        ),
    },
    {
        "id": "resolution",
        "label": "Hero's Resolution",
        "description": (
            "Increases your base statistics and the maximum that your base "
            "statistics can be increased by items and spells. Additionally, "
            "this ability increases your natural hit point, mana and endurance "
            "regeneration and the maximum that your hit point and mana "
            "regeneration can be increased by items."
        ),
    },
    {
        "id": "vitality",
        "label": "Hero's Vitality",
        "description": (
            "Increases your maximum hit points, mana, endurance as well as how "
            "far below zero your hit points can fall before you die."
        ),
    },
]

# Spreadsheet title (after apostrophe normalize) -> in-game dump title.
DUMP_CANONICAL: dict[str, str] = {
    "Savior of West Karana": "Savior of West Karana (Ethernere)",
    "Savior of West Karana II": "Savior of West Karana (Ethernere) II",
    "Savior of West Karana III": "Savior of West Karana (Ethernere) III",
    "Savior of Neriak": "Savior of Neriak - Fourth Gate",
    "Savior of Neriak II": "Savior of Neriak - Fourth Gate II",
    "Savior of Argin Hiz": "Savior of Argin-Hiz",
    "Savior of Katta (Deluge)": "Savior of Katta Castrum: Deluge",
    "Challenger of the Dark Seas": "Challenger of The Darkened Sea",
    "Lesser Spirit Armor (progression)": "Lesser Spirit Armor (Group)",
    "Savior of the Sul Vius, Demiplane of Life": (
        "Savior of the Sul Vius: Demiplane of Life"
    ),
    "Savior of the Sul Vius, Demiplane of Decay": (
        "Savior of the Sul Vius: Demiplane of Decay"
    ),
    "Hero of the Gorowyn": "Hero of Gorowyn",
    "Hero of the Veeshan's Peak": "Hero of Veeshan's Peak",
    "Savior of Empyre: Realm of Ash": "Savior of Empyr: Realms of Ash",
    "Hero of Empyre: Realm of Ash": "Hero of Empyr: Realms of Ash",
    "Savior of Doomfire: The Burning Lands": (
        "Savior of Doomfire, the Burning Lands (TBL)"
    ),
    "Hero of Doomfire: The Burning Lands": "Hero of Doomfire, the Burning Lands",
    "Hero of Shar Vahl: Mean Streets": "Hero of Shar Vahl, Divided: Mean Streets",
    "Savior of Shar Vahl: Mean Streets": (
        "Savior of Shar Vahl, Divided: Mean Streets"
    ),
    "Hero of Shadow Haven: When One Door Closes": (
        "Hero of Ruins of Shadow Haven: When One Door Closes"
    ),
    "Savior of Shadow Haven: When One Door Closes": (
        "Savior of Ruins of Shadow Haven: When One Door Closes"
    ),
    "Hero of Shar Vahl: Under Siege": "Hero of Shar Vahl, Divided: Under Siege",
    "Savior of Shar Vahl: Under Siege": (
        "Savior of Shar Vahl, Divided: Under Siege"
    ),
    "Savior of Thuliasaur": "Savior of Thuliasaur Island",
    "Savior of Degmar": "Savior of Degmar, the Lost Castle",
    "Savior of Caverns": "Savior of Caverns of Endless Song",
    "Savior of The Hero's Forge:Heroes Are Forged": (
        "Savior of The Hero's Forge: Heroes Are Forged"
    ),
    "Lesser Hero of the Darkened Sea": "Lesser Hero of The Darkened Sea",
    "Accomplished Hero of the Darkened Sea": (
        "Accomplished Hero of The Darkened Sea"
    ),
    "Greater Hero of the Darkened Sea": "Greater Hero of The Darkened Sea",
    "Legendary Hero of the Darkened Sea": "Legendary Hero of The Darkened Sea",
    "Novice Hunter of the Empires of Kunark": (
        "Novice Hunter of The Empires of Kunark"
    ),
    "Adept Hunter of the Empires of Kunark": (
        "Adept Hunter of The Empires of Kunark"
    ),
    "Veteran Hunter of the Empires of Kunark": (
        "Veteran Hunter of The Empires of Kunark"
    ),
    "Hero of the Overthere": "Hero of The Overthere",
    "Hero of the Skyfire Mountains": "Hero of The Skyfire Mountains",
    "Savior of the Plane of Smoke": "Savior of The Plane of Smoke",
    "Hero of the Plane of Smoke": "Hero of The Plane of Smoke",
}

_APOSTROPHE_RE = re.compile(r"[\u2018\u2019\u2032`]")
_TOTAL_RE = re.compile(r"^(total\s+|overall total$)", re.IGNORECASE)


def straighten_name(value: str) -> str:
    return _APOSTROPHE_RE.sub("'", value.strip())


def _flag(value: object) -> int:
    if value is None:
        return 0
    text = str(value).strip().casefold()
    return 1 if text in {"f", "r", "v", "1"} else 0


def _existing_eqresource_ids(path: Path) -> dict[str, int]:
    """Preserve EQ Resource ids across spreadsheet refreshes."""
    if not path.is_file():
        return {}
    try:
        from inventory_parser.heroic_aas import normalize_heroic_name
    except ImportError:
        # Allow running the script without an editable install.
        import sys

        sys.path.insert(0, str(ROOT / "src"))
        from inventory_parser.heroic_aas import normalize_heroic_name

    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, int] = {}
    for raw in data.get("achievements") or []:
        if not isinstance(raw, dict):
            continue
        eid = raw.get("eqresource_id")
        try:
            value = int(eid)
        except (TypeError, ValueError):
            continue
        if value <= 0:
            continue
        for label in (raw.get("name"), *(raw.get("aliases") or [])):
            if not isinstance(label, str) or not label.strip():
                continue
            out[normalize_heroic_name(label)] = value
    return out


def _apply_preserved_eqresource_ids(payload: dict, existing: dict[str, int]) -> None:
    if not existing:
        return
    try:
        from inventory_parser.heroic_aas import normalize_heroic_name
    except ImportError:
        import sys

        sys.path.insert(0, str(ROOT / "src"))
        from inventory_parser.heroic_aas import normalize_heroic_name

    for entry in payload.get("achievements") or []:
        keys = [normalize_heroic_name(entry.get("name", ""))]
        for alias in entry.get("aliases") or []:
            keys.append(normalize_heroic_name(alias))
        for key in keys:
            if key and key in existing:
                entry["eqresource_id"] = existing[key]
                break


def convert_workbook(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb.active
        achievements: list[dict] = []
        expansion = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_name, fort, reso, vita = (row + (None,) * 4)[:4]
            if raw_name is None or not str(raw_name).strip():
                continue
            name = straighten_name(str(raw_name))
            if _TOTAL_RE.match(name):
                continue
            f_flag = _flag(fort)
            r_flag = _flag(reso)
            v_flag = _flag(vita)
            if not f_flag and not r_flag and not v_flag:
                expansion = name
                continue
            canonical = DUMP_CANONICAL.get(name, name)
            aliases: list[str] = []
            if canonical != name:
                aliases.append(name)
            original = str(raw_name).strip()
            if original not in {canonical, name} and original not in aliases:
                aliases.append(original)
            entry = {
                "expansion": expansion,
                "name": canonical,
                "fortitude": f_flag,
                "resolution": r_flag,
                "vitality": v_flag,
            }
            if aliases:
                entry["aliases"] = aliases
            achievements.append(entry)
    finally:
        wb.close()

    totals = {
        "fortitude": sum(item["fortitude"] for item in achievements),
        "resolution": sum(item["resolution"] for item in achievements),
        "vitality": sum(item["vitality"] for item in achievements),
    }
    return {
        "source": xlsx_path.name,
        "credit": {
            "text": "Hero's Special AAs — Fanra's EverQuest Wiki",
            "url": FANRA_URL,
        },
        "intro": INTRO,
        "abilities": ABILITIES,
        "totals": totals,
        "achievements": achievements,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output", type=Path, default=OUTPUT_FILE)
    args = parser.parse_args(argv)
    if not args.input.is_file():
        raise SystemExit(f"Input not found: {args.input}")

    payload = convert_workbook(args.input)
    _apply_preserved_eqresource_ids(payload, _existing_eqresource_ids(args.output))
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    totals = payload["totals"]
    print(
        f"Wrote {args.output} ({len(payload['achievements'])} achievements; "
        f"F {totals['fortitude']} / R {totals['resolution']} / V {totals['vitality']})"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
