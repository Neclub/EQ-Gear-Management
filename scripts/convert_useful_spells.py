"""Convert Raccoo's useful-spells xlsx into bundled useful_spells.json."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "Examples" / "SpellData" / "SOR - Raccoo's list of useful spells.xlsx"
OUTPUT_FILE = ROOT / "src" / "inventory_parser" / "data" / "useful_spells.json"

# Raccoo sheet tab → EQ MissingSpells / GUI class abbreviation
CLASS_SHEET_ALIASES = {
    "SHK": "SHD",
}

SKIP_SHEETS = frozenset({"Intro", "TOV", "ROSTBL", "EOK", "TBM", "COTF"})

_NAME_HEADERS = frozenset(
    {
        "spell/disc name",
        "song/disc name",
        "spell name",
    }
)


def _header_index(headers: list[str], *candidates: str) -> int | None:
    wanted = {c.casefold() for c in candidates}
    for i, header in enumerate(headers):
        if header in wanted:
            return i
    return None


def _normalize_highest_rk(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    text = str(value).strip()
    if text.casefold() in {"n/a", "na", "none", ""}:
        return "n/a"
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def _normalize_level(value: object) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_class_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().casefold() if h is not None else "" for h in rows[0]]
    name_i = _header_index(headers, *_NAME_HEADERS)
    if name_i is None:
        raise ValueError(f"Sheet {ws.title!r}: no spell name column found in {headers!r}")
    rk_i = _header_index(headers, "highest rk")
    exp_i = _header_index(headers, "expansion")
    level_i = _header_index(headers, "level")
    comments_i = _header_index(headers, "comments")

    spells: list[dict] = []
    for row in rows[1:]:
        if not row or name_i >= len(row) or row[name_i] is None:
            continue
        name = str(row[name_i]).strip()
        if not name:
            continue
        entry = {
            "name": name,
            "level": _normalize_level(row[level_i] if level_i is not None and level_i < len(row) else None),
            "expansion": (
                str(row[exp_i]).strip()
                if exp_i is not None and exp_i < len(row) and row[exp_i] is not None
                else ""
            ),
            "highest_rk": _normalize_highest_rk(
                row[rk_i] if rk_i is not None and rk_i < len(row) else None
            ),
            "comments": (
                str(row[comments_i]).strip()
                if comments_i is not None and comments_i < len(row) and row[comments_i] is not None
                else ""
            ),
        }
        spells.append(entry)
    return spells


def convert_workbook(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    by_class: dict[str, list[dict]] = {}
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            class_abbr = CLASS_SHEET_ALIASES.get(sheet_name.upper(), sheet_name.upper())
            spells = parse_class_sheet(wb[sheet_name])
            if not spells:
                continue
            by_class[class_abbr] = spells
    finally:
        wb.close()
    return {
        "source": xlsx_path.name,
        "spells_by_class": by_class,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"Raccoo useful-spells xlsx (default: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=OUTPUT_FILE,
        help=f"Output JSON path (default: {OUTPUT_FILE})",
    )
    args = parser.parse_args(argv)
    if not args.input.is_file():
        raise SystemExit(f"Input not found: {args.input}")

    payload = convert_workbook(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    counts = {cls: len(spells) for cls, spells in payload["spells_by_class"].items()}
    total = sum(counts.values())
    print(f"Wrote {args.output} ({total} spells across {len(counts)} classes)")
    for cls in sorted(counts):
        print(f"  {cls}: {counts[cls]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
